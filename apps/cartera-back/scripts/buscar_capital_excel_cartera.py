"""
Calcula el CAPITAL INICIAL (desembolso) de un crédito buscando en el Excel
histórico de cartera "Cartera Préstamos (Cash-In) NUEVA 3.0.xlsx".

Estrategia de búsqueda (en orden):
1. Buscar por número SIFCO exacto.
2. Si no se encuentra y se pasó `--nombre <nombre>`: buscar por nombre del cliente.
   - Si solo aparece UN número SIFCO bajo ese nombre → usarlo.
   - Si aparecen VARIOS números SIFCO (con "_" indicando crédito partido) →
     calcular capital inicial para CADA SIFCO y sumarlos.

Para cada SIFCO encontrado:
- Si primera aparición es cuota 0 ó 1 → el capital de esa fila es el desembolso.
- Si primera aparición es cuota N > 1 → reconstruir hacia atrás hasta cuota 1
  con: capital_{k-1} = (cuota_fija - fijos + capital_k) / (1 + tasa * 1.12)

Estructura de cada hoja (fila 2 = headers):
    A=Fecha | B=# crédito SIFCO | C=# cuota | D=Nombre | E=Capital | F=% tasa
    G=Interés mes | H=Deuda Q | ... | P=Seguro | Q=GPS

Uso:
    python3 buscar_capital_excel_cartera.py <numero_sifco>
    python3 buscar_capital_excel_cartera.py <numero_sifco> --json
    python3 buscar_capital_excel_cartera.py <numero_sifco> --nombre "Cyntia Rodas" --json
"""

import argparse
import json
import os
import sys
import warnings
import unicodedata
from decimal import Decimal, getcontext
from typing import Any

import openpyxl

warnings.filterwarnings("ignore")
getcontext().prec = 28

EXCEL_DIR = "/home/daniel/Documentos"
EXCEL_NAME_HINT = "artera"


def find_excel_path() -> str:
    for f in os.listdir(EXCEL_DIR):
        if EXCEL_NAME_HINT in f and f.endswith(".xlsx"):
            return os.path.join(EXCEL_DIR, f)
    raise FileNotFoundError(f"No se encontró Excel de cartera en {EXCEL_DIR}")


_WB_CACHE: dict[str, Any] = {"wb": None}


def get_workbook():
    """Devuelve un workbook abierto, cacheado por invocación del script."""
    if _WB_CACHE["wb"] is None:
        _WB_CACHE["wb"] = openpyxl.load_workbook(
            find_excel_path(), read_only=True, data_only=True
        )
    return _WB_CACHE["wb"]


def close_workbook():
    if _WB_CACHE["wb"] is not None:
        _WB_CACHE["wb"].close()
        _WB_CACHE["wb"] = None


def normalizar(texto: str) -> str:
    """Normaliza texto para comparación: minúsculas, sin tildes, sin espacios extras."""
    if not texto:
        return ""
    s = unicodedata.normalize("NFKD", str(texto)).encode("ASCII", "ignore").decode()
    return " ".join(s.lower().split())


def matches_nombre(buscar: str, fila: str) -> bool:
    """
    Verifica si TODAS las palabras de `buscar` están en `fila`.
    Permite buscar 'Cyntia Rodas' y matchear 'Cyntia Paola Rodas Marcos'.
    """
    if not buscar or not fila:
        return False
    palabras_buscar = set(normalizar(buscar).split())
    palabras_fila = set(normalizar(fila).split())
    if not palabras_buscar:
        return False
    return palabras_buscar.issubset(palabras_fila)


def buscar_apariciones(
    numero_sifco: str | None = None,
    nombre: str | None = None,
) -> list[dict]:
    """
    Devuelve todas las filas donde aparece el crédito.
    - Si se pasa numero_sifco: match exacto en columna B.
    - Si se pasa nombre: match (subconjunto normalizado) en columna D.
    """
    wb = get_workbook()
    apariciones = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            sifco_celda = str(row[1]).strip() if row[1] is not None else ""
            nombre_celda = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

            match = False
            if numero_sifco and sifco_celda == numero_sifco:
                match = True
            elif nombre and matches_nombre(nombre, nombre_celda):
                match = True

            if not match:
                continue

            apariciones.append({
                "hoja": sheet_name,
                "fecha": row[0],
                "sifco": sifco_celda,
                "cuota": row[2] if len(row) > 2 else None,
                "nombre": nombre_celda,
                "capital": row[4] if len(row) > 4 else None,
                "tasa": row[5] if len(row) > 5 else None,
                "interes_mes": row[6] if len(row) > 6 else None,
                "deuda": row[7] if len(row) > 7 else None,
                "seguro": row[15] if len(row) > 15 else None,
                "gps": row[16] if len(row) > 16 else None,
            })

    return apariciones


def inferir_cuota_fija(a: dict, b: dict) -> dict:
    cap_a = Decimal(str(a["capital"]))
    cap_b = Decimal(str(b["capital"]))
    tasa = Decimal(str(a["tasa"]))
    seguro = Decimal(str(a["seguro"] or 0))
    gps = Decimal(str(a["gps"] or 0))

    abono_capital = cap_a - cap_b
    interes = cap_a * tasa
    iva = interes * Decimal("0.12")
    fijos = seguro + gps
    cuota_fija = abono_capital + interes + iva + fijos

    return {
        "abono_capital": abono_capital,
        "interes": interes,
        "iva": iva,
        "fijos": fijos,
        "cuota_fija": cuota_fija,
        "tasa": tasa,
    }


def rebobinar_hasta_cuota_uno(
    capital_actual: Decimal,
    cuota_actual: int,
    cuota_fija: Decimal,
    fijos: Decimal,
    tasa: Decimal,
) -> Decimal:
    factor = Decimal("1") + tasa * Decimal("1.12")
    capital = capital_actual
    for _ in range(cuota_actual, 1, -1):
        capital = (cuota_fija - fijos + capital) / factor
    return capital


def calcular_capital_para_un_sifco(apariciones: list[dict]) -> dict[str, Any] | None:
    """
    Dado un conjunto de apariciones de UN SIFCO específico (ya filtradas),
    devuelve el capital inicial calculado o None si no se pudo.
    """
    validas = [
        a for a in apariciones
        if isinstance(a["cuota"], int) and a["capital"] is not None and a["tasa"] is not None
    ]
    if not validas:
        return None

    validas.sort(key=lambda x: x["cuota"])
    primera = validas[0]

    if primera["cuota"] in (0, 1):
        return {
            "sifco": primera["sifco"],
            "capital_inicial": float(primera["capital"]),
            "fuente": "directo",
            "primera_cuota_excel": primera["cuota"],
            "primera_hoja": primera["hoja"],
            "nombre": primera["nombre"],
            "tasa": float(primera["tasa"]),
        }

    siguiente = next((a for a in validas if a["cuota"] == primera["cuota"] + 1), None)
    if siguiente is None:
        return None  # no podemos inferir cuota_fija sin dos consecutivas

    inferencia = inferir_cuota_fija(primera, siguiente)
    P = rebobinar_hasta_cuota_uno(
        capital_actual=Decimal(str(primera["capital"])),
        cuota_actual=primera["cuota"],
        cuota_fija=inferencia["cuota_fija"],
        fijos=inferencia["fijos"],
        tasa=inferencia["tasa"],
    )

    return {
        "sifco": primera["sifco"],
        "capital_inicial": float(P),
        "fuente": "calculado",
        "primera_cuota_excel": primera["cuota"],
        "primera_hoja": primera["hoja"],
        "nombre": primera["nombre"],
        "tasa": float(primera["tasa"]),
        "cuota_fija": float(inferencia["cuota_fija"]),
        "fijos": float(inferencia["fijos"]),
    }


def buscar_companions_con_underscore(numero_sifco_base: str) -> list[str]:
    """
    Busca todos los SIFCOs que empiezan con `numero_sifco_base + "_"` (créditos partidos).
    Por ejemplo: para "01010214109340" devuelve ["01010214109340_2", "01010214109340_3", ...]
    """
    wb = get_workbook()
    companions: set[str] = set()
    prefijo = numero_sifco_base + "_"
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2 or row[1] is None:
                continue
            v = str(row[1]).strip()
            if v.startswith(prefijo) and v != numero_sifco_base:
                companions.add(v)
    return sorted(companions)


def buscar_interes_en_apariciones(apariciones: list[dict], numero_cuota: int) -> float | None:
    """
    Busca en las apariciones ya cargadas la fila con cuota = `numero_cuota`
    y devuelve el `interes_mes`. No re-escanea el Excel.
    """
    for a in apariciones:
        if isinstance(a["cuota"], int) and a["cuota"] == numero_cuota:
            if a["interes_mes"] is not None:
                return float(a["interes_mes"])
    return None


def obtener_capital_inicial(
    numero_sifco: str,
    nombre: str | None = None,
    validar_cuota: int | None = None,
    capital_math: float | None = None,
) -> dict[str, Any]:
    # 1. Intentar por número SIFCO
    apariciones = buscar_apariciones(numero_sifco=numero_sifco)
    if apariciones:
        resultado = calcular_capital_para_un_sifco(apariciones)
        if resultado is None:
            raise ValueError(f"Solo una aparición de {numero_sifco}, no se puede inferir cuota fija")

        # 1.b Verificar si hay companions con "_N" (crédito partido) SOLO si el
        # capital del Excel es MENOR que el calculado por matemática (señal de
        # que faltan partes del crédito por sumar). Esto evita el escaneo
        # innecesario del Excel para créditos no partidos.
        capital_excel = resultado["capital_inicial"]
        debe_buscar_companions = (
            capital_math is None or capital_excel < capital_math - 1
        )

        companions = buscar_companions_con_underscore(numero_sifco) if debe_buscar_companions else []
        if companions:
            resultados = [resultado]
            for comp_sifco in companions:
                aps_comp = buscar_apariciones(numero_sifco=comp_sifco)
                r_comp = calcular_capital_para_un_sifco(aps_comp)
                if r_comp is not None:
                    resultados.append(r_comp)
            suma = sum(r["capital_inicial"] for r in resultados)
            res = {
                "numero_sifco": numero_sifco,
                "estrategia": "sifco_partido",
                "capital_inicial": suma,
                "fuente": "suma_creditos_partidos",
                "nombre_cliente": resultado["nombre"],
                "creditos_encontrados": resultados,
                "advertencia": f"Crédito partido: se sumaron {len(resultados)} SIFCOs ({numero_sifco} + {len(companions)} companion(s))",
            }
            if validar_cuota is not None:
                res["validacion"] = {
                    "cuota": validar_cuota,
                    "interes_excel": buscar_interes_en_apariciones(apariciones, validar_cuota),
                    "sifco_consultado": numero_sifco,
                }
            return res

        res = {
            "numero_sifco": numero_sifco,
            "estrategia": "sifco_directo",
            "capital_inicial": resultado["capital_inicial"],
            "fuente": resultado["fuente"],
            "primera_cuota_excel": resultado["primera_cuota_excel"],
            "primera_hoja": resultado["primera_hoja"],
            "nombre_cliente": resultado["nombre"],
            "creditos_encontrados": [resultado],
        }
        if validar_cuota is not None:
            res["validacion"] = {
                "cuota": validar_cuota,
                "interes_excel": buscar_interes_en_apariciones(apariciones, validar_cuota),
                "sifco_consultado": numero_sifco,
            }
        return res

    # 2. Fallback por nombre
    if not nombre:
        raise ValueError(f"{numero_sifco} no encontrado y no se proporcionó nombre")

    apariciones_nombre = buscar_apariciones(nombre=nombre)
    if not apariciones_nombre:
        raise ValueError(f"{numero_sifco} no encontrado, tampoco por nombre '{nombre}'")

    # Agrupar por SIFCO
    sifcos_unicos: dict[str, list[dict]] = {}
    for a in apariciones_nombre:
        sifcos_unicos.setdefault(a["sifco"], []).append(a)

    resultados_por_sifco = []
    for sifco, aps in sifcos_unicos.items():
        r = calcular_capital_para_un_sifco(aps)
        if r is not None:
            resultados_por_sifco.append(r)

    if not resultados_por_sifco:
        raise ValueError(f"Apariciones por nombre encontradas pero ninguna procesable")

    # Si hay un solo SIFCO → usarlo
    if len(resultados_por_sifco) == 1:
        r = resultados_por_sifco[0]
        return {
            "numero_sifco": numero_sifco,
            "estrategia": "nombre_unico",
            "capital_inicial": r["capital_inicial"],
            "fuente": r["fuente"],
            "primera_cuota_excel": r["primera_cuota_excel"],
            "primera_hoja": r["primera_hoja"],
            "nombre_cliente": r["nombre"],
            "creditos_encontrados": resultados_por_sifco,
        }

    # Múltiples SIFCOs → crédito partido. Verificar que tengan "_"
    sifcos_con_underscore = [r for r in resultados_por_sifco if "_" in r["sifco"]]
    if sifcos_con_underscore:
        suma = sum(r["capital_inicial"] for r in resultados_por_sifco)
        return {
            "numero_sifco": numero_sifco,
            "estrategia": "nombre_partido",
            "capital_inicial": suma,
            "fuente": "suma_creditos_partidos",
            "nombre_cliente": resultados_por_sifco[0]["nombre"],
            "creditos_encontrados": resultados_por_sifco,
            "advertencia": f"Crédito partido: se sumaron {len(resultados_por_sifco)} SIFCOs",
        }

    # Múltiples SIFCOs sin "_" → ambigüedad, no podemos decidir
    raise ValueError(
        f"Múltiples SIFCOs ({len(resultados_por_sifco)}) bajo el mismo nombre "
        f"sin marca '_' de partido: {[r['sifco'] for r in resultados_por_sifco]}"
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("numero_sifco")
    parser.add_argument("--nombre", default=None, help="Nombre del cliente (fallback)")
    parser.add_argument("--json", action="store_true", help="Salida en JSON")
    parser.add_argument(
        "--validar-cuota",
        type=int,
        default=None,
        dest="validar_cuota",
        help="Si se pasa, busca también el interes_mes del Excel para esa cuota (para validación post-reparación)",
    )
    parser.add_argument(
        "--capital-math",
        type=float,
        default=None,
        dest="capital_math",
        help="Hint: capital calculado por matemática. Si se pasa y el Excel devuelve un capital >= este, NO busca companions '_N' (ahorra tiempo).",
    )
    args = parser.parse_args()

    try:
        resultado = obtener_capital_inicial(
            args.numero_sifco,
            args.nombre,
            validar_cuota=args.validar_cuota,
            capital_math=args.capital_math,
        )
    except ValueError as e:
        if args.json:
            print(json.dumps({"error": str(e), "numero_sifco": args.numero_sifco}))
        else:
            print(f"❌ {e}")
        sys.exit(2)

    if args.json:
        print(json.dumps(resultado, ensure_ascii=False))
    else:
        print(f"\n🔍 Crédito: {resultado['numero_sifco']}")
        print(f"   Estrategia: {resultado['estrategia']}")
        print(f"   Cliente: {resultado['nombre_cliente']}")
        print(f"\n🎯 Capital inicial: Q{resultado['capital_inicial']:,.2f}")
        print(f"   Fuente: {resultado['fuente']}")
        if resultado.get("advertencia"):
            print(f"   ⚠️  {resultado['advertencia']}")
        print(f"\n   SIFCOs procesados:")
        for c in resultado["creditos_encontrados"]:
            print(f"     - {c['sifco']:25s} → Q{c['capital_inicial']:>12,.2f} ({c['fuente']})")


if __name__ == "__main__":
    try:
        main()
    finally:
        close_workbook()
