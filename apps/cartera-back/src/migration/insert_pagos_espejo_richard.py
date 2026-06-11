import openpyxl
import psycopg2
import re

# Config
EXCEL_PATH = r"src\migration\Richard Kachler.xlsx"
SHEET_NAME = "enero 2026"
DB_URL = ""
INVERSIONISTA_ID = 76
LIQUIDACION_ID = 27

# Mapeo de mes abreviado -> (mes_num, año)
MESES_MAP = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def parse_cuota_mes(cuota_mes_str):
    """Parsea 'ene. 26' -> (1, 2026), 'dic. 25' -> (12, 2025)"""
    cuota_mes_str = str(cuota_mes_str).strip().lower()
    match = re.match(r"(\w+)\.\s*(\d{2})", cuota_mes_str)
    if not match:
        return None, None
    mes_abr = match.group(1)
    anio = 2000 + int(match.group(2))
    mes_num = MESES_MAP.get(mes_abr)
    return mes_num, anio


def remove_accents(s):
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'ä': 'a', 'ë': 'e', 'ï': 'i', 'ö': 'o', 'ü': 'u',
        'ñ': 'n', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O',
        'Ú': 'U', 'Ñ': 'N',
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    return s


def buscar_credito_por_nombre(cur, nombre_cliente):
    """Busca crédito por nombre fuzzy en créditos del inversionista"""
    norm = remove_accents(nombre_cliente.strip().lower())
    palabras = [p for p in norm.split() if len(p) > 2]
    if not palabras:
        return None

    pattern = "%" + "%".join(palabras) + "%"
    cur.execute("""
        SELECT c.credito_id, c.numero_credito_sifco, u.nombre
        FROM cartera.creditos_inversionistas_espejo cie
        JOIN cartera.creditos c ON c.credito_id = cie.credito_id
        JOIN cartera.usuarios u ON u.usuario_id = c.usuario_id
        WHERE cie.inversionista_id = %s
          AND translate(lower(u.nombre),
              'áéíóúàèìòùäëïöüâêîôûñ',
              'aeiouaeiouaeiouaeioun') ILIKE %s
        LIMIT 5
    """, (INVERSIONISTA_ID, pattern))
    rows = cur.fetchall()

    if len(rows) == 1:
        return rows[0]
    elif len(rows) > 1:
        # Mejor match por score
        mejor = rows[0]
        mejor_score = 0
        for r in rows:
            nombre_norm = remove_accents(r[2].lower())
            score = sum(1 for p in palabras if p in nombre_norm)
            if score > mejor_score:
                mejor_score = score
                mejor = r
        return mejor

    # Fallback: buscar con menos palabras (solo las más largas)
    palabras_largas = sorted(palabras, key=len, reverse=True)[:3]
    pattern2 = "%" + "%".join(palabras_largas) + "%"
    cur.execute("""
        SELECT c.credito_id, c.numero_credito_sifco, u.nombre
        FROM cartera.creditos_inversionistas_espejo cie
        JOIN cartera.creditos c ON c.credito_id = cie.credito_id
        JOIN cartera.usuarios u ON u.usuario_id = c.usuario_id
        WHERE cie.inversionista_id = %s
          AND translate(lower(u.nombre),
              'áéíóúàèìòùäëïöüâêîôûñ',
              'aeiouaeiouaeiouaeioun') ILIKE %s
        LIMIT 5
    """, (INVERSIONISTA_ID, pattern2))
    rows2 = cur.fetchall()
    if rows2:
        return rows2[0]

    return None


def buscar_pago_por_cuota_mes(cur, credito_id, mes_num, anio):
    """Busca el pago_id de la cuota del mes/año indicado"""
    # Buscar cuota con fecha_vencimiento en ese mes/año
    cur.execute("""
        SELECT q.cuota_id, q.numero_cuota, q.fecha_vencimiento, p.pago_id
        FROM cartera.cuotas_credito q
        JOIN cartera.pagos_credito p ON p.cuota_id = q.cuota_id
        WHERE q.credito_id = %s
          AND EXTRACT(MONTH FROM q.fecha_vencimiento) = %s
          AND EXTRACT(YEAR FROM q.fecha_vencimiento) = %s
        ORDER BY q.numero_cuota DESC
        LIMIT 1
    """, (credito_id, mes_num, anio))
    row = cur.fetchone()
    if row:
        return row  # cuota_id, numero_cuota, fecha_vencimiento, pago_id

    # Fallback: si no encuentra por fecha exacta, buscar la cuota más cercana
    # no liquidada para inversionistas
    cur.execute("""
        SELECT q.cuota_id, q.numero_cuota, q.fecha_vencimiento, p.pago_id
        FROM cartera.cuotas_credito q
        JOIN cartera.pagos_credito p ON p.cuota_id = q.cuota_id
        WHERE q.credito_id = %s
          AND q.liquidado_inversionistas = false
        ORDER BY q.numero_cuota ASC
        LIMIT 1
    """, (credito_id,))
    return cur.fetchone()


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    nombre_inversionista = str(ws["C2"].value).strip()
    print(f"Inversionista: {nombre_inversionista} (ID: {INVERSIONISTA_ID})")
    print(f"Liquidación: {LIQUIDACION_ID}")
    print("=" * 70)

    total = 0
    exitosos = 0
    fallidos = 0
    errores = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        nombre_cliente = row[1].value        # B
        abono_interes = row[6].value         # G
        abono_iva = row[7].value             # H
        isr = row[8].value                   # I
        amort_capital = row[9].value         # J
        pct_inversor = row[4].value          # E (0.8, 0.7, etc)
        cuota_mes = row[12].value            # M

        if not nombre_cliente or amort_capital is None:
            continue

        nombre_cliente = str(nombre_cliente).strip()
        total += 1

        try:
            # 1. Buscar crédito por nombre
            credito = buscar_credito_por_nombre(cur, nombre_cliente)
            if not credito:
                raise Exception(f"No se encontró crédito para: {nombre_cliente}")
            credito_id, sifco, nombre_db = credito

            # 2. Parsear mes de cuota
            mes_num, anio = parse_cuota_mes(cuota_mes)
            if not mes_num:
                raise Exception(f"No se pudo parsear cuota_mes: {cuota_mes}")

            # 3. Buscar pago de esa cuota
            cuota_info = buscar_pago_por_cuota_mes(cur, credito_id, mes_num, anio)
            if not cuota_info:
                raise Exception(f"No se encontró pago para cuota {cuota_mes} del crédito {sifco}")
            cuota_id, numero_cuota, fecha_venc, pago_id = cuota_info

            # 4. Calcular valores
            cap = float(amort_capital)
            inter = float(abono_interes) if abono_interes else 0
            iva = float(abono_iva) if abono_iva else 0
            isr_val = float(isr) if isr else 0
            pct = float(pct_inversor) * 100 if pct_inversor else 80  # 0.8 -> 80
            # Cuota = capital + interes + iva - isr (facturación ajena)
            cuota_total = cap + inter + iva - isr_val

            # 5. Verificar que no exista ya (unique: pago_id + inversionista_id)
            cur.execute("""
                SELECT id FROM cartera.pagos_credito_inversionistas_espejo
                WHERE pago_id = %s AND inversionista_id = %s
            """, (pago_id, INVERSIONISTA_ID))
            if cur.fetchone():
                # Ya existe, hacer UPDATE
                cur.execute("""
                    UPDATE cartera.pagos_credito_inversionistas_espejo
                    SET abono_capital = %s, abono_interes = %s, abono_iva_12 = %s,
                        cuota = %s, estado_liquidacion = 'LIQUIDADO', liquidacion_id = %s
                    WHERE pago_id = %s AND inversionista_id = %s
                """, (cap, inter, iva, cuota_total, LIQUIDACION_ID, pago_id, INVERSIONISTA_ID))
                exitosos += 1
                print(f"  UPD [{total}] {nombre_cliente} -> {sifco} cuota#{numero_cuota} ({cuota_mes}) cap:{cap:.2f} int:{inter:.2f} iva:{iva:.2f} isr:{isr_val:.2f} = {cuota_total:.2f}")
            else:
                # INSERT
                cur.execute("""
                    INSERT INTO cartera.pagos_credito_inversionistas_espejo
                        (pago_id, inversionista_id, credito_id, abono_capital, abono_interes,
                         abono_iva_12, porcentaje_participacion, fecha_pago, estado_liquidacion,
                         cuota, liquidacion_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'LIQUIDADO', %s, %s)
                """, (pago_id, INVERSIONISTA_ID, credito_id, cap, inter, iva, pct,
                      '2026-02-10', cuota_total, LIQUIDACION_ID))
                exitosos += 1
                print(f"  INS [{total}] {nombre_cliente} -> {sifco} cuota#{numero_cuota} ({cuota_mes}) cap:{cap:.2f} int:{inter:.2f} iva:{iva:.2f} isr:{isr_val:.2f} = {cuota_total:.2f}")

        except Exception as e:
            fallidos += 1
            errores.append({"cliente": nombre_cliente, "error": str(e)})
            print(f"  FAIL [{total}] {nombre_cliente} - {e}")

    conn.commit()
    conn.close()

    print("\n" + "=" * 70)
    print("RESUMEN")
    print("=" * 70)
    print(f"Inversionista:    {nombre_inversionista} (ID: {INVERSIONISTA_ID})")
    print(f"Liquidación:      {LIQUIDACION_ID}")
    print(f"Total procesados: {total}")
    print(f"Exitosos:         {exitosos}")
    print(f"Fallidos:         {fallidos}")

    if errores:
        print(f"\nErrores:")
        for err in errores:
            print(f"  - {err['cliente']}: {err['error']}")


if __name__ == "__main__":
    main()
