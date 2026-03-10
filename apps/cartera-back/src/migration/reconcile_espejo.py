"""
Script para leer los Excel de Preparadas/ y llamar al endpoint /reconcile-espejo
por cada inversionista.

Uso:
    pip install openpyxl requests
    python reconcile_espejo.py

Configura HOJA_OBJETIVO para elegir que hoja leer (ej: "enero 2026", "Enero 2026")
Configura API_URL con la URL del backend.
"""

import os
import sys
import requests
import openpyxl

# --- CONFIGURACION ---
API_URL = "http://localhost:7000/reconcile-espejo"
CARPETA = os.path.join(os.path.dirname(__file__), "Preparadas")
HOJA_OBJETIVO = "enero 2026"  # case-insensitive
LIQUIDACION_ID = 6  # ID de la liquidacion de enero 2026
# ----------------------


def buscar_hoja(wb, nombre_objetivo):
    """Busca la hoja por nombre case-insensitive."""
    objetivo = nombre_objetivo.lower().strip()
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() == objetivo:
            return wb[sheet_name]
    return None


def extraer_creditos(ws):
    """
    Lee la hoja y extrae las filas de creditos.
    Estructura esperada:
      Fila 1 (0-indexed): Info inversionista
      Fila 3 (0-indexed): Headers
      Fila 4+: Datos de creditos

    Columnas relevantes:
      B (col 2) = CLIENTE
      M (col 13) = CUOTA DE MES
    """
    creditos = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        # Saltar filas de header/info (las primeras 4 filas, 0-indexed 0-3)
        if row_idx <= 4:
            continue

        # row es una tupla de valores
        if not row or len(row) < 13:
            continue

        cliente = row[1]   # columna B = CLIENTE
        cuota_mes = row[12] # columna M = CUOTA DE MES
        interes = row[6]   # columna G = INTERES INVERSOR
        iva = row[7]       # columna H = IVA
        isr = row[8]       # columna I = ISR
        amort_cap = row[9] # columna J = AMORTIZACION CAPITAL
        pct_inversor = row[4]  # columna E = % INVERSOR (0.8, 0.7, etc)

        # Validar que tenga datos reales (no fila de totales/vacios)
        if not cliente or not cuota_mes:
            continue

        # Filtrar filas de totales
        cliente_str = str(cliente).strip()
        if not cliente_str or cliente_str.lower() in ("gran total", "total"):
            continue

        cuota_str = str(cuota_mes).strip()
        if not cuota_str:
            continue

        # Calcular cuota total = capital + interes + iva - isr
        cap_val = float(amort_cap) if amort_cap else 0
        int_val = float(interes) if interes else 0
        iva_val = float(iva) if iva else 0
        isr_val = float(isr) if isr else 0
        cuota_total = cap_val + int_val + iva_val - isr_val
        pct_val = float(pct_inversor) * 100 if pct_inversor else 80

        creditos.append({
            "cliente": cliente_str,
            "cuota_mes": cuota_str,
            "abono_capital": str(round(cap_val, 2)),
            "abono_interes": str(round(int_val, 2)),
            "abono_iva_12": str(round(iva_val, 2)),
            "porcentaje_participacion": str(round(pct_val, 2)),
            "cuota": str(round(cuota_total, 2)),
            "liquidacion_id": LIQUIDACION_ID,
        })

    return creditos


def nombre_desde_archivo(filename):
    """Extrae el nombre del inversionista del nombre del archivo."""
    # Quitar todas las extensiones .xlsx
    nombre = filename
    while nombre.lower().endswith(".xlsx"):
        nombre = nombre[:-5]
    return nombre.strip()


def main():
    if not os.path.isdir(CARPETA):
        print(f"ERROR: No se encontro la carpeta {CARPETA}")
        sys.exit(1)

    archivos = sorted([f for f in os.listdir(CARPETA) if f.lower().endswith(".xlsx")])
    print(f"Encontrados {len(archivos)} archivos Excel en {CARPETA}")
    print(f"Hoja objetivo: '{HOJA_OBJETIVO}'")
    print("=" * 80)

    resumen_global = {
        "procesados": 0,
        "omitidos": 0,
        "errores": 0,
        "sin_hoja": 0,
        "sin_creditos": 0,
    }
    detalles_errores = []

    for i, archivo in enumerate(archivos, 1):
        inversionista = nombre_desde_archivo(archivo)
        ruta = os.path.join(CARPETA, archivo)

        print(f"\n[{i}/{len(archivos)}] {inversionista}")
        print("-" * 60)

        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ERROR al abrir Excel: {e}")
            resumen_global["errores"] += 1
            detalles_errores.append({"inversionista": inversionista, "error": str(e)})
            continue

        ws = buscar_hoja(wb, HOJA_OBJETIVO)
        if not ws:
            print(f"  Sin hoja '{HOJA_OBJETIVO}' - saltando")
            resumen_global["sin_hoja"] += 1
            wb.close()
            continue

        creditos = extraer_creditos(ws)
        wb.close()

        if not creditos:
            print(f"  Sin creditos en la hoja - saltando")
            resumen_global["sin_creditos"] += 1
            continue

        print(f"  {len(creditos)} creditos encontrados en Excel")

        # Llamar al endpoint
        payload = {
            "inversionista": inversionista,
            "creditos": creditos,
        }

        try:
            resp = requests.post(API_URL, json=payload, timeout=120)
            data = resp.json()
        except Exception as e:
            print(f"  ERROR en request: {e}")
            resumen_global["errores"] += 1
            detalles_errores.append({"inversionista": inversionista, "error": str(e)})
            continue

        if not data.get("success"):
            print(f"  FALLO: {data.get('message', 'sin mensaje')}")
            resumen_global["errores"] += 1
            detalles_errores.append({"inversionista": inversionista, "error": data.get("message")})
            continue

        # Resumen del inversionista
        resultados = data.get("resultados", [])
        omitidos = data.get("omitidos", [])

        resumen_global["procesados"] += len(resultados)
        resumen_global["omitidos"] += len(omitidos)

        print(f"  OK: {len(resultados)} procesados, {len(omitidos)} omitidos")

        for r in resultados:
            print(f"    + {r['cliente']} | cuota={r['cuota_mes']} | pago_id={r['pago_id_asignado']} | posteriores_no_liq={r['posteriores_a_no_liquidado']}")

        for o in omitidos:
            print(f"    - {o['cliente']}: {o['razon']}")

    # Resumen final
    print("\n" + "=" * 80)
    print("RESUMEN FINAL")
    print("=" * 80)
    print(f"  Creditos procesados:   {resumen_global['procesados']}")
    print(f"  Creditos omitidos:     {resumen_global['omitidos']}")
    print(f"  Inversionistas error:  {resumen_global['errores']}")
    print(f"  Sin hoja objetivo:     {resumen_global['sin_hoja']}")
    print(f"  Sin creditos en hoja:  {resumen_global['sin_creditos']}")

    if detalles_errores:
        print(f"\nERRORES ({len(detalles_errores)}):")
        for e in detalles_errores:
            print(f"  - {e['inversionista']}: {e['error']}")


if __name__ == "__main__":
    main()
