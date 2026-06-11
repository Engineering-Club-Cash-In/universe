import openpyxl
import requests
import time

# Config - Usando Excel de Flujocapital
EXCEL_PATH = r"src\migration\Flujocapital (1).xlsx"
SHEET_NAME = "Enero 2026"
API_URL = "http://localhost:7000/update-pagos-espejo"
HEADER_ROW = 4  # Fila 4 tiene los headers
# Fila 2, columna D = Nombre del inversionista (fijo para todo el Excel)
# Columnas datos (fila 5+): A=No.CREDITO, C=CLIENTE, H=INTERÉS, I=IVA, K=AMORTIZACIÓN CAPITAL

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Leer nombre del inversionista de la fila 2, columna D
    nombre_inversionista = str(ws["D2"].value).strip()
    print(f"Inversionista: {nombre_inversionista}")
    print("=" * 60)

    total = 0
    exitosos = 0
    fallidos = 0
    errores = []

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row, values_only=False):
        numero_credito = row[0].value           # Columna A (No. CRÉDITO)
        nombre_cliente = row[2].value            # Columna C (CLIENTE)
        abono_interes = row[7].value             # Columna H (INTERÉS INVERSOR)
        abono_iva = row[8].value                 # Columna I (IVA)
        amort_capital = row[10].value            # Columna K (AMORTIZACIÓN CAPITAL)

        # Saltar filas vacías
        if not numero_credito or amort_capital is None:
            continue

        numero_credito = str(numero_credito).strip().replace("_", "")
        nombre_cliente = str(nombre_cliente).strip() if nombre_cliente else "N/A"
        total += 1

        payload = {
            "numero_credito_sifco": numero_credito,
            "nombre_inversionista": nombre_inversionista,
            "abono_capital": float(amort_capital),
            "abono_interes": float(abono_interes) if abono_interes is not None else 0,
            "abono_iva": float(abono_iva) if abono_iva is not None else 0,
            "nombre_cliente": nombre_cliente,
        }

        try:
            response = requests.post(API_URL, json=payload, timeout=10)
            data = response.json()

            if response.status_code == 200 and data.get("success"):
                exitosos += 1
                registros = data.get("registrosActualizados", 0)
                int_val = float(abono_interes) if abono_interes else 0
                iva_val = float(abono_iva) if abono_iva else 0
                print(f"  OK  [{total}] {numero_credito} - {nombre_cliente} -> cap: {amort_capital:.2f} + int: {int_val:.2f} + iva: {iva_val:.2f} = {float(amort_capital) + int_val + iva_val:.2f} ({registros} reg)")
            else:
                fallidos += 1
                msg = data.get("message", "Error desconocido")
                errores.append({"credito": numero_credito, "cliente": nombre_cliente, "error": msg})
                print(f"  FAIL [{total}] {numero_credito} - {nombre_cliente} - {msg}")

        except Exception as e:
            fallidos += 1
            errores.append({"credito": numero_credito, "cliente": nombre_cliente, "error": str(e)})
            print(f"  ERROR [{total}] {numero_credito} - {nombre_cliente} - {e}")

    # Resumen
    print("\n" + "=" * 60)
    print("RESUMEN")
    print("=" * 60)
    print(f"Inversionista:    {nombre_inversionista}")
    print(f"Total procesados: {total}")
    print(f"Exitosos:         {exitosos}")
    print(f"Fallidos:         {fallidos}")

    if errores:
        print(f"\nErrores:")
        for err in errores:
            print(f"  - {err['credito']} ({err['cliente']}): {err['error']}")

if __name__ == "__main__":
    main()
