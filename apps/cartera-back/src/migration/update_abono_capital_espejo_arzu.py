import openpyxl
import requests

# Config - Javier Arzú
EXCEL_PATH = r"src\migration\javier Arzu.xlsx"
SHEET_NAME = "Enero 2026"
API_URL = "http://localhost:7000/update-pagos-espejo"
HEADER_ROW = 4
# Fila 2: C2 = Nombre inversionista
# Columnas datos (fila 5+): B=CLIENTE, G=INTERÉS INVERSOR, H=IVA, J=AMORTIZACIÓN CAPITAL

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    nombre_inversionista = str(ws["C2"].value).strip()
    print(f"Inversionista: {nombre_inversionista}")
    print("=" * 60)

    total = 0
    exitosos = 0
    fallidos = 0
    errores = []

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row, values_only=False):
        nombre_cliente = row[1].value            # Columna B (CLIENTE)
        abono_interes = row[6].value             # Columna G (INTERÉS INVERSOR)
        abono_iva = row[7].value                 # Columna H (IVA)
        amort_capital = row[9].value             # Columna J (AMORTIZACIÓN CAPITAL)

        # Saltar filas vacías
        if not nombre_cliente or amort_capital is None:
            continue

        nombre_cliente = str(nombre_cliente).strip()
        total += 1

        payload = {
            "numero_credito_sifco": "",
            "nombre_inversionista": nombre_inversionista,
            "abono_capital": float(amort_capital),
            "abono_interes": float(abono_interes) if abono_interes is not None else 0,
            "abono_iva": float(abono_iva) if abono_iva is not None else 0,
            "nombre_cliente": nombre_cliente,
        }

        try:
            response = requests.post(API_URL, json=payload, timeout=10)
            data = response.json()

            if response.status_code == 200 and data.get("success"):
                exitosos += 1
                registros = data.get("registrosActualizados", 0)
                int_val = float(abono_interes) if abono_interes else 0
                iva_val = float(abono_iva) if abono_iva else 0
                print(f"  OK  [{total}] {nombre_cliente} -> cap: {float(amort_capital):.2f} + int: {int_val:.2f} + iva: {iva_val:.2f} ({registros} reg)")
            else:
                fallidos += 1
                msg = data.get("message", "Error desconocido")
                errores.append({"cliente": nombre_cliente, "error": msg})
                print(f"  FAIL [{total}] {nombre_cliente} - {msg}")

        except Exception as e:
            fallidos += 1
            errores.append({"cliente": nombre_cliente, "error": str(e)})
            print(f"  ERROR [{total}] {nombre_cliente} - {e}")

    # Resumen
    print("\n" + "=" * 60)
    print("RESUMEN")
    print("=" * 60)
    print(f"Inversionista:    {nombre_inversionista}")
    print(f"Total procesados: {total}")
    print(f"Exitosos:         {exitosos}")
    print(f"Fallidos:         {fallidos}")

    if errores:
        print(f"\nErrores:")
        for err in errores:
            print(f"  - {err['cliente']}: {err['error']}")

if __name__ == "__main__":
    main()
