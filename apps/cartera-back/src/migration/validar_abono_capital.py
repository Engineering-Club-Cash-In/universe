import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# === CONFIG ===
CARPETA_EXCELS = r"C:\Users\Kelvin Palacios\Documents\analis de datos"
ARCHIVO_CARTERA = "Cartera Préstamos (Cash-In) NUEVA 3.0.xlsx"
HOJAS_CARTERA = ["Diciembre 2025", "Enero 2026"]

ARCHIVO_DIFERENCIAS = os.path.join(
    os.path.dirname(__file__),
    "diferencias_abono_capital.xlsx"
)
HOJA_DIFERENCIAS = "Diferencias Abono Capital"

ARCHIVO_SALIDA = os.path.join(
    os.path.dirname(__file__),
    "sin_coincidencia_abono_capital.xlsx"
)

def find_file(folder, name_fragment):
    for f in os.listdir(folder):
        if name_fragment in f and f.endswith('.xlsx') and not f.startswith('~'):
            return os.path.join(folder, f)
    return None

def main():
    # 1. Cargar Excel de Cartera y armar diccionario SIFCO -> abono_capital (solo Flujocapital)
    cartera_path = find_file(CARPETA_EXCELS, "NUEVA 3.0.xlsx")
    if not cartera_path or 'cierre' in cartera_path:
        # Buscar el que NO tiene "cierre"
        for f in os.listdir(CARPETA_EXCELS):
            if "NUEVA 3.0.xlsx" in f and not f.startswith("~") and "cierre" not in f:
                cartera_path = os.path.join(CARPETA_EXCELS, f)
                break

    print(f"Abriendo Cartera: {cartera_path}")
    wb_cartera = openpyxl.load_workbook(cartera_path, read_only=True, data_only=True)

    # Col B(1) = SIFCO, Col R(17) = Abono capital, Col AJ(35) = Inversionista
    # Guardar un mapa POR HOJA para poder intentar ambas
    cartera_por_hoja = {}  # hoja -> { sifco -> abono_capital }
    for hoja in HOJAS_CARTERA:
        print(f"  Leyendo hoja: {hoja}")
        ws_cartera = wb_cartera[hoja]
        hoja_map = {}
        for row in ws_cartera.iter_rows(min_row=3, max_col=40, values_only=True):
            sifco_raw = row[1]
            abono_capital = row[17]
            inversionista = row[35]

            if not sifco_raw or not inversionista:
                continue

            if str(inversionista).strip().lower() == "flujocapital":
                sifco = str(sifco_raw).strip()
                if "_" in sifco:
                    continue
                hoja_map[sifco] = float(abono_capital) if abono_capital else 0.0
        cartera_por_hoja[hoja] = hoja_map
        print(f"    Flujocapital encontrados: {len(hoja_map)}")

    wb_cartera.close()

    # 2. Cargar Excel de diferencias
    print(f"Abriendo diferencias: {ARCHIVO_DIFERENCIAS}")
    wb_dif = openpyxl.load_workbook(ARCHIVO_DIFERENCIAS, data_only=True)
    ws_dif = wb_dif[HOJA_DIFERENCIAS]

    sin_coincidencia = []
    coincidencias = 0
    no_encontrados = 0
    no_match = 0

    # Header row = 8, data starts at 9
    for row in ws_dif.iter_rows(min_row=9, max_col=7, values_only=True):
        nombre = row[0]
        abono_capital_tabla = row[1]
        abono_capital_json = row[2]
        diferencia = row[3]
        inversionista = row[5]
        sifco = row[6]

        if not sifco:
            continue

        sifco = str(sifco).strip()

        abono_json = float(abono_capital_json) if abono_capital_json else 0.0

        # Intentar matchear en cada hoja (Dic primero, luego Ene)
        matched = False
        for hoja in HOJAS_CARTERA:
            if sifco in cartera_por_hoja[hoja]:
                abono_cartera = cartera_por_hoja[hoja][sifco]
                if abs(abono_json - abono_cartera) <= 0.02:
                    coincidencias += 1
                    matched = True
                    break

        if matched:
            continue

        # No matcheó en ninguna hoja - reportar con los valores de ambas
        encontrado_en_alguna = any(sifco in cartera_por_hoja[h] for h in HOJAS_CARTERA)

        if not encontrado_en_alguna:
            no_encontrados += 1
            sin_coincidencia.append({
                "sifco": sifco,
                "nombre": nombre,
                "abono_json": round(abono_json, 2),
                "abono_cartera_dic": "NO ENCONTRADO",
                "abono_cartera_ene": "NO ENCONTRADO",
                "motivo": "No encontrado en Cartera",
            })
        else:
            no_match += 1
            val_dic = cartera_por_hoja["Diciembre 2025"].get(sifco)
            val_ene = cartera_por_hoja["Enero 2026"].get(sifco)
            sin_coincidencia.append({
                "sifco": sifco,
                "nombre": nombre,
                "abono_json": round(abono_json, 2),
                "abono_cartera_dic": round(val_dic, 2) if val_dic is not None else "N/A",
                "abono_cartera_ene": round(val_ene, 2) if val_ene is not None else "N/A",
                "motivo": "No coincide en ningún mes",
            })

    wb_dif.close()

    # 3. Resumen
    print(f"\n{'='*60}")
    print("RESUMEN")
    print(f"{'='*60}")
    print(f"Coincidencias OK:     {coincidencias}")
    print(f"No coinciden:         {no_match}")
    print(f"No encontrados:       {no_encontrados}")
    print(f"Total sin coincidencia: {len(sin_coincidencia)}")

    # 4. Generar Excel de salida
    if sin_coincidencia:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Sin Coincidencia"

        # Headers
        headers = ["No. Crédito SIFCO", "Nombre", "Abono Capital JSON", "Abono Capital Dic 2025", "Abono Capital Ene 2026", "Motivo"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws_out.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, item in enumerate(sin_coincidencia, 2):
            ws_out.cell(row=i, column=1, value=item["sifco"])
            ws_out.cell(row=i, column=2, value=item["nombre"])
            ws_out.cell(row=i, column=3, value=item["abono_json"])
            ws_out.cell(row=i, column=4, value=item["abono_cartera_dic"])
            ws_out.cell(row=i, column=5, value=item["abono_cartera_ene"])
            ws_out.cell(row=i, column=6, value=item["motivo"])

        # Auto-width
        for col in ws_out.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_out.column_dimensions[col[0].column_letter].width = max_len + 4

        wb_out.save(ARCHIVO_SALIDA)
        print(f"\nExcel generado: {ARCHIVO_SALIDA}")
    else:
        print("\nTodos coinciden, no se genera Excel.")

if __name__ == "__main__":
    main()
