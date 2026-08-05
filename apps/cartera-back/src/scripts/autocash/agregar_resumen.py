#!/usr/bin/env python3
"""Agrega una hoja 'Resumen' al inicio del comparativo Autocash, sin tocar las hojas existentes.
Los conteos se calculan leyendo el propio archivo, así que siempre reflejan lo que está a la vista.
"""
import os
import shutil
import warnings

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

RUTA = os.path.expanduser("~/Descargas/Comparativo_Autocash_DB_vs_Excel.xlsx")
TOL = 1.00  # Q; por debajo de esto es redondeo del Excel, no descuadre

H1, H2, RESUMEN = "Créditos comparativos", "No aparecen en sistema", "Resumen"

AZUL = PatternFill("solid", fgColor="1F4E79")
AZUL_CLARO = PatternFill("solid", fgColor="DDEBF7")
GRIS = PatternFill("solid", fgColor="F2F2F2")
ROJO = PatternFill("solid", fgColor="FFC7CE")
VERDE = PatternFill("solid", fgColor="C6EFCE")
AMBAR = PatternFill("solid", fgColor="FFEB9C")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def leer():
    wb = openpyxl.load_workbook(RUTA, data_only=True)
    ws = wb[H1]
    filas = [r for r in ws.iter_rows(values_only=True) if r[0]]
    hdr, datos = filas[0], filas[1:]
    idx = {h: i for i, h in enumerate(hdr) if h}

    def v(r, nombre):
        x = r[idx[nombre]]
        return float(x) if isinstance(x, (int, float)) else None

    regs = []
    for r in datos:
        cap_x = v(r, "CAPITAL Excel")
        regs.append({
            "sifco": r[0], "cliente": r[1], "status": r[2],
            "cap_db": v(r, "CAPITAL crédito (DB)"), "cap_inv": v(r, "CAPITAL suma inversionistas (DB)"),
            "cap_x": cap_x, "cuo_db": v(r, "CUOTA crédito (DB)"),
            "cuo_inv": v(r, "CUOTA suma inversionistas (DB)"), "cuo_x": v(r, "CUOTA Excel"),
            "d_cap_inv": v(r, "Dif capital: DB - inversionistas"), "d_cap_x": v(r, "Dif capital: DB - Excel"),
            "d_cuo_inv": v(r, "Dif cuota: DB - inversionistas"), "d_cuo_x": v(r, "Dif cuota: DB - Excel"),
            "liquidado_xls": cap_x is not None and abs(cap_x) < 0.01,
        })

    n_h2 = len([r for r in wb[H2].iter_rows(values_only=True) if r[0]]) - 1
    return wb, regs, n_h2


def cuadra(d, tol=TOL):
    return d is not None and abs(d) <= tol


def main():
    if not os.path.exists(RUTA):
        raise SystemExit(f"No existe {RUTA}")
    bkp = RUTA.replace(".xlsx", "_backup.xlsx")
    shutil.copy2(RUTA, bkp)
    print(f"backup -> {bkp}")

    wb, regs, n_h2 = leer()
    n = len(regs)

    cap_ok = [r for r in regs if cuadra(r["d_cap_x"])]
    cuo_ok = [r for r in regs if cuadra(r["d_cuo_x"])]
    ambas_ok = [r for r in regs if cuadra(r["d_cap_x"]) and cuadra(r["d_cuo_x"])]
    ambas_mal = [r for r in regs if not cuadra(r["d_cap_x"]) and not cuadra(r["d_cuo_x"])]
    alguna_mal = [r for r in regs if not (cuadra(r["d_cap_x"]) and cuadra(r["d_cuo_x"]))]
    exacto = [r for r in regs if cuadra(r["d_cap_x"], 0.0) and cuadra(r["d_cuo_x"], 0.0)]
    liq = [r for r in regs if r["liquidado_xls"]]
    liq_vivo = [r for r in liq if (r["cap_db"] or 0) > TOL]

    cap_inv_ok = [r for r in regs if cuadra(r["d_cap_inv"])]
    cuo_inv_ok = [r for r in regs if cuadra(r["d_cuo_inv"])]
    inv_ambas_ok = [r for r in regs if cuadra(r["d_cap_inv"]) and cuadra(r["d_cuo_inv"])]

    s = lambda k: sum(r[k] or 0 for r in regs)

    if RESUMEN in wb.sheetnames:
        del wb[RESUMEN]
    ws = wb.create_sheet(RESUMEN, 0)

    filas = []          # (texto, valor, tipo)  tipo: titulo|seccion|dato|bien|mal|neutro|nota|blanco
    A = filas.append
    A(("RESUMEN COMPARATIVO AUTOCASH — Sistema (DB cartera, prod) vs Excel de Cartera", None, "titulo"))
    A(("Comparación de capital y cuota. Se considera que CUADRA si la diferencia es "
       f"≤ Q{TOL:.2f} (por debajo de eso es redondeo del Excel).", None, "nota"))
    A(("", None, "blanco"))

    A(("UNIVERSO ANALIZADO", None, "seccion"))
    A(("Créditos que en el sistema tienen a Autocash S.A.", n, "dato"))
    A(("Créditos que YA NO están en el sistema con Autocash pero en el Excel 2026 sí", n_h2, "mal"))
    A(("Total de créditos revisados", n + n_h2, "dato"))
    A(("", None, "blanco"))

    A(("CUADRE SISTEMA vs EXCEL  (sobre los %d créditos Autocash del sistema)" % n, None, "seccion"))
    A(("Capital del crédito CUADRA con el Excel", len(cap_ok), "bien"))
    A(("Capital del crédito NO cuadra con el Excel", n - len(cap_ok), "mal"))
    A(("Cuota del crédito CUADRA con el Excel", len(cuo_ok), "bien"))
    A(("Cuota del crédito NO cuadra con el Excel", n - len(cuo_ok), "mal"))
    A(("CUADRAN LAS DOS (capital y cuota)", len(ambas_ok), "bien"))
    A(("De esas, cuadran exacto al centavo (diferencia = 0.00)", len(exacto), "neutro"))
    A(("NO cuadra al menos una de las dos", len(alguna_mal), "mal"))
    A(("NO cuadran ninguna de las dos", len(ambas_mal), "mal"))
    A(("", None, "blanco"))

    A(("CUADRE INTERNO DEL SISTEMA  (crédito vs suma de sus inversionistas)", None, "seccion"))
    A(("Capital del crédito CUADRA con la suma de inversionistas", len(cap_inv_ok), "bien"))
    A(("Capital del crédito NO cuadra con la suma de inversionistas", n - len(cap_inv_ok), "mal"))
    A(("Cuota del crédito CUADRA con la suma de inversionistas", len(cuo_inv_ok), "bien"))
    A(("Cuota del crédito NO cuadra con la suma de inversionistas", n - len(cuo_inv_ok), "mal"))
    A(("Cuadran las dos", len(inv_ambas_ok), "bien"))
    A(("", None, "blanco"))

    A(("CASO ESPECIAL: CRÉDITOS QUE EL EXCEL YA LIQUIDÓ", None, "seccion"))
    A(("Créditos con Capital restante = 0 en el Excel", len(liq), "neutro"))
    A(("De esos, el sistema TODAVÍA los tiene con capital vivo", len(liq_vivo), "mal"))
    A(("Capital que el sistema tiene vivo y el Excel ya dio por liquidado (Q)",
       round(sum(r["cap_db"] or 0 for r in liq_vivo), 2), "mal"))
    A(("Ojo: en esos créditos el Excel usa la columna Cuota para el monto de finiquito, "
       "no para la cuota mensual; por eso su 'cuota Excel' sale enorme.", None, "nota"))
    A(("", None, "blanco"))

    A(("NOTA DE MÉTODO: CRÉDITOS REPARTIDOS EN VARIOS SIFCO", None, "seccion"))
    A(("El Excel normalmente parte un crédito con el sufijo _2 / _3, pero en algunos casos usa "
       "SIFCO correlativos distintos. Se detectaron agrupando por (cliente, # de cuota) y sumando "
       "solo cuando un único SIFCO del grupo existe en el sistema. Dos créditos Autocash estaban "
       "afectados y ya quedaron corregidos aquí:", None, "nota"))
    A(("01010214105310 (Edmon Robinson) = filas 105310 + 105300 + 105290 del Excel", None, "neutro"))
    A(("01010214109450 (José Ángel Peralta) = filas 109450 + 109450_2 + 109470 del Excel", None, "neutro"))
    A(("Por eso 01010214105290 ya NO aparece como 'crédito que no está en el sistema': "
       "sí está, es la rebanada de Autocash del crédito 01010214105310.", None, "nota"))
    A(("", None, "blanco"))

    A(("MONTOS TOTALES (Q)", None, "seccion"))
    A(("Capital — sistema (crédito)", round(s("cap_db"), 2), "dato"))
    A(("Capital — sistema (suma inversionistas)", round(s("cap_inv"), 2), "dato"))
    A(("Capital — Excel", round(s("cap_x"), 2), "dato"))
    A(("Capital — diferencia sistema vs Excel", round(s("cap_db") - s("cap_x"), 2), "mal"))
    A(("Capital — diferencia crédito vs inversionistas", round(s("cap_db") - s("cap_inv"), 2), "mal"))
    A(("Cuota — sistema (crédito)", round(s("cuo_db"), 2), "dato"))
    A(("Cuota — sistema (suma inversionistas)", round(s("cuo_inv"), 2), "dato"))
    A(("Cuota — diferencia crédito vs inversionistas", round(s("cuo_db") - s("cuo_inv"), 2), "mal"))
    A(("", None, "blanco"))
    A(("Nota: el total de 'Cuota Excel' no se suma porque incluye los montos de finiquito "
       "de los créditos liquidados y no sería comparable.", None, "nota"))

    for i, (txt, val, tipo) in enumerate(filas, 1):
        c1, c2 = ws.cell(row=i, column=1, value=txt), ws.cell(row=i, column=2, value=val)
        if tipo == "titulo":
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            c1.fill = AZUL
            c1.font = Font(bold=True, color="FFFFFF", size=12)
            c1.alignment = Alignment(vertical="center")
            ws.row_dimensions[i].height = 26
        elif tipo == "seccion":
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            c1.fill = AZUL_CLARO
            c1.font = Font(bold=True, size=10, color="1F4E79")
            ws.row_dimensions[i].height = 20
            c1.alignment = Alignment(vertical="center")
        elif tipo == "nota":
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            c1.font = Font(italic=True, size=9, color="595959")
            c1.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 30
        elif tipo != "blanco":
            c1.font = Font(size=10, bold=(tipo in ("bien", "mal")))
            c1.fill = GRIS
            c1.border = BORDE
            c2.border = BORDE
            c2.font = Font(size=11, bold=True)
            c2.alignment = Alignment(horizontal="center")
            c2.number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
            if tipo == "bien" and val:
                c2.fill = VERDE
            elif tipo == "mal" and val:
                c2.fill = ROJO
            elif tipo == "neutro":
                c2.fill = AMBAR

    ws.column_dimensions["A"].width = 78
    ws.column_dimensions["B"].width = 20
    ws.sheet_view.showGridLines = False

    wb.active = 0
    wb.save(RUTA)

    print(f"\nHoja '{RESUMEN}' agregada al inicio de {RUTA}")
    print(f"  créditos Autocash en sistema: {n}   |   ya sin Autocash en sistema: {n_h2}")
    print(f"  capital cuadra {len(cap_ok)}/{n}  ·  cuota cuadra {len(cuo_ok)}/{n}  ·  ambas {len(ambas_ok)}/{n}"
          f"  ·  exacto {len(exacto)}/{n}")
    print(f"  no cuadra al menos una: {len(alguna_mal)}   |   no cuadran las dos: {len(ambas_mal)}")
    print(f"  liquidados en Excel: {len(liq)} (con capital vivo en el sistema: {len(liq_vivo)})")


if __name__ == "__main__":
    main()
