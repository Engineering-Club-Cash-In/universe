#!/usr/bin/env python3
"""¿De dónde salen los Q575.8 que reporta conta?

El comparativo normal toma como foto la ÚLTIMA hoja mensual donde el crédito
aparece cobrado. Conta puede estar usando otra base (la última hoja donde
aparece, o directamente la hoja de un mes fijo). Este script calcula el
descuadre de Autocash bajo TODAS esas bases y dice cuál cae en el número que
ellos reportan.

Solo mira el monto de Autocash (creditos_inversionistas.monto_aportado contra
la columna "Capital restante" de las filas cuyo Inversionista es Autocash).
El capital del crédito no entra.

El CSV de entrada se genera con el COPY del README y vive FUERA del repo
(~/Descargas/datos-cuadre-autocash/), porque trae nombres de clientes y montos.

Uso:
  python3 src/scripts/autocash/buscar_descuadre_autocash.py \
      --csv ~/Descargas/datos-cuadre-autocash/db_prod_inv89.csv --objetivo 575.8
"""
import argparse
import csv
import glob
import importlib.util
import os
import unicodedata
import warnings
from collections import defaultdict

import openpyxl

warnings.filterwarnings("ignore")

AQUI = os.path.dirname(os.path.abspath(__file__))
PAGADO_SI = {"si"}

_spec = importlib.util.spec_from_file_location(
    "base", os.path.join(AQUI, "comparativo_autocash.py"))
B = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(B)

# Los 4 que no están en Cartera.xlsx (viven en el libro propio de Autocash).
OMITIR = {"Centralcarga", "01010407301720", "Josecollia", "Pistolon"}


def resolver_excel(patron):
    if os.path.isfile(patron):
        return patron
    objetivo = unicodedata.normalize("NFC", os.path.basename(patron)).lower()
    d = os.path.dirname(os.path.expanduser(patron)) or "."
    for f in glob.glob(os.path.join(d, "*.xlsx")):
        if unicodedata.normalize("NFC", os.path.basename(f)).lower() == objetivo:
            return f
    raise SystemExit(f"No encontré el Excel: {patron}")


def cabecera(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [B.norm_hdr(v) for v in row]
        if "# credito sifco" in vals:
            hdr = {}
            for j, v in enumerate(vals):
                if v and v not in hdr:
                    hdr[v] = j
            return i, hdr
    return None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--excel", default="~/Descargas/Cartera.xlsx")
    ap.add_argument("--objetivo", type=float, default=575.8)
    # La salida trae nombres de clientes y montos: por defecto se escribe FUERA
    # del repo para que no termine versionada por error.
    ap.add_argument("--salida",
                    default="~/Descargas/datos-cuadre-autocash/diferencias_autocash.csv")
    args = ap.parse_args()

    es_inv = lambda n: "AUTOCASH" in B.norm_nombre(n)   # norm_nombre devuelve MAYÚSCULAS

    ruta_csv = os.path.expanduser(args.csv)
    if not os.path.isabs(ruta_csv):
        ruta_csv = os.path.join(os.getcwd(), ruta_csv)
    db = list(csv.DictReader(open(ruta_csv)))
    en_db = {r["numero_credito_sifco"] for r in db}
    objetivo = [r for r in db if r["es_objetivo"] == "1"
                and r["numero_credito_sifco"] not in OMITIR]
    print(f"Autocash en la DB : {len(objetivo)} créditos (omitidos {len(OMITIR)})")

    ruta = resolver_excel(os.path.expanduser(args.excel))
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = B.hojas_mes(wb)
    print(f"Excel             : {os.path.basename(ruta)}  —  {len(hojas)} hojas "
          f"({hojas[0][1]} → {hojas[-1][1]})\n")

    # ── companions + índice de nombres ────────────────────────────────────────
    comp, nombres = defaultdict(set), {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_nom, c_num = hdr.get("nombre"), hdr.get("#")
        grupos = defaultdict(set)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if not base:
                continue
            nom = B.norm_nombre(row[c_nom]) if (c_nom is not None and c_nom < len(row)) else ""
            if nom:
                nombres[base] = str(row[c_nom]).strip()
            n = row[c_num] if (c_num is not None and c_num < len(row)) else None
            if nom and n is not None:
                grupos[(nom, str(n))].add(base)
        for bases in grupos.values():
            if len(bases) < 2:
                continue
            dentro = [b for b in bases if b in en_db]
            if len(dentro) == 1:
                comp[dentro[0]] |= (bases - {dentro[0]})

    idx = defaultdict(list)
    for base, nom in nombres.items():
        idx[B.norm_nombre(nom)].append(base)

    dueno = {}
    for r in objetivo:
        s = r["numero_credito_sifco"]
        if s in nombres:
            base = s
        else:
            cands = idx.get(B.norm_nombre(r["cliente"]), [])
            base = cands[0] if len(cands) == 1 else None
        if base:
            for k in {base} | comp.get(base, set()):
                dueno[k] = s

    # ── recorrer hojas guardando el monto de Autocash por mes ─────────────────
    # por_mes[hoja][sifco] = capital restante de Autocash en esa hoja
    por_mes, cobrado_en, visto_en = {}, defaultdict(dict), defaultdict(dict)
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_cap, c_cuo = hdr.get("capital restante"), hdr.get("cuota")
        c_pag, c_inv = hdr.get("pagado"), hdr.get("inversionista")
        c_bol = hdr.get("monto boleta")
        agg = defaultdict(lambda: {"mio": None, "cuota": 0.0, "boleta": 0.0, "pagado": []})
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if base not in dueno:
                continue
            a = agg[dueno[base]]
            g = lambda c: (B.num(row[c]) if (c is not None and c < len(row)) else None)
            cap, cuo, bol = g(c_cap), g(c_cuo), g(c_bol)
            a["cuota"] += cuo or 0.0
            a["boleta"] += bol or 0.0
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(B.norm_hdr(row[c_pag]))
            nm = str(row[c_inv]).strip() if (c_inv is not None and c_inv < len(row) and row[c_inv]) else ""
            if nm and es_inv(nm):
                a["mio"] = (a["mio"] or 0.0) + (cap or 0.0)
        por_mes[hoja] = {}
        for s, a in agg.items():
            if a["mio"] is None:
                continue
            por_mes[hoja][s] = a["mio"]
            visto_en[s][hoja] = a["mio"]
            if any(p in PAGADO_SI for p in a["pagado"]) or (
                    a["cuota"] > 0 and abs(a["boleta"] - a["cuota"]) < 0.01):
                cobrado_en[s][hoja] = a["mio"]

    orden = [h for _, h, _a in hojas]
    pos = {h: i for i, h in enumerate(orden)}
    ultimo = lambda d: max(d, key=lambda h: pos[h]) if d else None

    db_monto = {r["numero_credito_sifco"]: float(r["inv_padre"] or 0) for r in objetivo}
    total_db = sum(db_monto.values())

    def descuadre(getter):
        tot, n = 0.0, 0
        for s, v in db_monto.items():
            x = getter(s)
            if x is None:
                continue
            tot += v - x
            n += 1
        return tot, n

    print("=" * 96)
    print(f"DESCUADRE DE AUTOCASH (DB − Excel) SEGÚN LA BASE QUE SE USE — objetivo Q{args.objetivo:,.2f}")
    print("=" * 96)
    print(f"{'BASE':<44}{'CRÉDITOS':>10}{'DESCUADRE':>18}{'|dif vs obj|':>18}")
    print("-" * 96)

    filas = []
    g_cob = lambda s: cobrado_en[s].get(ultimo(cobrado_en[s]))
    g_vis = lambda s: visto_en[s].get(ultimo(visto_en[s]))

    def sin_estructurales(g):
        """Descarta los créditos donde un lado está en 0 y el otro no: son bajas
        que el Excel ya liquidó (o al revés), no descuadres de reparto."""
        def h(s):
            x = g(s)
            if x is None:
                return None
            v = db_monto[s]
            if (abs(x) < 0.01) != (abs(v) < 0.01):
                return None
            return x
        return h

    for et, g in (("última hoja COBRADA (criterio del comparativo)", g_cob),
                  ("última hoja COBRADA, sin bajas de un solo lado", sin_estructurales(g_cob)),
                  ("última hoja donde APARECE", g_vis),
                  ("última hoja donde APARECE, sin bajas de un solo lado", sin_estructurales(g_vis))):
        t, n = descuadre(g)
        filas.append((et, n, t))
    for h in orden:
        t, n = descuadre(lambda s, h=h: por_mes[h].get(s))
        filas.append((f"hoja fija: {h}", n, t))

    for et, n, t in filas:
        d = min(abs(t - args.objetivo), abs(t + args.objetivo))
        marca = "  ← 🎯" if d < 1.0 else ""
        print(f"{et:<44}{n:>10}{t:>18,.2f}{d:>18,.2f}{marca}")

    print("-" * 96)
    print(f"{'TOTAL Autocash en la DB':<44}{len(db_monto):>10}{total_db:>18,.2f}")

    # Volcado del detalle bajo la base "sin bajas de un solo lado": es la que sirve
    # para discutir con conta, porque las bajas no son descuadres de reparto.
    g_real = sin_estructurales(g_cob)
    nombre_de = {r["numero_credito_sifco"]: r["cliente"] for r in objetivo}
    detalle = []
    for s, v in db_monto.items():
        x = g_real(s)
        if x is None or abs(v - x) < 0.005:
            continue
        detalle.append((round(v - x, 2), s, nombre_de[s], v, x))
    detalle.sort(key=lambda t: -abs(t[0]))
    import csv as _csv
    salida = os.path.expanduser(args.salida)
    os.makedirs(os.path.dirname(salida) or ".", exist_ok=True)
    with open(salida, "w", newline="") as fh:
        w = _csv.writer(fh)
        w.writerow(["dif", "sifco", "cliente", "db", "excel"])
        w.writerows(detalle)
    print(f"\n(detalle de {len(detalle)} diferencias → {salida})")

    mejor = min(filas, key=lambda f: min(abs(f[2] - args.objetivo), abs(f[2] + args.objetivo)))
    print(f"\nLa base más cercana a Q{args.objetivo:,.2f} es «{mejor[0]}» con {mejor[2]:,.2f}")

    # Desglose de la mejor base
    if mejor[0].startswith("hoja fija: "):
        h = mejor[0][len("hoja fija: "):]
        g = lambda s, h=h: por_mes[h].get(s)
    else:
        g = g_cob if mejor[0].startswith("última hoja COBRADA") else g_vis
    det = []
    for s, v in db_monto.items():
        x = g(s)
        if x is None or abs(v - x) < 0.005:
            continue
        det.append((v - x, s, x, v))
    det.sort(key=lambda t: -abs(t[0]))
    nombre_de = {r["numero_credito_sifco"]: r["cliente"] for r in objetivo}
    print(f"\n{'CLIENTE':<36}{'SIFCO':<17}{'DB':>15}{'EXCEL':>15}{'DIF':>13}")
    print("-" * 96)
    for d, s, x, v in det[:25]:
        print(f"{nombre_de[s][:35]:<36}{s[:16]:<17}{v:>15,.2f}{x:>15,.2f}{d:>13,.2f}")


if __name__ == "__main__":
    main()
