#!/usr/bin/env python3
"""
Comparativo Autocash: DB de cartera (prod, solo lectura) vs Excel "Cartera Préstamos (Cash-In) NUEVA 3.0".

Hoja 1 "Créditos comparativos": los 82 créditos que en la DB tienen a Autocash S.A. como
inversionista, con las 6 métricas:
    capital  -> creditos.capital | SUM(creditos_inversionistas.monto_aportado) | Excel col "Capital restante"
    cuota    -> creditos.cuota   | SUM(creditos_inversionistas.cuota_inversionista) | Excel col "Cuota"

Hoja 2 "No aparecen en sistema": créditos que en alguna hoja de 2026 figuran con inversionista
Autocash en el Excel pero que en la DB ya no tienen a Autocash.

Del Excel, un crédito son 1..N filas (SIFCO base + sufijo _2, _3 por inversionista); se suman.
La foto se toma de la ÚLTIMA hoja cronológica donde alguna fila del crédito tenga Pagado ∈ {Si, Atrasado}.
"""
import csv
import glob
import os
import re
import sys
import unicodedata
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

AQUI = os.path.dirname(os.path.abspath(__file__))
DB_CSV = os.path.join(AQUI, "db_creditos.csv")
SALIDA = os.path.expanduser("~/Descargas/Comparativo_Autocash_DB_vs_Excel.xlsx")
TOLERANCIA = 1.00  # Q; diferencias por debajo se consideran redondeo

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
PAGADO_OK = {"si", "atrasado"}


def sin_tildes(s):
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c))


def norm_hdr(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", sin_tildes(v)).strip().lower()


def norm_nombre(v):
    """Normaliza nombre de cliente para el match de respaldo."""
    if v is None:
        return ""
    s = sin_tildes(v).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def base_sifco(v):
    """Quita el sufijo _N (una fila por inversionista) y deja el número base."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "#n/a", "#ref!", "0"):
        return None
    return re.sub(r"_\d+$", "", s)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("Q", "").replace(" ", "")
    if not s or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------- 1. DB
def cargar_db():
    autocash, todos = {}, {}
    with open(DB_CSV, newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            reg = {
                "sifco": (r["numero_credito_sifco"] or "").strip(),
                "cliente": r["cliente"] or "",
                "status": r["status"] or "",
                "capital": num(r["capital"]) or 0.0,
                "cuota": num(r["cuota"]) or 0.0,
                "n_inv": int(r["n_inv"]) if r["n_inv"] else 0,
                "cap_inv": num(r["suma_aportado"]) or 0.0,
                "cuota_inv": num(r["suma_cuota_inv"]) or 0.0,
                "inversionistas": r["inversionistas"] or "",
                "es_autocash": r["es_autocash"] == "1",
            }
            todos[reg["sifco"]] = reg
            if reg["es_autocash"]:
                autocash[reg["sifco"]] = reg
    return autocash, todos


# ------------------------------------------------------------- 2. Excel
def hojas_mes(wb):
    """Devuelve [(orden, nombre_hoja, anio)] de las hojas mensuales, en orden cronológico.
    Se excluyen las hojas con sufijo '_' (duplicados en formato viejo) y las no-mensuales."""
    out = []
    for name in wb.sheetnames:
        if name.strip().endswith("_"):
            continue
        m = re.fullmatch(r"\s*([A-Za-zÁÉÍÓÚáéíóú]+)\s+(\d{4})\s*", name)
        if not m:
            continue
        mes = MESES.get(sin_tildes(m.group(1)).lower())
        if not mes:
            continue
        anio = int(m.group(2))
        out.append((anio * 100 + mes, name, anio))
    out.sort()
    return out


def leer_excel(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = hojas_mes(wb)
    print(f"  hojas mensuales: {len(hojas)}  ({hojas[0][1]} → {hojas[-1][1]})")

    ultima_pagada = {}    # base -> foto de la última hoja con Pagado ∈ {Si, Atrasado}
    ultima_vista = {}     # base -> foto de la última hoja donde aparece (respaldo)
    autocash_2026 = defaultdict(list)   # base -> [hojas de 2026 con inversionista Autocash]
    nombres = {}          # base -> último nombre visto
    avisos = []

    for _, hoja, anio in hojas:
        ws = wb[hoja]
        hdr, hdr_row = {}, None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
            vals = [norm_hdr(v) for v in row]
            if "# credito sifco" in vals:
                hdr_row = i
                for j, v in enumerate(vals):
                    if v and v not in hdr:
                        hdr[v] = j
                break
        if hdr_row is None:
            avisos.append(f"{hoja}: sin encabezado, se omite")
            continue

        c_sifco = hdr["# credito sifco"]
        c_nom = hdr.get("nombre")
        c_cap = hdr.get("capital restante")
        c_cuo = hdr.get("cuota")
        c_pag = hdr.get("pagado")
        c_inv = hdr.get("inversionista")
        c_pdm = hdr.get("pago del mes")

        agg = defaultdict(lambda: {"cap": None, "cuota": None, "inv": [], "filas": 0,
                                   "pagado": [], "pdm": 0.0, "nombre": None, "sifcos": []})
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if c_sifco >= len(row):
                continue
            base = base_sifco(row[c_sifco])
            if not base:
                continue
            a = agg[base]
            a["filas"] += 1
            a["sifcos"].append(str(row[c_sifco]).strip())
            if c_nom is not None and c_nom < len(row) and row[c_nom] and not a["nombre"]:
                a["nombre"] = str(row[c_nom]).strip()
            if c_cap is not None and c_cap < len(row):
                v = num(row[c_cap])
                if v is not None:
                    a["cap"] = (a["cap"] or 0.0) + v
            if c_cuo is not None and c_cuo < len(row):
                v = num(row[c_cuo])
                if v is not None:
                    a["cuota"] = (a["cuota"] or 0.0) + v
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(norm_hdr(row[c_pag]))
            if c_inv is not None and c_inv < len(row) and row[c_inv]:
                nm = str(row[c_inv]).strip()
                if nm not in a["inv"]:
                    a["inv"].append(nm)
            if c_pdm is not None and c_pdm < len(row):
                a["pdm"] += num(row[c_pdm]) or 0.0

        for base, a in agg.items():
            if a["nombre"]:
                nombres[base] = a["nombre"]
            foto = {"hoja": hoja, "cap": a["cap"], "cuota": a["cuota"], "filas": a["filas"],
                    "inv": list(a["inv"]), "sifcos": a["sifcos"],
                    "tiene_col_cuota": c_cuo is not None, "tiene_col_pagado": c_pag is not None,
                    "pagado": list(a["pagado"]), "pdm": a["pdm"]}
            ultima_vista[base] = foto
            if any(p in PAGADO_OK for p in a["pagado"]):
                ultima_pagada[base] = foto
            if anio == 2026 and any("autocash" in i.lower() for i in a["inv"]):
                autocash_2026[base].append(hoja)

    return ultima_pagada, ultima_vista, autocash_2026, nombres, avisos


# ------------------------------------------------------------- 3. match
def construir_indice_nombres(nombres):
    idx = defaultdict(list)
    for base, nom in nombres.items():
        n = norm_nombre(nom)
        if n:
            idx[n].append(base)
    return idx


def emparejar(sifco_db, cliente_db, ultima_vista, idx_nombres):
    """-> (base_excel | None, descripción de la llave usada)"""
    s = (sifco_db or "").strip()
    if s in ultima_vista:
        return s, "SIFCO"
    sin_ceros = s.lstrip("0")
    if sin_ceros and sin_ceros != s:
        for cand in (sin_ceros, sin_ceros.zfill(14)):
            if cand in ultima_vista:
                return cand, "SIFCO (ceros a la izquierda)"
    cands = idx_nombres.get(norm_nombre(cliente_db), [])
    if len(cands) == 1:
        return cands[0], "Nombre"
    if len(cands) > 1:
        return None, f"Nombre ambiguo ({len(cands)} candidatos)"
    return None, "Sin match"


FUTURAS = {"Septiembre 2026", "Octubre 2026"}  # posteriores al mes en curso (agosto 2026)


def foto_de(base, ultima_pagada, ultima_vista):
    if base in ultima_pagada:
        f = ultima_pagada[base]
        crit = "Última hoja pagada"
        if f["hoja"] in FUTURAS:
            crit += " (¡hoja futura!)"
        return f, crit
    if base in ultima_vista:
        f = ultima_vista[base]
        crit = ("Última aparición (hoja sin col. Pagado)" if not f["tiene_col_pagado"]
                else "Última aparición (nunca marcado Pagado)")
        return f, crit
    return None, "No está en el Excel"


# ------------------------------------------------------------ 4. salida
AZUL = PatternFill("solid", fgColor="1F4E79")
GRIS = PatternFill("solid", fgColor="D9D9D9")
ROJO = PatternFill("solid", fgColor="FFC7CE")
AMBAR = PatternFill("solid", fgColor="FFEB9C")
VERDE = PatternFill("solid", fgColor="C6EFCE")


def escribir_hoja(ws, encabezados, filas, cols_dinero, cols_dif):
    ws.append(encabezados)
    for c in range(1, len(encabezados) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = AZUL
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 34
    for f in filas:
        ws.append(f)
    for r in range(2, ws.max_row + 1):
        for c in cols_dinero:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        for c in cols_dif:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.fill = VERDE if abs(cell.value) <= TOLERANCIA else (
                    AMBAR if abs(cell.value) <= 100 else ROJO)
            elif cell.value:
                cell.fill = GRIS
    anchos = {}
    for r in ws.iter_rows(values_only=True):
        for j, v in enumerate(r, 1):
            anchos[j] = min(46, max(anchos.get(j, 10), len(str(v)) + 2 if v is not None else 10))
    for j, w in anchos.items():
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}{ws.max_row}"


def main():
    cand = [f for f in glob.glob(os.path.expanduser("~/Descargas/*.xlsx")) if "NUEVA 3.0" in f]
    if not cand:
        sys.exit("No encontré el Excel 'NUEVA 3.0' en ~/Descargas")
    ruta = cand[0]
    print(f"Excel : {ruta}")
    print(f"DB CSV: {DB_CSV}")

    autocash_db, todos_db = cargar_db()
    print(f"  DB: {len(todos_db)} créditos, {len(autocash_db)} con Autocash S.A.")

    ultima_pagada, ultima_vista, autocash_2026, nombres, avisos = leer_excel(ruta)
    print(f"  Excel: {len(ultima_vista)} créditos distintos, "
          f"{len(ultima_pagada)} con hoja pagada, {len(autocash_2026)} con Autocash en 2026")
    for a in avisos:
        print(f"    aviso: {a}")

    idx_nombres = construir_indice_nombres(nombres)

    # ---------------- Hoja 1
    h1, emparejados = [], set()
    for sifco, d in sorted(autocash_db.items()):
        base, llave = emparejar(sifco, d["cliente"], ultima_vista, idx_nombres)
        if base:
            emparejados.add(base)
        foto, criterio = foto_de(base, ultima_pagada, ultima_vista) if base else (None, "Sin match en Excel")

        cap_x = foto["cap"] if foto else None
        cuo_x = foto["cuota"] if foto else None
        if foto and not foto["tiene_col_cuota"]:
            cuo_x = None

        d_cap_inv = round(d["capital"] - d["cap_inv"], 2)
        d_cap_x = round(d["capital"] - cap_x, 2) if cap_x is not None else "sin dato Excel"
        d_cuo_inv = round(d["cuota"] - d["cuota_inv"], 2)
        d_cuo_x = round(d["cuota"] - cuo_x, 2) if cuo_x is not None else "sin dato Excel"

        # El Excel marca la liquidación poniendo Capital restante = 0 y el monto de
        # finiquito en la columna Cuota; ahí "cuota Excel" NO es la cuota mensual.
        liquidado_xls = cap_x is not None and abs(cap_x) < 0.01

        alertas = []
        if liquidado_xls and d["capital"] > TOLERANCIA:
            alertas.append("Excel lo da por LIQUIDADO (cap. restante 0) y la DB lo tiene vivo; "
                           "la 'cuota Excel' es el monto de finiquito, no la mensual")
        elif liquidado_xls:
            alertas.append("Excel lo da por liquidado; la 'cuota Excel' es el finiquito, no la mensual")
        if abs(d_cap_inv) > TOLERANCIA:
            alertas.append("capital DB ≠ suma inversionistas")
        if isinstance(d_cap_x, float) and abs(d_cap_x) > TOLERANCIA and not liquidado_xls:
            alertas.append("capital DB ≠ Excel")
        if abs(d_cuo_inv) > TOLERANCIA:
            alertas.append("cuota DB ≠ suma inversionistas")
        if isinstance(d_cuo_x, float) and abs(d_cuo_x) > TOLERANCIA and not liquidado_xls:
            alertas.append("cuota DB ≠ Excel")
        if foto and not any("autocash" in i.lower() for i in foto["inv"]):
            alertas.append("en el Excel ya no figura Autocash")
        if base and base not in autocash_2026:
            alertas.append("sin ninguna hoja de 2026 con Autocash en el Excel")
        if llave.startswith("Nombre"):
            alertas.append(f"match por {llave.lower()}")
        if criterio.endswith("(¡hoja futura!)"):
            alertas.append("la hoja usada es un mes futuro")

        ult = ultima_vista.get(base) if base else None
        h1.append([
            sifco, d["cliente"], d["status"], d["n_inv"], d["inversionistas"],
            llave, foto["hoja"] if foto else "", criterio,
            foto["filas"] if foto else "", " | ".join(foto["inv"]) if foto else "",
            round(d["capital"], 2), round(d["cap_inv"], 2), round(cap_x, 2) if cap_x is not None else "sin dato",
            d_cap_inv, d_cap_x,
            round(d["cuota"], 2), round(d["cuota_inv"], 2), round(cuo_x, 2) if cuo_x is not None else "sin dato",
            d_cuo_inv, d_cuo_x,
            "; ".join(alertas) or "OK",
            ult["hoja"] if ult else "",
            round(ult["cap"], 2) if (ult and ult["cap"] is not None) else "",
        ])

    # ---------------- Hoja 2
    h2 = []
    for base, hojas26 in sorted(autocash_2026.items()):
        if base in emparejados:
            continue
        foto, criterio = foto_de(base, ultima_pagada, ultima_vista)
        cap_x = foto["cap"] if foto else None
        cuo_x = foto["cuota"] if (foto and foto["tiene_col_cuota"]) else None

        en_db = todos_db.get(base)
        if en_db:
            nota_db = "Sí, mismo SIFCO (pero ya sin Autocash)"
        else:
            n = norm_nombre(nombres.get(base, ""))
            hits = [v for v in todos_db.values() if norm_nombre(v["cliente"]) == n] if n else []
            if not hits:
                nota_db = "NO existe en la DB (ni por SIFCO ni por nombre)"
            elif len(hits) == 1:
                en_db, nota_db = hits[0], f"Ese SIFCO no existe; por nombre pega con {hits[0]['sifco']}"
            else:
                nota_db = ("Ese SIFCO no existe; el nombre tiene "
                           f"{len(hits)} créditos en la DB: "
                           + ", ".join(f"{h['sifco']} [{h['status']}]" for h in hits))

        liquidado_xls = cap_x is not None and abs(cap_x) < 0.01
        nota_xls = ("Excel lo da por liquidado (cap. restante 0); la 'cuota Excel' es el finiquito"
                    if liquidado_xls else "")

        h2.append([
            base, nombres.get(base, ""), foto["hoja"] if foto else "", criterio,
            len(hojas26), " | ".join(hojas26[-3:]),
            " | ".join(foto["inv"]) if foto else "",
            round(cap_x, 2) if cap_x is not None else "sin dato",
            round(cuo_x, 2) if cuo_x is not None else "sin dato",
            nota_xls,
            nota_db,
            en_db["status"] if en_db else "",
            en_db["inversionistas"] if en_db else "",
            round(en_db["capital"], 2) if en_db else "",
            round(en_db["cuota"], 2) if en_db else "",
        ])

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Créditos comparativos"
    escribir_hoja(ws1, [
        "SIFCO (DB)", "Cliente (DB)", "Status (DB)", "# inv (DB)", "Inversionistas (DB)",
        "Llave de match", "Hoja Excel usada", "Criterio de la hoja", "Filas en Excel",
        "Inversionistas (Excel)",
        "CAPITAL crédito (DB)", "CAPITAL suma inversionistas (DB)", "CAPITAL Excel",
        "Dif capital: DB - inversionistas", "Dif capital: DB - Excel",
        "CUOTA crédito (DB)", "CUOTA suma inversionistas (DB)", "CUOTA Excel",
        "Dif cuota: DB - inversionistas", "Dif cuota: DB - Excel",
        "Alertas",
        "Ref: última hoja donde aparece", "Ref: capital en esa hoja",
    ], h1, cols_dinero=[11, 12, 13, 16, 17, 18, 23], cols_dif=[14, 15, 19, 20])

    ws2 = wb.create_sheet("No aparecen en sistema")
    escribir_hoja(ws2, [
        "SIFCO (Excel)", "Nombre (Excel)", "Hoja Excel usada", "Criterio de la hoja",
        "# hojas 2026 c/ Autocash", "Últimas hojas 2026 c/ Autocash", "Inversionistas (Excel)",
        "CAPITAL Excel", "CUOTA Excel", "Nota del Excel",
        "¿Está en la DB?", "Status (DB)", "Inversionistas (DB)", "Capital (DB)", "Cuota (DB)",
    ], h2, cols_dinero=[8, 9, 14, 15], cols_dif=[])

    wb.save(SALIDA)
    print(f"\nListo: {SALIDA}")
    print(f"  Hoja 1 'Créditos comparativos': {len(h1)} créditos")
    print(f"  Hoja 2 'No aparecen en sistema': {len(h2)} créditos")


if __name__ == "__main__":
    main()
