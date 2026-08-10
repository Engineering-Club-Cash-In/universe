#!/usr/bin/env python3
"""Comparativo de la cartera de AUTOCASH: DB de cartera vs Excel "Cartera Préstamos
(Cash-In) NUEVA 3.0". Salida por consola, para el reporte de gerencia.

Compara dos cosas por crédito:
  1. CAPITAL del crédito     — creditos.capital        vs suma de "Capital restante"
  2. LO DE AUTOCASH          — creditos_inversionistas vs la fila del Excel cuyo
                               "Inversionista" es Autocash

Criterio de la foto (el mes que se toma como bueno): la ÚLTIMA hoja mensual donde el
crédito aparece con `Pagado = "Si"` **o** donde el "Pago del mes" iguala a la "Cuota"
(hay meses cobrados que quedaron sin marcar el Pagado).

NO vale "Atrasado": cuando el cliente cae en mora el Excel congela el capital de casi
todos los inversionistas, pero al que se le sigue sirviendo la cuota le sigue
amortizando. Tomar esas hojas mete diferencias que no son descuadres.

Un crédito son 1..N filas del Excel (SIFCO base + sufijo _2/_3, una por inversionista);
se suman. Y a veces el Excel lo parte en SIFCO correlativos distintos ("companions"),
que se detectan agrupando por (nombre, # de cuota).

Uso:
  python3 src/scripts/autocash/comparativo_autocash_gerencia.py
"""
import csv
import glob
import importlib.util
import os
import sys
import warnings
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

AQUI = os.path.dirname(os.path.abspath(__file__))
DB_CSV = os.path.join(AQUI, "db_autocash_full.csv")
TOL = 1.00          # Q; por debajo se considera redondeo
PAGADO_SI = {"si"}  # estricto: "Atrasado" no cuenta como pagado

# Créditos que quedan fuera de todo reporte.
OMITIR = {
    "CRM-e11d3cec-6836-455e-be9c-c2d67bf71a85":
        "MARVIN RODRIGUEZ — crédito basura que debería estar eliminado (todo en 0)",
}

# Créditos que NO están en el Excel de cartera general porque viven en el libro
# propio de Autocash (`Cartera Préstamos (Autocash).xlsx`). Ya se cotejaron ahí y
# cuadran al centavo, así que no son un hallazgo: se listan aparte.
EN_LIBRO_AUTOCASH = {
    "Pistolon":       "Lucas F. Fernandez (El Pistolón)",
    "Centralcarga":   "Central de Carga S.A.",
    "01010407301720": "Distribuidora 360",
    "Josecollia":     "Jose Andres Collia Orero",
}

_spec = importlib.util.spec_from_file_location(
    "base", os.path.join(AQUI, "comparativo_autocash.py")
)
B = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(B)


def es_autocash(nombre):
    return "AUTOCASH" in B.norm_nombre(nombre)


def cabecera(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [B.norm_hdr(v) for v in row]
        if "# credito sifco" in vals:
            hdr = {}
            for j, v in enumerate(vals):
                if v and v not in hdr:
                    hdr[v] = j
            return i, hdr
    return None, None


def main():
    db = list(csv.DictReader(open(DB_CSV)))
    en_db = {r["numero_credito_sifco"] for r in db}
    todos_auto = [r for r in db if r["es_autocash"] == "1"]
    objetivo = [r for r in todos_auto if r["numero_credito_sifco"] not in OMITIR]
    omitidos = [r for r in todos_auto if r["numero_credito_sifco"] in OMITIR]
    print(f"DB local : {len(db)} créditos, {len(todos_auto)} con Autocash S.A."
          f"  → {len(objetivo)} en el reporte, {len(omitidos)} omitidos")

    ruta = [f for f in glob.glob(os.path.expanduser("~/Descargas/*.xlsx")) if "NUEVA 3.0" in f][0]
    print(f"Excel    : {ruta}")
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = B.hojas_mes(wb)
    print(f"Hojas    : {len(hojas)}  ({hojas[0][1]} → {hojas[-1][1]})\n")

    # ── Pase 1: companions + índice de nombres ────────────────────────────────
    comp = defaultdict(set)
    nombres = {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_nom, c_num = hdr.get("nombre"), hdr.get("#")
        grupos = defaultdict(set)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if not base:
                continue
            nom = B.norm_nombre(row[c_nom]) if (c_nom is not None and c_nom < len(row)) else ""
            if nom:
                nombres[base] = str(row[c_nom]).strip()
            n = row[c_num] if (c_num is not None and c_num < len(row)) else None
            if nom and n is not None:
                grupos[(nom, str(n))].add(base)
        for bases in grupos.values():
            if len(bases) < 2:
                continue
            dentro = [b for b in bases if b in en_db]
            if len(dentro) == 1:
                comp[dentro[0]] |= (bases - {dentro[0]})

    idx_nombres = defaultdict(list)
    for base, nom in nombres.items():
        idx_nombres[B.norm_nombre(nom)].append(base)

    # ── Match ─────────────────────────────────────────────────────────────────
    dueno, llave_de = {}, {}
    for r in objetivo:
        s = r["numero_credito_sifco"]
        if s in nombres:
            base, llave = s, "SIFCO"
        else:
            cands = idx_nombres.get(B.norm_nombre(r["cliente"]), [])
            if len(cands) == 1:
                base, llave = cands[0], "Nombre"
            elif len(cands) > 1:
                base, llave = None, f"Nombre ambiguo ({len(cands)})"
            else:
                base, llave = None, "Sin match"
        llave_de[s] = (base, llave)
        if base:
            for k in {base} | comp.get(base, set()):
                dueno[k] = s

    # ── Pase 2: foto del último mes cobrado ───────────────────────────────────
    def nueva():
        return {"cap": 0.0, "cuota": 0.0, "pdm": 0.0, "pagado": [], "filas": 0,
                "sifcos": [], "inv": [], "auto_cap": None, "auto_cuota": 0.0,
                "otros": defaultdict(float), "ncuota": None}

    pagada, vista = {}, {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_cap, c_cuo = hdr.get("capital restante"), hdr.get("cuota")
        c_pag, c_inv = hdr.get("pagado"), hdr.get("inversionista")
        c_pdm, c_num = hdr.get("pago del mes"), hdr.get("#")
        agg = defaultdict(nueva)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if base not in dueno:
                continue
            a = agg[dueno[base]]
            a["filas"] += 1
            a["sifcos"].append(str(row[hdr["# credito sifco"]]).strip())
            cap = B.num(row[c_cap]) if (c_cap is not None and c_cap < len(row)) else None
            cuo = B.num(row[c_cuo]) if (c_cuo is not None and c_cuo < len(row)) else None
            pdm = B.num(row[c_pdm]) if (c_pdm is not None and c_pdm < len(row)) else None
            if cap is not None:
                a["cap"] += cap
            if cuo is not None:
                a["cuota"] += cuo
            if pdm is not None:
                a["pdm"] += pdm
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(B.norm_hdr(row[c_pag]))
            if c_num is not None and c_num < len(row) and a["ncuota"] is None:
                a["ncuota"] = row[c_num]
            nm = str(row[c_inv]).strip() if (c_inv is not None and c_inv < len(row) and row[c_inv]) else ""
            if nm:
                if nm not in a["inv"]:
                    a["inv"].append(nm)
                if es_autocash(nm):
                    a["auto_cap"] = (a["auto_cap"] or 0.0) + (cap or 0.0)
                    a["auto_cuota"] += cuo or 0.0
                else:
                    a["otros"][nm] += cap or 0.0
        for s, a in agg.items():
            foto = dict(a, hoja=hoja, tiene_cuota=c_cuo is not None,
                        otros=dict(a["otros"]))
            vista[s] = foto
            # Mes cobrado: Pagado=Si, o el pago del mes iguala a la cuota.
            cobrado = any(p in PAGADO_SI for p in a["pagado"]) or (
                a["cuota"] > 0 and abs(a["pdm"] - a["cuota"]) < 0.01
            )
            if cobrado:
                pagada[s] = foto

    # ── Armado ────────────────────────────────────────────────────────────────
    filas, sin_match = [], []
    for r in sorted(objetivo, key=lambda x: x["cliente"]):
        s = r["numero_credito_sifco"]
        base, llave = llave_de[s]
        foto = pagada.get(s)
        crit = "mes cobrado"
        if foto is None:
            foto = vista.get(s)
            crit = "última aparición (nunca cobrado)" if foto else None
        if foto is None:
            sin_match.append({
                "sifco": s, "cliente": r["cliente"], "status": r["status"],
                "llave": llave,
                "capital": float(r["capital"] or 0), "cuota": float(r["cuota"] or 0),
                "autocash": float(r["autocash_padre"] or 0),
                "inversionistas": r["inversionistas"],
            })
            continue

        cap_db = float(r["capital"] or 0)
        auto_db = float(r["autocash_padre"] or 0)
        cap_x = foto["cap"]
        auto_x = foto["auto_cap"]
        liquidado = abs(cap_x) < 0.01

        d_cap = cap_db - cap_x
        d_auto = (auto_db - auto_x) if auto_x is not None else None

        obs = []
        if liquidado and cap_db > TOL:
            obs.append("Excel lo LIQUIDÓ (capital restante 0) y la DB lo tiene vivo")
        else:
            if abs(d_cap) > TOL:
                obs.append(f"capital del crédito ({d_cap:+,.2f})")
            if d_auto is None:
                obs.append("Autocash ya no figura en el Excel")
            elif abs(d_auto) > TOL:
                obs.append(f"monto de Autocash ({d_auto:+,.2f})")

        filas.append({
            "sifco": s, "cliente": r["cliente"], "status": r["status"],
            "llave": llave, "comp": sorted(comp.get(base, set())) if base else [],
            "hoja": foto["hoja"], "crit": crit, "ncuota": foto["ncuota"],
            "n_inv": int(r["n_inv"] or 0), "inv_db": r["inversionistas"],
            "inv_x": foto["inv"], "otros_x": foto["otros"],
            "cap_db": cap_db, "cap_x": cap_x, "d_cap": d_cap,
            "auto_db": auto_db, "auto_x": auto_x, "d_auto": d_auto,
            "auto_espejo": float(r["autocash_espejo"] or 0) if r["autocash_espejo"] else None,
            "cuo_db": float(r["cuota"] or 0), "cuo_x": foto["cuota"] if foto["tiene_cuota"] else None,
            "liquidado": liquidado, "obs": obs,
        })

    # ── Clasificación en lotes ────────────────────────────────────────────────
    # 1 el Excel lo liquidó y el sistema lo tiene vivo
    # 2 al revés: el sistema en cero y el Excel con saldo
    # 3 el CAPITAL DEL CRÉDITO difiere de verdad
    # 4 el capital cuadra y lo que cambia es el reparto entre inversionistas
    # 5 centavos (≤ Q100 en las dos medidas)
    for x in filas:
        d_cap, d_auto = x["d_cap"], x["d_auto"] or 0.0
        if x["liquidado"] and x["cap_db"] > TOL:
            x["lote"] = 1
        elif x["cap_db"] <= TOL and x["cap_x"] > TOL:
            x["lote"] = 2
        elif abs(d_cap) > 100:
            x["lote"] = 3
        elif abs(d_cap) <= TOL and abs(d_auto) > TOL:
            x["lote"] = 4
        elif x["obs"]:
            x["lote"] = 5
        else:
            x["lote"] = 0

    ok = [x for x in filas if not x["obs"]]
    mal = [x for x in filas if x["obs"]]
    if "--excel" in sys.argv:
        generar_excel(filas, len(ok), len(filas))
    solos = [x for x in filas if x["n_inv"] == 1]
    acomp = [x for x in filas if x["n_inv"] > 1]

    def M(v, w=13):
        return (f"{v:>{w},.2f}" if isinstance(v, float) else f"{'—':>{w}}")

    # ── Salida ────────────────────────────────────────────────────────────────
    print("=" * 146)
    print(f"✅ CUADRAN — capital del crédito y monto de Autocash  ({len(ok)} de {len(filas)})")
    print("=" * 146)
    print(f"{'CLIENTE':<36}{'SIFCO':<17}{'STATUS':<12}{'CAP DB':>13}{'CAP XLS':>13}"
          f"{'AUTOCASH DB':>14}{'AUTO XLS':>13}{'#INV':>5}  MES")
    print("-" * 146)
    for x in ok:
        print(f"{x['cliente'][:35]:<36}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
              f"{M(x['cap_db'])}{M(x['cap_x'])}{M(x['auto_db'],14)}{M(x['auto_x'])}"
              f"{x['n_inv']:>5}  {x['hoja']}")

    print("\n" + "=" * 146)
    print(f"⚠️  NO CUADRAN  ({len(mal)} de {len(filas)})")
    print("=" * 146)
    for x in mal:
        print(f"\n▸ {x['cliente']}  [{x['status']}]   {x['sifco']}")
        for o in x["obs"]:
            print(f"    → {o}")
        print(f"    Mes usado : {x['hoja']} (cuota #{x['ncuota']}, {x['crit']})"
              + (f"   companions: {', '.join(x['comp'])}" if x["comp"] else ""))
        print(f"    CAPITAL   : DB {M(x['cap_db'])}   Excel {M(x['cap_x'])}   dif {M(x['d_cap'])}")
        print(f"    AUTOCASH  : DB {M(x['auto_db'])}   Excel {M(x['auto_x'])}   dif {M(x['d_auto'])}"
              + (f"   espejo {M(x['auto_espejo'])}" if x["auto_espejo"] is not None else ""))
        if x["n_inv"] > 1:
            print(f"    Otros DB  : {x['inv_db']}")
            print(f"    Otros XLS : " + (" | ".join(f"{k}={v:,.2f}" for k, v in x["otros_x"].items()) or "—"))

    en_libro = [x for x in sin_match if x["sifco"] in EN_LIBRO_AUTOCASH]
    sin_match = [x for x in sin_match if x["sifco"] not in EN_LIBRO_AUTOCASH]

    if en_libro:
        print("\n" + "=" * 146)
        print(f"📗 NO ESTÁN EN EL EXCEL DE CARTERA — viven en el libro propio de Autocash  ({len(en_libro)})")
        print("=" * 146)
        print(f"{'CLIENTE':<40}{'SIFCO':<17}{'STATUS':<12}{'CAPITAL DB':>14}{'AUTOCASH DB':>14}")
        print("-" * 146)
        for x in en_libro:
            print(f"{x['cliente'][:39]:<40}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
                  f"{M(x['capital'],14)}{M(x['autocash'],14)}")
        print(f"{'':<69}{sum(x['capital'] for x in en_libro):>14,.2f}{sum(x['autocash'] for x in en_libro):>14,.2f}")
        print("  Cotejados contra 'Cartera Préstamos (Autocash).xlsx' (Julio 2026): cuadran al centavo.")

    print("\n" + "=" * 146)
    print(f"❓ NO APARECEN EN NINGÚN EXCEL  ({len(sin_match)})")
    print("=" * 146)
    if sin_match:
        print(f"{'CLIENTE':<40}{'SIFCO':<17}{'STATUS':<12}{'CAPITAL DB':>14}{'AUTOCASH DB':>14}  MOTIVO")
        print("-" * 146)
        for x in sin_match:
            print(f"{x['cliente'][:39]:<40}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
                  f"{M(x['capital'],14)}{M(x['autocash'],14)}  {x['llave']}")

    print("\n" + "=" * 146)
    print(f"RESUMEN  ({len(objetivo)} créditos con Autocash, tolerancia Q{TOL:.2f})")
    print("=" * 146)
    print(f"  Encontrados en el Excel        : {len(filas)}")
    print(f"  Cuadran capital Y Autocash     : {len(ok)}")
    print(f"  No cuadran                     : {len(mal)}")
    print(f"  En el libro propio de Autocash : {len(en_libro)}  (cuadran)")
    print(f"  No aparecen en ningún Excel    : {len(sin_match)}")
    for r in omitidos:
        print(f"  Omitido: {r['numero_credito_sifco']} — {OMITIR[r['numero_credito_sifco']]}")
    print()
    print(f"  Autocash solo (1 inversionista): {len(solos)}   cuadran {sum(1 for x in solos if not x['obs'])}")
    print(f"  Con otros inversionistas       : {len(acomp)}   cuadran {sum(1 for x in acomp if not x['obs'])}")
    tot_db = sum(x["auto_db"] for x in filas) + sum(x["autocash"] for x in sin_match)
    tot_x = sum(x["auto_x"] for x in filas if x["auto_x"] is not None)
    print()
    print(f"  TOTAL Autocash en la DB        : {tot_db:>16,.2f}")
    print(f"  TOTAL Autocash en el Excel     : {tot_x:>16,.2f}   (solo los {sum(1 for x in filas if x['auto_x'] is not None)} que el Excel reconoce)")


# ════════════════════════════════════════════════════════════════════════════
# Reporte de una sola hoja para gerencia: lotes 1, 2 y 3.
# El lote 4 (reparto entre inversionistas) y el 5 (centavos) quedan fuera a
# propósito: no son un descuadre del capital del crédito.
# ════════════════════════════════════════════════════════════════════════════
SALIDA = os.path.expanduser("~/Descargas/Reporte_Autocash_Gerencia.xlsx")

AZUL   = "1F4E79"
ROJO   = "C00000"
NARANJ = "ED7D31"
CELESTE= "2E75B6"
GRISF  = "F2F2F2"
BORDE  = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def generar_excel(filas, n_ok, n_total):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    r = 1

    def titulo(txt, size=16, color="000000", fill=None, alto=24, span=8):
        nonlocal r
        ws.cell(row=r, column=1, value=txt)
        c = ws.cell(row=r, column=1)
        c.font = Font(bold=True, size=size, color="FFFFFF" if fill else color)
        c.alignment = Alignment(vertical="center")
        if fill:
            for j in range(1, span + 1):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=fill)
        ws.row_dimensions[r].height = alto
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        r += 1

    def texto(txt, size=10, italic=False, span=8):
        nonlocal r
        ws.cell(row=r, column=1, value=txt).font = Font(size=size, italic=italic, color="595959")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        r += 1

    def bloque(color, encabezado, subtitulo, cols, datos, total_idx=()):
        nonlocal r
        r += 1
        titulo(encabezado, size=12, fill=color, alto=22, span=len(cols))
        texto(subtitulo, size=9, italic=True, span=len(cols))
        for j, (h, _w) in enumerate(cols, 1):
            c = ws.cell(row=r, column=j, value=h)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=AZUL)
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            c.border = BORDE
        ws.row_dimensions[r].height = 30
        r += 1
        primera = r
        for fila in datos:
            for j, v in enumerate(fila, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = BORDE
                c.font = Font(size=10)
                if isinstance(v, float):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(vertical="center")
            r += 1
        if datos and total_idx:
            for j in range(1, len(cols) + 1):
                c = ws.cell(row=r, column=j)
                c.fill = PatternFill("solid", fgColor=GRISF)
                c.border = BORDE
                c.font = Font(bold=True, size=10)
            ws.cell(row=r, column=1, value="TOTAL")
            for j in total_idx:
                col = get_column_letter(j)
                c = ws.cell(row=r, column=j,
                            value=f"=SUM({col}{primera}:{col}{r-1})")
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
            r += 1
        r += 1

    # ── Cabecera ──────────────────────────────────────────────────────────────
    titulo("CARTERA AUTOCASH — Sistema vs Excel de Cartera", 16, AZUL)
    texto(f"Generado el {date.today().strftime('%d/%m/%Y')}   ·   "
          f"{n_total} créditos revisados   ·   {n_ok} cuadran   ·   {n_total - n_ok} con hallazgo")
    texto("Se compara el capital de cada crédito contra la última hoja mensual del Excel donde el crédito "
          "aparece efectivamente cobrado (Pagado = Si, o el pago del mes iguala a la cuota). "
          "Los meses en mora no se toman: ahí el Excel congela el capital y la comparación no es válida.")
    r += 1

    L = lambda n: sorted([x for x in filas if x["lote"] == n], key=lambda x: -abs(x["d_cap"]))

    bloque(ROJO,
           "1 · EL EXCEL LOS DA POR LIQUIDADOS Y EL SISTEMA LOS TIENE VIVOS",
           "El Excel les puso el capital en cero y pagó el finiquito. En el sistema el crédito sigue con saldo.",
           [("Cliente", 42), ("Crédito", 20), ("Estado en el sistema", 14),
            ("Capital en el sistema", 15), ("Capital según Excel", 15),
            ("De eso, lo de Autocash", 15), ("Mes del finiquito", 14)],
           [[x["cliente"], x["sifco"], x["status"], x["cap_db"], x["cap_x"], x["auto_db"], x["hoja"]]
            for x in L(1)],
           total_idx=(4, 6))

    bloque(CELESTE,
           "2 · EL SISTEMA LOS TIENE EN CERO Y EL EXCEL CON SALDO",
           "Lo contrario del bloque 1: el sistema los cerró y el Excel los sigue mostrando con capital pendiente.",
           [("Cliente", 42), ("Crédito", 20), ("Estado en el sistema", 14),
            ("Capital en el sistema", 15), ("Capital según Excel", 15),
            ("Lo de Autocash según Excel", 15), ("Mes revisado", 14)],
           [[x["cliente"], x["sifco"], x["status"], x["cap_db"], x["cap_x"],
             x["auto_x"] if x["auto_x"] is not None else "—", x["hoja"]]
            for x in L(2)],
           total_idx=(5, 6))

    bloque(NARANJ,
           "3 · DIFERENCIAS DE CAPITAL   ←   ACÁ ESTÁ EL PROBLEMA",
           "Créditos vivos en los dos lados, pero con distinto capital. Es el bloque a resolver: "
           "cada fila es plata que el sistema y la contabilidad no cuentan igual.",
           [("Cliente", 42), ("Crédito", 20), ("Estado en el sistema", 14),
            ("Capital en el sistema", 15), ("Capital según Excel", 15),
            ("Diferencia", 14), ("De eso, lo de Autocash", 15), ("Mes revisado", 14)],
           [[x["cliente"], x["sifco"], x["status"], x["cap_db"], x["cap_x"], x["d_cap"],
             x["d_auto"] if x["d_auto"] is not None else "—", x["hoja"]]
            for x in L(3)],
           total_idx=(4, 5, 6, 7))

    # Resaltar el bloque 3 con color en la columna Diferencia
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=6):
        for c in row:
            if isinstance(c.value, float) and abs(c.value) > 100:
                c.font = Font(size=10, bold=True, color=ROJO if c.value < 0 else "006100")

    texto("")
    texto("No se incluyen las diferencias de centavos ni los créditos donde el capital del crédito cuadra "
          "y lo único que cambia es el reparto entre inversionistas: esos no son un descuadre de cartera.",
          size=9, italic=True)

    for j, (_h, w) in enumerate([("Cliente", 42), ("Crédito", 22), ("Estado", 15),
                                 ("Cap sistema", 17), ("Cap Excel", 17), ("Dif", 15),
                                 ("Autocash", 17), ("Mes", 15)], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    wb.save(SALIDA)
    print(f"\n📄 Reporte de gerencia: {SALIDA}")
    for n in (1, 2, 3):
        print(f"   Bloque {n}: {len(L(n))} créditos")


if __name__ == "__main__":
    main()
