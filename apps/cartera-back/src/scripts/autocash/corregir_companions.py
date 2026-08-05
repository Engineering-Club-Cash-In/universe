#!/usr/bin/env python3
"""Corrige el comparativo Autocash por 'companions': créditos que el Excel reparte en varios
SIFCO correlativos (no con el sufijo _N) en vez de en una sola familia.

Regla de detección: dentro de una hoja, se agrupan las filas por (nombre del cliente, # de cuota).
Si en el grupo hay EXACTAMENTE UN SIFCO que existe en la DB y los demás no existen, esos otros
son rebanadas del mismo crédito y se suman. Si hay dos o más que sí existen en la DB, son
créditos distintos del mismo cliente y NO se tocan.

Actualiza el archivo del usuario en sitio, respetando las columnas que ya recortó.
"""
import csv
import glob
import importlib.util
import os
import shutil
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore")

AQUI = os.path.dirname(os.path.abspath(__file__))
RUTA = os.path.expanduser("~/Descargas/Comparativo_Autocash_DB_vs_Excel.xlsx")
TOL = 1.00

_spec = importlib.util.spec_from_file_location("base", os.path.join(AQUI, "comparativo_autocash.py"))
B = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(B)

VERDE = PatternFill("solid", fgColor="C6EFCE")
AMBAR = PatternFill("solid", fgColor="FFEB9C")
ROJO = PatternFill("solid", fgColor="FFC7CE")


def cabecera(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [B.norm_hdr(v) for v in row]
        if "# credito sifco" in vals:
            hdr = {}
            for j, v in enumerate(vals):
                if v and v not in hdr:
                    hdr[v] = j
            return i, hdr
    return None, None


def main():
    db = {r["numero_credito_sifco"]: r for r in csv.DictReader(open(os.path.join(AQUI, "db_creditos.csv")))}
    en_db = set(db)

    ruta_xls = [f for f in glob.glob(os.path.expanduser("~/Descargas/*.xlsx")) if "NUEVA 3.0" in f][0]
    wbx = openpyxl.load_workbook(ruta_xls, read_only=True, data_only=True)
    hojas = B.hojas_mes(wbx)

    # ---- Pase 1: descubrir companions
    comp = defaultdict(set)          # sifco_db -> {sifcos del Excel que son rebanadas suyas}
    for _, hoja, _a in hojas:
        ws = wbx[hoja]
        hr, hdr = cabecera(ws)
        if hr is None or "#" not in hdr or "nombre" not in hdr:
            continue
        grupos = defaultdict(set)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if not base:
                continue
            nom = B.norm_nombre(row[hdr["nombre"]] if hdr["nombre"] < len(row) else None)
            n = row[hdr["#"]] if hdr["#"] < len(row) else None
            if not nom or n is None:
                continue
            grupos[(nom, str(n))].add(base)
        for bases in grupos.values():
            if len(bases) < 2:
                continue
            dentro = [b for b in bases if b in en_db]
            if len(dentro) != 1:
                continue                       # 0 o ≥2 en la DB -> no se asume nada
            comp[dentro[0]] |= (bases - {dentro[0]})

    autocash = sorted(k for k, v in db.items() if v["es_autocash"] == "1")
    afectados = {s: sorted(comp[s]) for s in autocash if comp.get(s)}
    print("Créditos Autocash con rebanadas en SIFCO aparte:")
    for s, c in afectados.items():
        print(f"  {s} -> {c}")

    interes = set()
    for s in autocash:
        interes.add(s)
        interes |= comp.get(s, set())

    # ---- Pase 2: recomputar la foto (última hoja pagada) con las rebanadas incluidas
    de_interes = {s: {c for c in [s] + list(comp.get(s, set()))} for s in autocash}
    dueno = {}
    for s, ks in de_interes.items():
        for k in ks:
            dueno[k] = s

    pagada, vista = {}, {}
    for _, hoja, _a in hojas:
        ws = wbx[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_cap, c_cuo = hdr.get("capital restante"), hdr.get("cuota")
        c_pag, c_inv = hdr.get("pagado"), hdr.get("inversionista")
        agg = defaultdict(lambda: {"cap": None, "cuota": None, "inv": [], "pagado": [], "sifcos": []})
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if base not in dueno:
                continue
            a = agg[dueno[base]]
            a["sifcos"].append(str(row[hdr["# credito sifco"]]).strip())
            if c_cap is not None and c_cap < len(row):
                v = B.num(row[c_cap])
                if v is not None:
                    a["cap"] = (a["cap"] or 0.0) + v
            if c_cuo is not None and c_cuo < len(row):
                v = B.num(row[c_cuo])
                if v is not None:
                    a["cuota"] = (a["cuota"] or 0.0) + v
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(B.norm_hdr(row[c_pag]))
            if c_inv is not None and c_inv < len(row) and row[c_inv]:
                nm = str(row[c_inv]).strip()
                if nm not in a["inv"]:
                    a["inv"].append(nm)
        for s, a in agg.items():
            foto = {"hoja": hoja, "cap": a["cap"], "cuota": a["cuota"] if c_cuo is not None else None,
                    "inv": list(a["inv"]), "sifcos": a["sifcos"]}
            vista[s] = foto
            if any(p in B.PAGADO_OK for p in a["pagado"]):
                pagada[s] = foto

    # ---- Aplicar al archivo del usuario
    bkp = RUTA.replace(".xlsx", "_backup_pre_companions.xlsx")
    shutil.copy2(RUTA, bkp)
    print(f"\nbackup -> {bkp}")

    wb = openpyxl.load_workbook(RUTA)
    ws1 = wb["Créditos comparativos"]
    hdr1 = [c.value for c in ws1[1]]
    ix = {h: i + 1 for i, h in enumerate(hdr1) if h}

    cambios = []
    for r in range(2, ws1.max_row + 1):
        sifco = ws1.cell(row=r, column=1).value
        if not sifco or sifco not in afectados:
            continue
        foto = pagada.get(sifco) or vista.get(sifco)
        if not foto:
            continue
        cap_db = float(ws1.cell(row=r, column=ix["CAPITAL crédito (DB)"]).value or 0)
        cuo_db = float(ws1.cell(row=r, column=ix["CUOTA crédito (DB)"]).value or 0)
        antes = (ws1.cell(row=r, column=ix["CAPITAL Excel"]).value,
                 ws1.cell(row=r, column=ix["CUOTA Excel"]).value)

        cap_x = round(foto["cap"], 2) if foto["cap"] is not None else None
        cuo_x = round(foto["cuota"], 2) if foto["cuota"] is not None else None
        ws1.cell(row=r, column=ix["Hoja Excel usada"], value=foto["hoja"])
        ws1.cell(row=r, column=ix["Inversionistas (Excel)"], value=" | ".join(foto["inv"]))
        ws1.cell(row=r, column=ix["CAPITAL Excel"], value=cap_x if cap_x is not None else "sin dato")
        ws1.cell(row=r, column=ix["CUOTA Excel"], value=cuo_x if cuo_x is not None else "sin dato")
        d_cap = round(cap_db - cap_x, 2) if cap_x is not None else "sin dato Excel"
        d_cuo = round(cuo_db - cuo_x, 2) if cuo_x is not None else "sin dato Excel"
        for col, val in ((ix["Dif capital: DB - Excel"], d_cap), (ix["Dif cuota: DB - Excel"], d_cuo)):
            c = ws1.cell(row=r, column=col, value=val)
            if isinstance(val, float):
                c.number_format = '#,##0.00'
                c.fill = VERDE if abs(val) <= TOL else (AMBAR if abs(val) <= 100 else ROJO)
        uv = vista.get(sifco)
        if uv and "Ref: última hoja donde aparece" in ix:
            ws1.cell(row=r, column=ix["Ref: última hoja donde aparece"], value=uv["hoja"])
            ws1.cell(row=r, column=ix["Ref: capital en esa hoja"],
                     value=round(uv["cap"], 2) if uv["cap"] is not None else "")
        for col in (ix["CAPITAL Excel"], ix["CUOTA Excel"]):
            ws1.cell(row=r, column=col).number_format = '#,##0.00'
        cambios.append((sifco, antes, (cap_x, cuo_x), d_cap, d_cuo, afectados[sifco]))

    # Hoja 2: quitar los que resultaron ser rebanadas de un crédito que sí está en la DB
    ws2 = wb["No aparecen en sistema"]
    rebanadas = {c for cs in afectados.values() for c in cs}
    quitar = [r for r in range(ws2.max_row, 1, -1)
              if ws2.cell(row=r, column=1).value in rebanadas]
    quitados = []
    for r in quitar:
        quitados.append((ws2.cell(row=r, column=1).value, ws2.cell(row=r, column=2).value))
        ws2.delete_rows(r)
    n2 = len([r for r in ws2.iter_rows(values_only=True) if r[0]]) - 1
    ws2.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws2.max_column)}{n2 + 1}"

    wb.save(RUTA)

    print("\nFilas corregidas en 'Créditos comparativos':")
    for s, antes, ahora, dc, dq, cs in cambios:
        print(f"  {s}  (+{', '.join(cs)})")
        print(f"     CAPITAL Excel: {antes[0]:>12,.2f} -> {ahora[0]:>12,.2f}   dif DB-Excel ahora: {dc}")
        print(f"     CUOTA   Excel: {antes[1]:>12,.2f} -> {ahora[1]:>12,.2f}   dif DB-Excel ahora: {dq}")
    print(f"\nQuitados de 'No aparecen en sistema': {quitados}  -> quedan {n2}")


if __name__ == "__main__":
    main()
