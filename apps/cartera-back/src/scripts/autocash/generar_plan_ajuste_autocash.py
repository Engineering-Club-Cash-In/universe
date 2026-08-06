#!/usr/bin/env python3
"""
⚠️  NO USAR — SUPERSEDIDO. Se conserva solo como referencia.

Genera el plan de la propuesta ORIGINAL de cuadre (capital Y cuota alineados al Excel en
los 72 créditos vivos). Contabilidad revisó el comparativo y autorizó otra cosa: un
subconjunto, solo capital, y con reglas por color. El flujo vivo es:

    preview_ajuste_conta.py  →  plan_ajuste_conta.json  →  aplicarAjusteConta.ts

Ver README.md.

────────────────────────────────────────────────────────────────────────────────

Genera el PLAN de ajuste de los créditos de Autocash contra el Excel de cartera.

Este script NO escribe nada en la base de datos: solo hace SELECT y produce un JSON
(el "plan") que después consume `aplicarPlanAjusteAutocash.ts`.

Reglas de negocio (acordadas):
  1. Solo se ajustan créditos en estado ACTIVO, MOROSO o EN_CONVENIO.
  2. `creditos.capital` y `creditos.cuota` se alinean a lo que dice el Excel.
  3. Del reparto por inversionista SOLO se mueve la fila de Autocash
     (monto_aportado y cuota_inversionista). Ningún otro inversionista se toca.
  4. Si el Excel indica que otro inversionista necesita ajuste, va a
     `revisar_manual` para evaluarlo a mano.
  5. Los créditos que el Excel da por liquidados (Capital restante = 0) se EXCLUYEN:
     ahí el Excel usa la columna Cuota para el finiquito, no para la cuota mensual.

Del Excel se toma la foto de la ÚLTIMA hoja mensual donde el crédito aparece con
Pagado ∈ {Si, Atrasado}, sumando todas sus filas (una por inversionista).

Uso:
    python3 generar_plan_ajuste_autocash.py \
        --excel "~/Descargas/Cartera Préstamos (Cash-In) NUEVA 3.0.xlsx" \
        --salida plan_ajuste_autocash.json

La cadena de conexión sale de SUPABASE_DB_URL (o --db-url). Se usa solo para leer.
"""
import argparse
import csv
import glob
import io
import json
import os
import re
import subprocess
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl:  pip install openpyxl")

# ─────────────────────────────────────────────────────────────── configuración
INVERSIONISTA_OBJETIVO = "Autocash S.A."
ESTADOS_AJUSTABLES = ("ACTIVO", "MOROSO", "EN_CONVENIO")
TOLERANCIA = Decimal("1.00")      # Q; por debajo de esto no vale la pena tocar el registro
PAGADO_OK = {"si", "atrasado"}    # qué cuenta como "el crédito pagó ese mes"

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

Q = Decimal("0.01")


# ────────────────────────────────────────────────────────────────── utilidades
def sin_tildes(s):
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c))


def norm_hdr(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", sin_tildes(v)).strip().lower()


def norm_nombre(v):
    if v is None:
        return ""
    s = sin_tildes(v).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def base_sifco(v):
    """Quita el sufijo _N con que el Excel parte un crédito entre inversionistas."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "#n/a", "#ref!", "0"):
        return None
    return re.sub(r"_\d+$", "", s)


def dec(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "").replace("Q", "").replace(" ", "")
    if not s or s.startswith("#"):
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def q2(v):
    return None if v is None else v.quantize(Q)


def f(v):
    return None if v is None else float(v)


# ──────────────────────────────────────────────────────────────────── base
def consultar(db_url, sql):
    """Corre un SELECT y devuelve list[dict]. Solo lectura."""
    out = subprocess.run(
        ["psql", db_url, "-v", "ON_ERROR_STOP=1", "--no-psqlrc", "-c",
         f"COPY ({sql}) TO STDOUT WITH CSV HEADER"],
        capture_output=True, text=True,
    )
    if out.returncode != 0:
        sys.exit(f"Error consultando la base:\n{out.stderr}")
    return list(csv.DictReader(io.StringIO(out.stdout)))


SQL_CREDITOS = """
SELECT c.credito_id,
       c.numero_credito_sifco,
       u.nombre  AS cliente,
       c."statusCredit" AS status,
       c.capital,
       c.cuota,
       c.cuota_interes,
       c.porcentaje_interes,
       c.seguro_10_cuotas,
       c.gps,
       c.membresias_pago,
       c.plazo,
       c.no_amortiza_capital
FROM cartera.creditos c
JOIN cartera.usuarios u ON u.usuario_id = c.usuario_id
"""

SQL_PARTICIPACIONES = """
SELECT c.numero_credito_sifco,
       ci.id           AS credito_inversionista_id,
       ci.inversionista_id,
       i.nombre        AS inversionista,
       ci.monto_aportado,
       ci.cuota_inversionista,
       ci.porcentaje_participacion_inversionista,
       ci.porcentaje_cash_in,
       e.id            AS espejo_id,
       e.monto_aportado      AS espejo_monto_aportado,
       e.cuota_inversionista AS espejo_cuota_inversionista
FROM cartera.creditos_inversionistas ci
JOIN cartera.creditos c       ON c.credito_id = ci.credito_id
JOIN cartera.inversionistas i ON i.inversionista_id = ci.inversionista_id
LEFT JOIN cartera.creditos_inversionistas_espejo e
       ON e.credito_id = ci.credito_id AND e.inversionista_id = ci.inversionista_id
"""


# ──────────────────────────────────────────────────────────────────── excel
def hojas_mes(wb):
    out = []
    for name in wb.sheetnames:
        if name.strip().endswith("_"):      # duplicados en formato viejo
            continue
        m = re.fullmatch(r"\s*([A-Za-zÁÉÍÓÚáéíóú]+)\s+(\d{4})\s*", name)
        if not m:
            continue
        mes = MESES.get(sin_tildes(m.group(1)).lower())
        if mes:
            out.append((int(m.group(2)) * 100 + mes, name))
    out.sort()
    return out


def cabecera(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [norm_hdr(v) for v in row]
        if "# credito sifco" in vals:
            hdr = {}
            for j, v in enumerate(vals):
                if v and v not in hdr:
                    hdr[v] = j
            return i, hdr
    return None, None


def detectar_companions(wb, hojas, sifcos_db):
    """El Excel a veces parte un crédito en SIFCO correlativos en vez de usar _N.
    Se agrupa por (cliente, # de cuota) dentro de la hoja: si en el grupo hay
    EXACTAMENTE un SIFCO que existe en la base, los demás son rebanadas suyas.
    Con dos o más en la base son créditos distintos del mismo cliente: no se tocan."""
    comp = defaultdict(set)
    for _, hoja in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None or "#" not in hdr or "nombre" not in hdr:
            continue
        grupos = defaultdict(set)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            b = base_sifco(row[hdr["# credito sifco"]])
            if not b:
                continue
            nom = norm_nombre(row[hdr["nombre"]] if hdr["nombre"] < len(row) else None)
            n = row[hdr["#"]] if hdr["#"] < len(row) else None
            if not nom or n is None:
                continue
            grupos[(nom, str(n))].add(b)
        for bases in grupos.values():
            if len(bases) < 2:
                continue
            dentro = [b for b in bases if b in sifcos_db]
            if len(dentro) == 1:
                comp[dentro[0]] |= (bases - {dentro[0]})
    return comp


def fotos_excel(ruta, sifcos_db):
    """-> {sifco_db: {hoja, capital, cuota, filas:[{sifco, inversionista, capital, cuota}]}}
    con la foto de la última hoja mensual donde el crédito aparece pagado."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = hojas_mes(wb)
    comp = detectar_companions(wb, hojas, sifcos_db)

    dueno = {}
    for s in sifcos_db:
        dueno[s] = s
        for c in comp.get(s, ()):
            dueno[c] = s

    pagada, vista = {}, {}
    for _, hoja in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_cap, c_cuo = hdr.get("capital restante"), hdr.get("cuota")
        c_pag, c_inv = hdr.get("pagado"), hdr.get("inversionista")

        agg = defaultdict(lambda: {"capital": None, "cuota": None, "filas": [], "pagado": []})
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            b = base_sifco(row[hdr["# credito sifco"]])
            s = dueno.get(b)
            if not s:
                continue
            a = agg[s]
            cap = dec(row[c_cap]) if c_cap is not None and c_cap < len(row) else None
            cuo = dec(row[c_cuo]) if c_cuo is not None and c_cuo < len(row) else None
            if cap is not None:
                a["capital"] = (a["capital"] or Decimal(0)) + cap
            if cuo is not None:
                a["cuota"] = (a["cuota"] or Decimal(0)) + cuo
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(norm_hdr(row[c_pag]))
            a["filas"].append({
                "sifco": str(row[hdr["# credito sifco"]]).strip(),
                "inversionista": (str(row[c_inv]).strip()
                                  if c_inv is not None and c_inv < len(row) and row[c_inv] else None),
                "capital": cap, "cuota": cuo,
            })

        for s, a in agg.items():
            foto = {"hoja": hoja, "capital": a["capital"],
                    "cuota": a["cuota"] if c_cuo is not None else None,
                    "filas": a["filas"], "tiene_col_cuota": c_cuo is not None}
            vista[s] = foto
            if any(p in PAGADO_OK for p in a["pagado"]):
                pagada[s] = foto

    return pagada, vista, {k: sorted(v) for k, v in comp.items()}


# ───────────────────────────────────────────────────────────────────── plan
def main():
    ap = argparse.ArgumentParser(description="Genera el plan de ajuste Autocash (solo lectura)")
    ap.add_argument("--excel", help="Ruta del Excel de cartera. Por defecto busca 'NUEVA 3.0' en ~/Descargas")
    ap.add_argument("--salida", default="plan_ajuste_autocash.json")
    ap.add_argument("--db-url", default=os.environ.get("SUPABASE_DB_URL"))
    ap.add_argument("--inversionista", default=INVERSIONISTA_OBJETIVO)
    ap.add_argument("--tolerancia", type=Decimal, default=TOLERANCIA)
    args = ap.parse_args()

    if not args.db_url:
        sys.exit("Falta la cadena de conexión: exportá SUPABASE_DB_URL o pasá --db-url")

    ruta = os.path.expanduser(args.excel) if args.excel else None
    if not ruta:
        cand = [f for f in glob.glob(os.path.expanduser("~/Descargas/*.xlsx")) if "NUEVA 3.0" in f]
        if not cand:
            sys.exit("No encontré el Excel; pasalo con --excel")
        ruta = cand[0]
    print(f"Excel: {ruta}")

    creditos = {r["numero_credito_sifco"]: r for r in consultar(args.db_url, SQL_CREDITOS)}
    parts = defaultdict(list)
    for r in consultar(args.db_url, SQL_PARTICIPACIONES):
        parts[r["numero_credito_sifco"]].append(r)
    print(f"DB: {len(creditos)} créditos, {sum(len(v) for v in parts.values())} participaciones")

    objetivo_norm = norm_nombre(args.inversionista)
    del_objetivo = {s for s, rows in parts.items()
                    if any(norm_nombre(r["inversionista"]) == objetivo_norm for r in rows)}
    print(f"Créditos con {args.inversionista}: {len(del_objetivo)}")

    pagada, vista, comp = fotos_excel(ruta, set(creditos))

    aplicar, excluidos, revisar = [], [], []
    tol = args.tolerancia

    for sifco in sorted(del_objetivo):
        c = creditos[sifco]
        filas = parts[sifco]
        auto = next((r for r in filas if norm_nombre(r["inversionista"]) == objetivo_norm), None)

        def excluir(motivo, **extra):
            excluidos.append({"numero_credito_sifco": sifco, "cliente": c["cliente"],
                              "status": c["status"], "motivo": motivo, **extra})

        if c["status"] not in ESTADOS_AJUSTABLES:
            excluir(f"status {c['status']} no ajustable")
            continue

        foto = pagada.get(sifco)
        if not foto:
            excluir("sin foto en el Excel (nunca aparece marcado como pagado)")
            continue
        if foto["capital"] is None:
            excluir("el Excel no trae Capital restante en esa hoja", hoja=foto["hoja"])
            continue
        if not foto["tiene_col_cuota"] or foto["cuota"] is None:
            excluir("el Excel no trae columna Cuota en esa hoja (hojas previas a Julio 2024)",
                    hoja=foto["hoja"])
            continue

        cap_obj, cuo_obj = q2(foto["capital"]), q2(foto["cuota"])

        # Regla 5: el Excel marca la liquidación con Capital restante = 0 y mete el
        # finiquito en la columna Cuota. Ahí no se ajusta nada automáticamente.
        if cap_obj == 0:
            excluir("el Excel lo da por liquidado (Capital restante = 0); la columna Cuota "
                    "trae el finiquito, no la cuota mensual",
                    hoja=foto["hoja"], capital_db=f(dec(c["capital"])),
                    cuota_excel_finiquito=f(cuo_obj))
            continue

        cap_act, cuo_act = q2(dec(c["capital"])), q2(dec(c["cuota"]))
        sum_aport = q2(sum((dec(r["monto_aportado"]) or Decimal(0)) for r in filas))
        sum_cuota_inv = q2(sum((dec(r["cuota_inversionista"]) or Decimal(0)) for r in filas))

        # Autocash absorbe la diferencia para que la suma cuadre con el crédito.
        auto_aport_act = q2(dec(auto["monto_aportado"]))
        auto_cuota_act = q2(dec(auto["cuota_inversionista"]))
        auto_aport_obj = q2(auto_aport_act + (cap_obj - sum_aport))
        auto_cuota_obj = q2(auto_cuota_act + (cuo_obj - sum_cuota_inv))

        if auto_aport_obj < 0:
            excluir("el ajuste dejaría el monto_aportado de Autocash en negativo",
                    hoja=foto["hoja"], monto_aportado_resultante=f(auto_aport_obj))
            continue
        if auto_cuota_obj < 0:
            excluir("el ajuste dejaría la cuota_inversionista de Autocash en negativo",
                    hoja=foto["hoja"], cuota_resultante=f(auto_cuota_obj))
            continue

        # ── Regla 4: contrastar cada OTRO inversionista contra su fila del Excel.
        #    No se toca nada; solo se reporta para evaluarlo a mano.
        por_inv_excel = defaultdict(lambda: {"capital": Decimal(0), "cuota": Decimal(0), "sifcos": []})
        for fila in foto["filas"]:
            if not fila["inversionista"]:
                continue
            k = norm_nombre(fila["inversionista"])
            por_inv_excel[k]["capital"] += (fila["capital"] or Decimal(0))
            por_inv_excel[k]["cuota"] += (fila["cuota"] or Decimal(0))
            por_inv_excel[k]["sifcos"].append(fila["sifco"])

        otros = []
        for r in filas:
            k = norm_nombre(r["inversionista"])
            if k == objetivo_norm:
                continue
            x = por_inv_excel.get(k)
            fila_otro = {
                "inversionista": r["inversionista"],
                "monto_aportado_db": f(q2(dec(r["monto_aportado"]))),
                "cuota_inversionista_db": f(q2(dec(r["cuota_inversionista"]))),
                "monto_aportado_excel": f(q2(x["capital"])) if x else None,
                "cuota_inversionista_excel": f(q2(x["cuota"])) if x else None,
                "en_excel": x is not None,
            }
            if x is not None:
                d_cap = q2(dec(r["monto_aportado"])) - q2(x["capital"])
                d_cuo = q2(dec(r["cuota_inversionista"])) - q2(x["cuota"])
                fila_otro["dif_capital"] = f(d_cap)
                fila_otro["dif_cuota"] = f(d_cuo)
                if abs(d_cap) > tol or abs(d_cuo) > tol:
                    revisar.append({
                        "numero_credito_sifco": sifco, "cliente": c["cliente"],
                        "hoja": foto["hoja"], "inversionista": r["inversionista"],
                        "monto_aportado_db": f(q2(dec(r["monto_aportado"]))),
                        "monto_aportado_excel": f(q2(x["capital"])),
                        "dif_capital": f(d_cap),
                        "cuota_inversionista_db": f(q2(dec(r["cuota_inversionista"]))),
                        "cuota_inversionista_excel": f(q2(x["cuota"])),
                        "dif_cuota": f(d_cuo),
                        "nota": "Otro inversionista fuera de cuadre. El script NO lo toca: ajustar a mano.",
                    })
            else:
                revisar.append({
                    "numero_credito_sifco": sifco, "cliente": c["cliente"],
                    "hoja": foto["hoja"], "inversionista": r["inversionista"],
                    "monto_aportado_db": f(q2(dec(r["monto_aportado"]))),
                    "monto_aportado_excel": None, "dif_capital": None,
                    "cuota_inversionista_db": f(q2(dec(r["cuota_inversionista"]))),
                    "cuota_inversionista_excel": None, "dif_cuota": None,
                    "nota": "Está en la base pero no tiene fila en el Excel de esa hoja.",
                })
        for k, x in por_inv_excel.items():
            if k != objetivo_norm and not any(norm_nombre(r["inversionista"]) == k for r in filas):
                revisar.append({
                    "numero_credito_sifco": sifco, "cliente": c["cliente"],
                    "hoja": foto["hoja"], "inversionista": x["sifcos"] and k or k,
                    "monto_aportado_db": None, "monto_aportado_excel": f(q2(x["capital"])),
                    "dif_capital": None, "cuota_inversionista_db": None,
                    "cuota_inversionista_excel": f(q2(x["cuota"])), "dif_cuota": None,
                    "nota": "Está en el Excel pero no tiene participación en la base.",
                })

        avisos = []
        if auto["espejo_id"]:
            e_ap = q2(dec(auto["espejo_monto_aportado"]))
            e_cu = q2(dec(auto["espejo_cuota_inversionista"]))
            avisos.append(
                f"Fila ESPEJO de Autocash (monto {e_ap}, cuota {e_cu}): se empareja con el padre."
            )
        else:
            avisos.append(
                "Autocash NO tiene fila espejo en este crédito. El script no la crea: revisar a mano."
            )
        if comp.get(sifco):
            avisos.append("El Excel reparte este crédito en varios SIFCO: "
                          + ", ".join([sifco] + comp[sifco]))
        if abs(cap_act - cap_obj) <= tol and abs(cuo_act - cuo_obj) <= tol:
            avisos.append("Ya estaba cuadrado dentro de la tolerancia; el ajuste es cosmético.")

        aplicar.append({
            "numero_credito_sifco": sifco,
            "credito_id": int(c["credito_id"]),
            "cliente": c["cliente"],
            "status": c["status"],
            "hoja_excel": foto["hoja"],
            "sifcos_excel": sorted({fl["sifco"] for fl in foto["filas"]}),
            "capital": {"actual": f(cap_act), "objetivo": f(cap_obj), "delta": f(q2(cap_obj - cap_act))},
            "cuota": {"actual": f(cuo_act), "objetivo": f(cuo_obj), "delta": f(q2(cuo_obj - cuo_act))},
            "autocash": {
                "inversionista_id": int(auto["inversionista_id"]),
                "credito_inversionista_id": int(auto["credito_inversionista_id"]),
                "monto_aportado": {"actual": f(auto_aport_act), "objetivo": f(auto_aport_obj),
                                   "delta": f(q2(auto_aport_obj - auto_aport_act))},
                "cuota_inversionista": {"actual": f(auto_cuota_act), "objetivo": f(auto_cuota_obj),
                                        "delta": f(q2(auto_cuota_obj - auto_cuota_act))},
                "tiene_espejo": bool(auto["espejo_id"]),
            },
            "suma_inversionistas_antes": {"monto_aportado": f(sum_aport), "cuota": f(sum_cuota_inv)},
            "otros_inversionistas": otros,
            "avisos": avisos,
        })

    # ── FASE 2 (solo relevamiento, no se ejecuta todavía) ──────────────────────
    # Créditos que en el Excel figuran con Autocash pero que en la base ya no
    # tienen participación de Autocash. Mismos estados ajustables.
    fase2 = []
    for sifco, foto in pagada.items():
        if sifco in del_objetivo:
            continue
        c = creditos.get(sifco)
        if not c or c["status"] not in ESTADOS_AJUSTABLES:
            continue
        invs_excel = {norm_nombre(fl["inversionista"]): fl["inversionista"]
                      for fl in foto["filas"] if fl["inversionista"]}
        if objetivo_norm not in invs_excel:
            continue
        cap_auto = sum((fl["capital"] or Decimal(0)) for fl in foto["filas"]
                       if norm_nombre(fl["inversionista"]) == objetivo_norm)
        cuo_auto = sum((fl["cuota"] or Decimal(0)) for fl in foto["filas"]
                       if norm_nombre(fl["inversionista"]) == objetivo_norm)
        fase2.append({
            "numero_credito_sifco": sifco,
            "credito_id": int(c["credito_id"]),
            "cliente": c["cliente"],
            "status": c["status"],
            "hoja_excel": foto["hoja"],
            "capital_credito_db": f(q2(dec(c["capital"]))),
            "cuota_credito_db": f(q2(dec(c["cuota"]))),
            "capital_credito_excel": f(q2(foto["capital"])) if foto["capital"] is not None else None,
            "cuota_credito_excel": f(q2(foto["cuota"])) if foto["cuota"] is not None else None,
            "autocash_en_excel": {"capital": f(q2(cap_auto)), "cuota": f(q2(cuo_auto))},
            "inversionistas_db": [r["inversionista"] for r in parts.get(sifco, [])],
            "inversionistas_excel": sorted(invs_excel.values()),
            "nota": "El Excel lo tiene con Autocash y la base no. Requiere alta/reemplazo de "
                    "participación, no un simple ajuste de montos. Ver README (Fase 2).",
        })
    fase2.sort(key=lambda x: x["numero_credito_sifco"])

    plan = {
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "excel": ruta,
        "inversionista_objetivo": args.inversionista,
        "estados_ajustables": list(ESTADOS_AJUSTABLES),
        "tolerancia": float(tol),
        "resumen": {
            "creditos_del_inversionista": len(del_objetivo),
            "a_aplicar": len(aplicar),
            "excluidos": len(excluidos),
            "inversionistas_a_revisar_manual": len(revisar),
            "fase2_candidatos": len(fase2),
            "capital_total_delta": round(sum(x["capital"]["delta"] for x in aplicar), 2),
            "cuota_total_delta": round(sum(x["cuota"]["delta"] for x in aplicar), 2),
        },
        "aplicar": aplicar,
        "excluidos": excluidos,
        "revisar_manual": revisar,
        "fase2_candidatos": fase2,
    }

    with open(args.salida, "w", encoding="utf-8") as fh:
        json.dump(plan, fh, ensure_ascii=False, indent=2)

    csv_path = os.path.splitext(args.salida)[0] + "_revisar_manual.csv"
    if revisar:
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(revisar[0].keys()))
            w.writeheader()
            w.writerows(revisar)

    print(f"\nPlan escrito en {args.salida}")
    for k, v in plan["resumen"].items():
        print(f"  {k}: {v}")
    if revisar:
        print(f"  (detalle de inversionistas a revisar en {csv_path})")
    print("\nEste script no modificó nada. Para aplicar:  bun run src/scripts/autocash/aplicarPlanAjusteAutocash.ts")


if __name__ == "__main__":
    main()
