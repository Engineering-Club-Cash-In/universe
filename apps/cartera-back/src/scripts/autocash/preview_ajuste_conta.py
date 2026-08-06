#!/usr/bin/env python3
"""
PREVIEW del ajuste autorizado por contabilidad (lote Autocash).

Lee el Excel revisado por conta (`Comparativo_Autocash_DB_vs_Excel (1).xlsx`), cruza contra
la base y muestra CÓMO QUEDARÍA cada crédito. No escribe absolutamente nada: es el insumo
para pedir la autorización por correo.

Leyenda de colores de conta (columna T `capital correcto`):
    AMARILLO (FFFFFF00) -> se corrige solo el ESPEJO
    VERDE    (FF92D050) -> se corrige solo el PADRE
    CELESTE  (FF00B0F0) -> se corrige PADRE y ESPEJO
    NARANJA  (FFFFC000) -> se corrige PADRE y ESPEJO   (confirmado con Daniel: es un celeste)
    ROJO     (FFFF0000) -> crédito CANCELADO: el ESPEJO va a 0, el padre NO se toca

Reglas acordadas:
  · Si la fila trae un segundo monto (columna U), ESE es el capital correcto y manda sobre T.
  · Solo se mueve el `monto_aportado` de Autocash para que la suma cuadre con el capital
    correcto. Excepción autorizada: Crhistian Herrera, donde conta dio los montos de
    Autocash y de Cube por separado (columnas W y X) y se aplican los dos.
  · La CUOTA no se toca en este lote: conta solo autorizó capital.
"""
import argparse
import glob
import json
import os
import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

Q = Decimal("0.01")
AUTOCASH = "Autocash S.A."

# color de relleno -> qué registros se corrigen
ACCION = {
    "FFFFFF00": ("espejo", "Amarillo: corregir solo el espejo"),
    "FF92D050": ("padre", "Verde: corregir solo el padre"),
    "FF00B0F0": ("ambos", "Celeste: corregir padre y espejo"),
    "FFFFC000": ("ambos", "Naranja: corregir padre y espejo"),
    "FFFF0000": ("cancelado", "Rojo CANCELADO: espejo a 0, el padre no se toca"),
    "FFC00000": ("cancelado", "Rojo CANCELADO: espejo a 0, el padre no se toca"),
}

# Crhistian Herrera: conta dio los dos montos por separado (Autocash y Cube).
SIFCO_DOS_MONTOS = "01010214115040"

# Créditos que conta marcó pero que salen del lote (con el motivo que va al correo).
# Los tres primeros venían marcados con la PRIMERA versión del comparativo, antes de
# corregir los SIFCO correlativos: ahí Autocash ya está bien de los dos lados.
EXCLUIDOS = {
    "01010214105310": "Autocash ya está correcto en padre y espejo (Q21,654.56 en ambos). "
                      "El descuadre del espejo es de Javier Arzu y Diego Furlan.",
    "01010214109450": "Autocash ya está correcto en padre y espejo (Q19,575.98 en ambos). "
                      "El descuadre del espejo es de Anna Lisseth Lorenzo y Aura Floridalma.",
    "01010214117240": "Autocash ya está correcto en padre y espejo (Q23,158.51 en ambos). "
                      "El descuadre del espejo es de Richard Kachler.",
    "01010214108630": "Diferencia de centavos (Q0.04): no amerita movimiento.",
}

# ── LOTE 2: saneamiento interno ────────────────────────────────────────────────
# Créditos que conta no revisó y que están internamente incoherentes. Acá NO se
# cambia el capital del crédito: se mueve a Autocash para que SUM(padre) == capital,
# y se alinea su espejo al padre. Se puede alinear el espejo de Autocash sin riesgo
# porque este inversionista nunca ha sido liquidado (0 filas en cartera.liquidaciones).
SANEAMIENTO = {
    "01010214108640": {"opcional": False, "nota": "Solo Autocash en el crédito: el espejo estaba Q1,260.44 arriba del padre."},
    "01010214117420": {"opcional": False, "nota": "Solo Autocash en el crédito: el espejo estaba Q942.46 arriba del padre."},
    "01010214119450": {"opcional": False, "nota": "Solo Autocash en el crédito: el espejo estaba Q743.84 arriba del padre."},
    "01010214116550": {"opcional": False, "nota": "3 inversionistas. Autocash ya estaba alineado; absorbe los Q12.41 que "
                                                  "descuadran el padre. El espejo de Rodrigo Estrada (Q292.62) NO se toca."},
    "CRM-c554294a-1f1c-47dd-86cb-9a8f6afd204c":
                      {"opcional": True,  "nota": "2 inversionistas. Autocash ya estaba alineado; absorbe los Q42.79 que "
                                                  "descuadran el padre. El espejo de Bibiana Moyano (Q3.21) NO se toca."},
}

# Filas que conta dejó sin pintar y que se resolvieron aparte con Daniel.
ACCION_MANUAL = {
    # José Carlos Motta: se corrige todo. Ragnar Magaña no se toca, así que Autocash
    # absorbe el resto hasta llegar al capital correcto (en padre y en espejo).
    "01010214119500": ("ambos", "Celeste: corregir padre y espejo"),
}


def dec(v):
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "").replace("Q", "").replace(" ", "")
    try:
        return Decimal(s)
    except Exception:
        return None


def q2(v):
    return None if v is None else v.quantize(Q)


def fill_de(cell):
    if not cell.fill or cell.fill.fill_type is None:
        return None
    rgb = cell.fill.fgColor.rgb
    return rgb if isinstance(rgb, str) else None


def leer_conta(ruta):
    """-> list[dict] con lo que conta marcó, en orden de aparición."""
    wb = openpyxl.load_workbook(ruta)
    ws = wb["Créditos comparativos"]
    filas = []
    for r in range(2, ws.max_row + 1):
        sifco = ws.cell(row=r, column=1).value
        if not sifco:
            continue
        cT, cU = ws.cell(row=r, column=20), ws.cell(row=r, column=21)
        vT, vU = cT.value, cU.value
        if vT in (None, "") and vU in (None, ""):
            continue

        # El color operativo es el de la celda que trae el monto final; si esa
        # celda no tiene relleno, se hereda el de T.
        objetivo, celda_val = (dec(vU), cU) if dec(vU) is not None else (dec(vT), cT)
        color = fill_de(celda_val) or fill_de(cT)
        accion, leyenda = ACCION.get(color, (None, f"SIN COLOR RECONOCIDO ({color})"))

        es_cancelado = isinstance(vT, str) and "CANCELADO" in vT.upper()
        if es_cancelado:
            accion, leyenda, objetivo = "cancelado", ACCION["FFFF0000"][1], Decimal(0)
        elif str(sifco).strip() in ACCION_MANUAL:
            accion, leyenda = ACCION_MANUAL[str(sifco).strip()]

        filas.append({
            "fila_excel": r,
            "sifco": str(sifco).strip(),
            "cliente": str(ws.cell(row=r, column=2).value or ""),
            "objetivo": q2(objetivo) if objetivo is not None else None,
            "origen_monto": "columna U (segundo monto)" if dec(vU) is not None else "columna T",
            "accion": accion,
            "leyenda": leyenda,
            "color": color,
            "monto_T": q2(dec(vT)),
            "monto_U": q2(dec(vU)),
            "extra_W": q2(dec(ws.cell(row=r, column=23).value)),
            "extra_X": q2(dec(ws.cell(row=r, column=24).value)),
            "nota_conta": ws.cell(row=r, column=21).value if isinstance(ws.cell(row=r, column=21).value, str) else None,
        })
    return filas


def consultar(db_url, sql):
    out = subprocess.run(
        ["psql", db_url, "-v", "ON_ERROR_STOP=1", "--no-psqlrc", "-tAF", "\x1f", "-c", sql],
        capture_output=True, text=True,
    )
    if out.returncode != 0:
        sys.exit(f"Error consultando la base:\n{out.stderr}")
    return [l.split("\x1f") for l in out.stdout.splitlines() if l.strip()]


def cargar_db(db_url, sifcos):
    lista = ",".join("'" + s.replace("'", "''") + "'" for s in sifcos)
    creditos = {}
    for row in consultar(db_url, f"""
        SELECT c.numero_credito_sifco, c."statusCredit", c.capital, c.cuota, u.nombre
        FROM cartera.creditos c
        JOIN cartera.usuarios u ON u.usuario_id = c.usuario_id
        WHERE c.numero_credito_sifco IN ({lista})"""):
        creditos[row[0]] = {"status": row[1], "capital": q2(dec(row[2])), "cuota": q2(dec(row[3])),
                            "cliente": row[4]}

    invs = defaultdict(list)
    for row in consultar(db_url, f"""
        SELECT c.numero_credito_sifco, i.nombre, ci.monto_aportado, e.monto_aportado
        FROM cartera.creditos c
        JOIN cartera.creditos_inversionistas ci ON ci.credito_id = c.credito_id
        JOIN cartera.inversionistas i USING(inversionista_id)
        LEFT JOIN cartera.creditos_inversionistas_espejo e
               ON e.credito_id = ci.credito_id AND e.inversionista_id = ci.inversionista_id
        WHERE c.numero_credito_sifco IN ({lista})
        ORDER BY c.numero_credito_sifco, ci.monto_aportado DESC"""):
        invs[row[0]].append({
            "nombre": row[1],
            "padre": q2(dec(row[2])) or Decimal(0),
            "espejo": q2(dec(row[3])) if row[3] not in ("", None) else None,
        })
    return creditos, invs


def calcular(f, cred, inversionistas):
    """Devuelve el movimiento propuesto para un crédito."""
    obj = f["objetivo"]
    accion = f["accion"]
    auto = next((x for x in inversionistas if x["nombre"] == AUTOCASH), None)
    suma_padre = sum(x["padre"] for x in inversionistas)
    suma_espejo = sum((x["espejo"] or Decimal(0)) for x in inversionistas)

    mov = {
        "status": cred["status"],
        "n_inv": len(inversionistas),
        "otros": [x["nombre"] for x in inversionistas if x["nombre"] != AUTOCASH],
        "cap_padre_antes": cred["capital"],
        "suma_padre_antes": q2(suma_padre),
        "suma_espejo_antes": q2(suma_espejo),
        "auto_padre_antes": auto["padre"] if auto else None,
        "auto_espejo_antes": (auto["espejo"] if auto else None),
        "cap_padre_despues": cred["capital"],
        "suma_padre_despues": q2(suma_padre),
        "suma_espejo_despues": q2(suma_espejo),
        "auto_padre_despues": auto["padre"] if auto else None,
        "auto_espejo_despues": (auto["espejo"] if auto else None),
        "otros_movidos": [],
        "avisos": [],
    }
    if auto is None:
        mov["avisos"].append("Este crédito NO tiene participación de Autocash")
        return mov

    toca_padre = accion in ("padre", "ambos")
    toca_espejo = accion in ("espejo", "ambos", "cancelado")

    if f["sifco"] == SIFCO_DOS_MONTOS and f["extra_W"] is not None and f["extra_X"] is not None:
        # Autorización específica de conta: los dos inversionistas van a un monto dado.
        otro = next((x for x in inversionistas if x["nombre"] != AUTOCASH), None)
        if toca_padre:
            mov["auto_padre_despues"] = f["extra_W"]
            mov["cap_padre_despues"] = obj
            mov["suma_padre_despues"] = q2(f["extra_W"] + f["extra_X"])
        if toca_espejo:
            mov["auto_espejo_despues"] = f["extra_W"]
            mov["suma_espejo_despues"] = q2(f["extra_W"] + f["extra_X"])
        if otro:
            mov["otros_movidos"].append({
                "nombre": otro["nombre"],
                "padre_antes": otro["padre"], "padre_despues": f["extra_X"] if toca_padre else otro["padre"],
                "espejo_antes": otro["espejo"], "espejo_despues": f["extra_X"] if toca_espejo else otro["espejo"],
            })
        mov["avisos"].append(
            "EXCEPCIÓN: conta dio montos separados para Autocash y Cube; se mueven los dos.")
        return mov

    if accion == "cancelado":
        # El espejo de Autocash va a 0. El padre no se toca.
        mov["auto_espejo_despues"] = Decimal(0)
        mov["suma_espejo_despues"] = q2(suma_espejo - (auto["espejo"] or Decimal(0)))
        mov["avisos"].append(
            "El crédito se va a cancelar. Por el momento solo se deja el espejo en 0 para que "
            f"no se tome en liquidación; el padre sigue con Q{cred['capital']:,} (status {cred['status']}).")
        return mov

    if toca_padre:
        mov["cap_padre_despues"] = obj
        mov["auto_padre_despues"] = q2(auto["padre"] + (obj - suma_padre))
        mov["suma_padre_despues"] = obj
    if toca_espejo:
        mov["auto_espejo_despues"] = q2((auto["espejo"] or Decimal(0)) + (obj - suma_espejo))
        mov["suma_espejo_despues"] = obj

    # ¿El descuadre del espejo era de otro inversionista?
    if toca_espejo and len(inversionistas) > 1:
        dif_auto = (auto["padre"] - (auto["espejo"] or Decimal(0)))
        ajenos = [x["nombre"] for x in inversionistas
                  if x["nombre"] != AUTOCASH and x["espejo"] is not None
                  and abs(x["padre"] - x["espejo"]) > Decimal("0.01")]
        if ajenos and abs(dif_auto) <= Decimal("0.01"):
            mov["avisos"].append(
                "OJO: el descuadre padre↔espejo es de " + ", ".join(ajenos) +
                ", no de Autocash. Por decisión de Daniel, Autocash absorbe la diferencia.")

    for x in inversionistas:
        if x["nombre"] == AUTOCASH:
            continue
        if mov["auto_padre_despues"] is not None and toca_padre:
            pass
    return mov


def negativos(mov):
    for k in ("auto_padre_despues", "auto_espejo_despues"):
        v = mov.get(k)
        if v is not None and v < 0:
            return True
    return False


# ─────────────────────────────────────────────────────────────────── salida
AZUL = PatternFill("solid", fgColor="1F4E79")
GRIS = PatternFill("solid", fgColor="F2F2F2")
AMBAR = PatternFill("solid", fgColor="FFEB9C")
ROJO = PatternFill("solid", fgColor="FFC7CE")
VERDE = PatternFill("solid", fgColor="C6EFCE")


def escribir_excel(ruta, filas, movs, creditos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Preview del ajuste"
    enc = [
        "SIFCO", "Cliente", "Status", "Qué se corrige", "Capital correcto (conta)", "De dónde sale",
        "PADRE capital antes", "PADRE capital después",
        "PADRE Autocash antes", "PADRE Autocash después", "PADRE Autocash movimiento",
        "ESPEJO suma antes", "ESPEJO suma después",
        "ESPEJO Autocash antes", "ESPEJO Autocash después", "ESPEJO Autocash movimiento",
        "# inversionistas", "Otros inversionistas (no se tocan salvo nota)", "Observaciones",
    ]
    ws.append(enc)
    for c in range(1, len(enc) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = AZUL, Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 42

    def mov_txt(antes, despues):
        if antes is None or despues is None:
            return ""
        d = despues - antes
        if abs(d) < Decimal("0.005"):
            return "sin cambio"
        return f"{'+' if d > 0 else ''}{d:,.2f}"

    for f in filas:
        m = movs[f["sifco"]]
        otros = ", ".join(m["otros"]) or "—"
        for om in m["otros_movidos"]:
            otros += f"  ⚠ {om['nombre']}: {om['padre_antes']:,.2f} → {om['padre_despues']:,.2f}"
        ws.append([
            f["sifco"], f["cliente"], m["status"], f["leyenda"],
            float(f["objetivo"]) if f["objetivo"] is not None else "", f["origen_monto"],
            float(m["cap_padre_antes"]), float(m["cap_padre_despues"]),
            float(m["auto_padre_antes"]) if m["auto_padre_antes"] is not None else "",
            float(m["auto_padre_despues"]) if m["auto_padre_despues"] is not None else "",
            mov_txt(m["auto_padre_antes"], m["auto_padre_despues"]),
            float(m["suma_espejo_antes"]), float(m["suma_espejo_despues"]),
            float(m["auto_espejo_antes"]) if m["auto_espejo_antes"] is not None else "",
            float(m["auto_espejo_despues"]) if m["auto_espejo_despues"] is not None else "",
            mov_txt(m["auto_espejo_antes"], m["auto_espejo_despues"]),
            m["n_inv"], otros, "; ".join(m["avisos"]),
        ])

    dinero = [5, 7, 8, 9, 10, 12, 13, 14, 15]
    for r in range(2, ws.max_row + 1):
        for c in dinero:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        for c in (11, 16):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str) and cell.value not in ("", "sin cambio"):
                cell.fill = AMBAR
            elif cell.value == "sin cambio":
                cell.fill = GRIS
        obs = ws.cell(row=r, column=19)
        if obs.value:
            obs.fill = ROJO if "OJO" in str(obs.value) or "EXCEPCIÓN" in str(obs.value) else AMBAR
            obs.alignment = Alignment(wrap_text=True, vertical="top")

    anchos = {}
    for row in ws.iter_rows(values_only=True):
        for j, v in enumerate(row, 1):
            anchos[j] = min(52, max(anchos.get(j, 11), len(str(v)) + 2 if v is not None else 11))
    for j, w in anchos.items():
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(enc))}{ws.max_row}"
    wb.save(ruta)


ACCION_CORTA = {
    "padre": "Padre",
    "espejo": "Espejo",
    "ambos": "Padre y espejo",
    "cancelado": "Espejo → 0",
}

NOTA_MANUAL = {
    "01010214119500": "Ragnar Magaña no se toca; Autocash toma el monto restante "
                      "hasta llegar al capital correcto.",
}


def nota_corta(m, f):
    if f["sifco"] in NOTA_MANUAL:
        return NOTA_MANUAL[f["sifco"]]
    if f["accion"] is None:
        return "Falta el color de conta: sin definir"
    if f["accion"] == "cancelado":
        return ("Este crédito se va a cancelar. Por el momento solo se deja el espejo en 0 "
                "para que no se tome en liquidación.")
    for a in m["avisos"]:
        if a.startswith("EXCEPCIÓN"):
            otro = m["otros_movidos"][0]["nombre"] if m["otros_movidos"] else "otro inversionista"
            return f"También se mueve {otro} (autorizado por conta)"
        if a.startswith("OJO"):
            return "Autocash absorbe un descuadre que es de otros inversionistas"
    return ""


def escribir_excel_corto(ruta, filas, movs):
    """Versión de una sola mirada para gerencia: 7 columnas y totales."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    enc = ["Crédito", "Cliente", "Se corrige", "Capital hoy", "Capital correcto",
           "Ajuste Autocash PADRE", "Ajuste Autocash ESPEJO", "Nota"]

    ws.append(["AJUSTE DE CARTERA AUTOCASH — RESUMEN PARA AUTORIZACIÓN"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(enc))
    t = ws.cell(row=1, column=1)
    t.fill, t.font = AZUL, Font(bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append([f"{len(filas)} créditos autorizados por Contabilidad. "
               "Solo se mueve el capital (monto aportado) de Autocash; la cuota no se toca."])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(enc))
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="595959")
    ws.append([])

    hdr_row = 4
    ws.append(enc)
    for c in range(1, len(enc) + 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.fill, cell.font = AZUL, Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[hdr_row].height = 32

    tot_p = tot_e = Decimal(0)
    for f in filas:
        m = movs[f["sifco"]]
        dp = ((m["auto_padre_despues"] or Decimal(0)) - (m["auto_padre_antes"] or Decimal(0)))
        de = ((m["auto_espejo_despues"] or Decimal(0)) - (m["auto_espejo_antes"] or Decimal(0)))
        tot_p += dp
        tot_e += de
        ws.append([
            f["sifco"], f["cliente"], ACCION_CORTA.get(f["accion"], "Sin definir"),
            float(m["cap_padre_antes"]),
            float(f["objetivo"]) if f["objetivo"] is not None else "",
            float(dp) if dp != 0 else "—",
            float(de) if de != 0 else "—",
            nota_corta(m, f),
        ])

    fin = ws.max_row
    ws.append(["", "TOTAL", "", "", "", float(tot_p), float(tot_e), ""])
    for c in range(1, len(enc) + 1):
        cell = ws.cell(row=fin + 1, column=c)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    for r in range(hdr_row + 1, ws.max_row + 1):
        for c in (4, 5, 6, 7):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        nota = ws.cell(row=r, column=8)
        if nota.value:
            nota.fill = AMBAR
            nota.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEFGH", (17, 38, 15, 14, 15, 15, 16, 46)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.sheet_view.showGridLines = False
    return wb


def hoja_excluidos(wb, filas_fuera, movs):
    ws = wb.create_sheet("Fuera del lote")
    ws.append(["CRÉDITOS QUE CONTA MARCÓ PERO QUE NO NECESITAN AJUSTE DE AUTOCASH"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t = ws.cell(row=1, column=1)
    t.fill, t.font = AZUL, Font(bold=True, color="FFFFFF", size=12)
    ws.row_dimensions[1].height = 26
    ws.append([])
    enc = ["Crédito", "Cliente", "Capital correcto", "Por qué no entra"]
    ws.append(enc)
    for c in range(1, 5):
        cell = ws.cell(row=3, column=c)
        cell.fill, cell.font = AZUL, Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for f in filas_fuera:
        ws.append([f["sifco"], f["cliente"],
                   float(f["objetivo"]) if f["objetivo"] is not None else "",
                   EXCLUIDOS.get(f["sifco"], "")])
    for r in range(4, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="center")
    for col, w in zip("ABCD", (17, 38, 16, 78)):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False


def item_plan(sifco, cred, invs, m, lote, accion, objetivo, opcional=False, notas=None):
    """Arma un item del plan JSON: qué queda en el padre y qué en el espejo."""
    auto = next((x for x in invs if x["nombre"] == AUTOCASH), None)
    toca_padre = m["cap_padre_despues"] != m["cap_padre_antes"] or m["auto_padre_despues"] != m["auto_padre_antes"]
    toca_espejo = m["auto_espejo_despues"] != m["auto_espejo_antes"]
    return {
        "numero_credito_sifco": sifco,
        "cliente": cred["cliente"] if "cliente" in cred else "",
        "status": cred["status"],
        "lote": lote,
        "opcional": opcional,
        "accion": accion,
        "capital_objetivo": float(objetivo) if objetivo is not None else None,
        "padre": None if not toca_padre else {
            "capital_antes": float(m["cap_padre_antes"]), "capital_despues": float(m["cap_padre_despues"]),
            "autocash_antes": float(m["auto_padre_antes"]), "autocash_despues": float(m["auto_padre_despues"]),
        },
        "espejo": None if not toca_espejo else {
            "autocash_antes": float(m["auto_espejo_antes"] or 0),
            "autocash_despues": float(m["auto_espejo_despues"] or 0),
        },
        "otros_movidos": [{
            "nombre": o["nombre"],
            "padre_antes": float(o["padre_antes"]), "padre_despues": float(o["padre_despues"]),
            "espejo_antes": float(o["espejo_antes"] or 0), "espejo_despues": float(o["espejo_despues"] or 0),
        } for o in m["otros_movidos"]],
        # Los pagos solo se recalculan si cambia el capital del crédito: es de ahí
        # que recalcularPagosCredito amortiza. Si solo se movió el espejo, no aplica.
        "recalcular_pagos": m["cap_padre_despues"] != m["cap_padre_antes"],
        "notas": (notas or []) + m["avisos"],
    }


def calcular_saneamiento(cred, invs):
    """Sin tocar el capital del crédito: Autocash cuadra el padre y su espejo se alinea."""
    auto = next((x for x in invs if x["nombre"] == AUTOCASH), None)
    suma_padre = sum(x["padre"] for x in invs)
    cap = cred["capital"]
    auto_padre_new = q2(auto["padre"] + (cap - suma_padre))
    mov = {
        "status": cred["status"], "n_inv": len(invs),
        "otros": [x["nombre"] for x in invs if x["nombre"] != AUTOCASH],
        "cap_padre_antes": cap, "cap_padre_despues": cap,        # el capital NO se toca
        "suma_padre_antes": q2(suma_padre), "suma_padre_despues": cap,
        "auto_padre_antes": auto["padre"], "auto_padre_despues": auto_padre_new,
        "suma_espejo_antes": q2(sum((x["espejo"] or Decimal(0)) for x in invs)),
        "suma_espejo_despues": None,
        "auto_espejo_antes": auto["espejo"], "auto_espejo_despues": auto_padre_new,
        "otros_movidos": [], "avisos": [],
    }
    ajenos = [x["nombre"] for x in invs
              if x["nombre"] != AUTOCASH and x["espejo"] is not None
              and abs(x["padre"] - x["espejo"]) > Decimal("0.01")]
    if ajenos:
        mov["avisos"].append("El espejo de " + ", ".join(ajenos) + " queda descuadrado a propósito: no es de Autocash.")
    return mov


def main():
    ap = argparse.ArgumentParser(description="Preview del ajuste autorizado por conta (no escribe nada)")
    ap.add_argument("--excel", help="Excel revisado por conta")
    ap.add_argument("--db-url", default=os.environ.get("SUPABASE_DB_URL"))
    ap.add_argument("--salida", default=os.path.expanduser("~/Descargas/Preview_Ajuste_Autocash_conta.xlsx"))
    ap.add_argument("--salida-corta", default=os.path.expanduser("~/Descargas/Resumen_Ajuste_Autocash_gerencia.xlsx"))
    ap.add_argument("--plan", default="plan_ajuste_conta.json", help="JSON que consume el aplicador")
    ap.add_argument("--inversionista", default=AUTOCASH)
    args = ap.parse_args()

    ruta = os.path.expanduser(args.excel) if args.excel else None
    if not ruta:
        cand = [f for f in glob.glob(os.path.expanduser("~/Descargas/*.xlsx"))
                if "Comparativo_Autocash" in f and "(1)" in f]
        if not cand:
            sys.exit("No encontré el Excel de conta; pasalo con --excel")
        ruta = cand[0]
    if not args.db_url:
        sys.exit("Falta SUPABASE_DB_URL o --db-url")

    print(f"Excel de conta: {ruta}")
    filas = leer_conta(ruta)
    print(f"Créditos marcados por conta: {len(filas)}\n")

    creditos, invs = cargar_db(args.db_url, [f["sifco"] for f in filas] + list(SANEAMIENTO))
    movs = {f["sifco"]: calcular(f, creditos[f["sifco"]], invs[f["sifco"]]) for f in filas
            if f["sifco"] in creditos}

    faltan = [f["sifco"] for f in filas if f["sifco"] not in creditos]
    if faltan:
        print(f"⚠ No están en la base: {faltan}\n")

    tot_p = tot_e = Decimal(0)
    for f in filas:
        m = movs.get(f["sifco"])
        if not m:
            continue
        dp = (m["auto_padre_despues"] - m["auto_padre_antes"]) if m["auto_padre_antes"] is not None else Decimal(0)
        de = ((m["auto_espejo_despues"] or Decimal(0)) - (m["auto_espejo_antes"] or Decimal(0)))
        tot_p += dp
        tot_e += de
        print(f"{f['sifco']}  {f['cliente'][:34]:36}[{m['status']}]  {f['leyenda']}")
        print(f"    capital correcto: Q{f['objetivo']:,}  ({f['origen_monto']})")
        print(f"    PADRE  capital {m['cap_padre_antes']:>12,} → {m['cap_padre_despues']:>12,}   "
              f"Autocash {m['auto_padre_antes']:>12,} → {m['auto_padre_despues']:>12,}  ({dp:+,.2f})")
        print(f"    ESPEJO suma    {m['suma_espejo_antes']:>12,} → {m['suma_espejo_despues']:>12,}   "
              f"Autocash {(m['auto_espejo_antes'] or 0):>12,} → {(m['auto_espejo_despues'] or 0):>12,}  ({de:+,.2f})")
        if m["otros"]:
            print(f"    otros inversionistas: {', '.join(m['otros'])}")
        for om in m["otros_movidos"]:
            print(f"      ⚠ TAMBIÉN se mueve {om['nombre']}: {om['padre_antes']:,} → {om['padre_despues']:,}")
        for a in m["avisos"]:
            print(f"    ⚠ {a}")
        if negativos(m):
            print("    ⛔ EL RESULTADO DEJA A AUTOCASH EN NEGATIVO")
        print()

    print("──────── totales ────────")
    print(f"  créditos en el lote: {len(movs)}")
    print(f"  movimiento neto Autocash en el PADRE:  {tot_p:+,.2f}")
    print(f"  movimiento neto Autocash en el ESPEJO: {tot_e:+,.2f}")

    utiles = [f for f in filas if f["sifco"] in movs]
    dentro = [f for f in utiles if f["sifco"] not in EXCLUIDOS]
    fuera = [f for f in utiles if f["sifco"] in EXCLUIDOS]

    escribir_excel(args.salida, utiles, movs, creditos)
    wb = escribir_excel_corto(args.salida_corta, dentro, movs)
    if fuera:
        hoja_excluidos(wb, fuera, movs)
    wb.save(args.salida_corta)
    print(f"\nEn el lote de conta: {len(dentro)}   ·   fuera del lote: {len(fuera)}")

    # ── Plan JSON: lote de conta + lote de saneamiento ────────────────────────
    items = []
    for f in dentro:
        c, iv, m = creditos[f["sifco"]], invs[f["sifco"]], movs[f["sifco"]]
        c = {**c, "cliente": f["cliente"]}
        items.append(item_plan(f["sifco"], c, iv, m, "conta", f["accion"], f["objetivo"],
                               notas=[f["leyenda"]]))

    print("\n── Lote de saneamiento interno ──")
    for sifco, cfg in SANEAMIENTO.items():
        if sifco not in creditos:
            print(f"  ⚠ {sifco} no está en la base, se omite")
            continue
        c, iv = creditos[sifco], invs[sifco]
        m = calcular_saneamiento(c, iv)
        dp = m["auto_padre_despues"] - m["auto_padre_antes"]
        de = (m["auto_espejo_despues"] or Decimal(0)) - (m["auto_espejo_antes"] or Decimal(0))
        etiqueta = " [OPCIONAL]" if cfg["opcional"] else ""
        print(f"  {sifco}{etiqueta}  [{c['status']}]  capital {c['capital']:,} (no se toca)")
        print(f"      PADRE  Autocash {m['auto_padre_antes']:>12,} → {m['auto_padre_despues']:>12,}  ({dp:+,.2f})")
        print(f"      ESPEJO Autocash {(m['auto_espejo_antes'] or 0):>12,} → {(m['auto_espejo_despues'] or 0):>12,}  ({de:+,.2f})")
        print(f"      {cfg['nota']}")
        items.append(item_plan(sifco, c, iv, m, "saneamiento", "saneamiento",
                               c["capital"], opcional=cfg["opcional"], notas=[cfg["nota"]]))

    plan = {
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "excel_conta": ruta,
        "inversionista": args.inversionista if hasattr(args, "inversionista") else AUTOCASH,
        "resumen": {
            "lote_conta": sum(1 for i in items if i["lote"] == "conta"),
            "lote_saneamiento": sum(1 for i in items if i["lote"] == "saneamiento"),
            "opcionales": sum(1 for i in items if i["opcional"]),
            "con_recalculo_de_pagos": sum(1 for i in items if i["recalcular_pagos"]),
        },
        "items": items,
    }
    with open(args.plan, "w", encoding="utf-8") as fh:
        json.dump(plan, fh, ensure_ascii=False, indent=2)
    print(f"\nPlan: {args.plan}")
    for k, v in plan["resumen"].items():
        print(f"  {k}: {v}")
    print(f"\nPreview detallado: {args.salida}")
    print(f"Resumen gerencia:  {args.salida_corta}")
    print("No se modificó nada en la base.")


if __name__ == "__main__":
    main()
