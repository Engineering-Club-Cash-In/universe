#!/usr/bin/env python3
"""Comparativo Blokfund: DB de cartera vs Excel "Cartera Préstamos (Cash-In) NUEVA 3.0".

Mismo método que el comparativo de Autocash, pero para los 12 créditos de Blokfund y
sacando el resultado por consola (no genera Excel).

Cómo se lee el Excel:
  · Un crédito son 1..N filas: SIFCO base + sufijo _2/_3 (una por inversionista). Se SUMAN.
  · A veces el Excel parte el crédito en SIFCO correlativos distintos ("companions"). Se
    detectan agrupando por (nombre, # de cuota): si en el grupo hay EXACTAMENTE UN SIFCO
    que existe en la DB, los demás son rebanadas suyas. Con dos o más, son créditos
    distintos del mismo cliente y no se juntan.
  · La foto se toma de la ÚLTIMA hoja mensual donde el crédito aparece con Pagado = "Si".
    Columnas: "Capital restante" y "Cuota".

    OJO: acá NO vale "Atrasado". Cuando el cliente entra en mora, el Excel deja la fila en
    Atrasado y congela el capital de casi todos los inversionistas, pero al inversionista
    que se le sigue sirviendo la cuota le sigue amortizando ("Pago del mes" fijo mes a mes).
    Tomar esas hojas mete diferencias que no son descuadres: son meses no cobrados al
    cliente. La foto buena es la del último mes efectivamente pagado.
  · Los créditos CRM- no tienen SIFCO en el Excel: se buscan por nombre del cliente.

Uso:
  python3 src/scripts/espejo/comparativo_blokfund.py
"""
import csv
import glob
import importlib.util
import os
import warnings
from collections import defaultdict

import openpyxl

warnings.filterwarnings("ignore")

AQUI = os.path.dirname(os.path.abspath(__file__))
DB_CSV = os.path.join(AQUI, "db_blokfund.csv")
TOL = 1.00        # Q; por debajo se considera redondeo
PAGADO_SI = {"si"}  # estricto: "Atrasado" NO cuenta como pagado

# Se reutilizan los helpers del comparativo de Autocash (norm_nombre, base_sifco, num, ...)
_spec = importlib.util.spec_from_file_location(
    "base", os.path.join(AQUI, "..", "autocash", "comparativo_autocash.py")
)
B = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(B)


def cabecera(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [B.norm_hdr(v) for v in row]
        if "# credito sifco" in vals:
            hdr = {}
            for j, v in enumerate(vals):
                if v and v not in hdr:
                    hdr[v] = j
            return i, hdr
    return None, None


def main():
    db = list(csv.DictReader(open(DB_CSV)))
    en_db = {r["numero_credito_sifco"] for r in db}
    objetivo = [r for r in db if r["es_blokfund"] == "1"]
    print(f"DB local: {len(db)} créditos, {len(objetivo)} con Blokfund S.A.")

    ruta = [f for f in glob.glob(os.path.expanduser("~/Descargas/*.xlsx")) if "NUEVA 3.0" in f][0]
    print(f"Excel   : {ruta}\n")
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = B.hojas_mes(wb)
    print(f"Hojas mensuales: {len(hojas)}  ({hojas[0][1]} → {hojas[-1][1]})")

    # ── Pase 1: companions + índice de nombres ────────────────────────────────
    comp = defaultdict(set)
    nombres = {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_nom, c_num = hdr.get("nombre"), hdr.get("#")
        grupos = defaultdict(set)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if not base:
                continue
            nom = B.norm_nombre(row[c_nom]) if (c_nom is not None and c_nom < len(row)) else ""
            if nom:
                nombres[base] = str(row[c_nom]).strip()
            n = row[c_num] if (c_num is not None and c_num < len(row)) else None
            if nom and n is not None:
                grupos[(nom, str(n))].add(base)
        for bases in grupos.values():
            if len(bases) < 2:
                continue
            dentro = [b for b in bases if b in en_db]
            if len(dentro) == 1:
                comp[dentro[0]] |= (bases - {dentro[0]})

    idx_nombres = defaultdict(list)
    for base, nom in nombres.items():
        idx_nombres[B.norm_nombre(nom)].append(base)

    # ── Match de los 12 ───────────────────────────────────────────────────────
    dueno, llave_de = {}, {}
    for r in objetivo:
        s = r["numero_credito_sifco"]
        if s in nombres:
            base, llave = s, "SIFCO"
        else:
            cands = idx_nombres.get(B.norm_nombre(r["cliente"]), [])
            if len(cands) == 1:
                base, llave = cands[0], "Nombre"
            elif len(cands) > 1:
                base, llave = None, f"Nombre ambiguo ({len(cands)}: {', '.join(sorted(cands))})"
            else:
                base, llave = None, "Sin match"
        llave_de[s] = (base, llave)
        if base:
            for k in {base} | comp.get(base, set()):
                dueno[k] = s

    # ── Pase 2: foto de la última hoja pagada, sumando todas las rebanadas ────
    pagada, vista = {}, {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_cap, c_cuo = hdr.get("capital restante"), hdr.get("cuota")
        c_pag, c_inv = hdr.get("pagado"), hdr.get("inversionista")
        c_pdm, c_num = hdr.get("pago del mes"), hdr.get("#")
        c_fpago = hdr.get("pago")
        agg = defaultdict(lambda: {"cap": None, "cuota": None, "inv": [], "pagado": [],
                                   "sifcos": [], "pdm": 0.0, "ncuota": None, "filas": 0,
                                   "fpago": None})
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if base not in dueno:
                continue
            a = agg[dueno[base]]
            a["filas"] += 1
            a["sifcos"].append(str(row[hdr["# credito sifco"]]).strip())
            for col, key in ((c_cap, "cap"), (c_cuo, "cuota")):
                if col is not None and col < len(row):
                    v = B.num(row[col])
                    if v is not None:
                        a[key] = (a[key] or 0.0) + v
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(B.norm_hdr(row[c_pag]))
            if c_inv is not None and c_inv < len(row) and row[c_inv]:
                nm = str(row[c_inv]).strip()
                if nm not in a["inv"]:
                    a["inv"].append(nm)
            if c_pdm is not None and c_pdm < len(row):
                a["pdm"] += B.num(row[c_pdm]) or 0.0
            if c_num is not None and c_num < len(row) and a["ncuota"] is None:
                a["ncuota"] = row[c_num]
            if c_fpago is not None and c_fpago < len(row) and a["fpago"] is None:
                a["fpago"] = row[c_fpago]
        for s, a in agg.items():
            foto = dict(a, hoja=hoja, tiene_cuota=c_cuo is not None)
            vista[s] = foto
            if any(p in PAGADO_SI for p in a["pagado"]):
                pagada[s] = foto

    # ── Reporte ───────────────────────────────────────────────────────────────
    filas = []
    for r in sorted(objetivo, key=lambda x: x["cliente"]):
        s = r["numero_credito_sifco"]
        base, llave = llave_de[s]
        foto = pagada.get(s)
        crit = "Última hoja con Pagado=Si"
        if foto is None:
            foto = vista.get(s)
            crit = "Última aparición (nunca marcado Pagado=Si)" if foto else "No está en el Excel"
        cap_db, cuo_db = float(r["capital"]), float(r["cuota"])
        cap_x = foto["cap"] if foto else None
        cuo_x = foto["cuota"] if (foto and foto["tiene_cuota"]) else None
        ult_vista = vista.get(s)

        d_cap = (cap_db - cap_x) if cap_x is not None else None
        d_cuo = (cuo_db - cuo_x) if cuo_x is not None else None
        liquidado_xls = cap_x is not None and abs(cap_x) < 0.01

        obs = []
        if liquidado_xls and cap_db > TOL:
            obs.append("Excel lo LIQUIDÓ (cap. restante 0, finiquito en la col. Cuota) y la DB lo tiene con capital vivo")
        else:
            if d_cap is not None and abs(d_cap) > TOL:
                obs.append(f"capital DB ≠ Excel ({d_cap:+,.2f})")
            if d_cuo is not None and abs(d_cuo) > TOL:
                obs.append(f"cuota DB ≠ Excel ({d_cuo:+,.2f})")
        if abs(cap_db - float(r["suma_aportado"] or 0)) > TOL:
            obs.append("capital DB ≠ suma inversionistas")

        filas.append({
            "sifco": s, "cliente": r["cliente"], "status": r["status"],
            "llave": llave, "base": base, "comp": sorted(comp.get(base, set())) if base else [],
            "hoja": foto["hoja"] if foto else "—", "crit": crit,
            "filas_xls": foto["filas"] if foto else 0,
            "sifcos": foto["sifcos"] if foto else [],
            "inv_xls": foto["inv"] if foto else [],
            "ncuota": foto["ncuota"] if foto else None,
            "fpago": foto["fpago"] if foto else None,
            "hoja_ult": ult_vista["hoja"] if ult_vista else "—",
            "ncuota_ult": ult_vista["ncuota"] if ult_vista else None,
            "pagado_ult": ", ".join(sorted(set(ult_vista["pagado"]))) if ult_vista else "",
            "cap_db": cap_db, "cap_x": cap_x, "cuo_db": cuo_db, "cuo_x": cuo_x,
            "d_cap": d_cap, "d_cuo": d_cuo,
            "cap_inv": float(r["suma_aportado"] or 0),
            "cuo_inv": float(r["suma_cuota_inv"] or 0),
            "inv_db": r["inversionistas"],
            "db_ult_cuota": r["ult_cuota_pagada"], "db_ult_vence": r["ult_cuota_vence"],
            "db_parcial": r["cuota_parcial"], "db_parcial_monto": r["parcial_boleta"],
            "obs": obs,
        })

    def f(v):
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10] if v is not None else "—"

    def m(v):
        return f"{v:>12,.2f}" if isinstance(v, float) else f"{'sin dato':>12}"

    def m2(v):
        """Como m(), pero para la columna Cuota, que no existe antes de Julio 2024."""
        return f"{v:>12,.2f}" if isinstance(v, float) else f"{'sin col.':>12}"

    ok = [x for x in filas if not x["obs"]]
    malos = [x for x in filas if x["obs"]]

    print("\n" + "=" * 132)
    print(f"✅ CUADRAN EXACTO  ({len(ok)} de {len(filas)})")
    print("=" * 132)
    print(f"{'CLIENTE':<34}{'SIFCO':<17}{'STATUS':<11}{'CAP DB':>13}{'CAP XLS':>13}"
          f"{'CUOTA DB':>12}{'CUOTA XLS':>12}   HOJA (última pagada)")
    print("-" * 132)
    for x in ok:
        print(f"{x['cliente'][:33]:<34}{x['sifco'][:16]:<17}{x['status']:<11}"
              f"{m(x['cap_db'])}{m(x['cap_x'])}{x['cuo_db']:>12,.2f}{m2(x['cuo_x'])}   "
              f"{x['hoja']} (#{x['ncuota']})")

    print("\n" + "=" * 132)
    print(f"⚠️  A OBSERVACIÓN  ({len(malos)} de {len(filas)})")
    print("=" * 132)
    for x in malos:
        print(f"\n▸ {x['cliente']}  [{x['status']}]   {x['sifco']}")
        for o in x["obs"]:
            print(f"    → {o}")
        print(f"    Excel  : hoja {x['hoja']} (cuota #{x['ncuota']}, pago {f(x['fpago'])})  [{x['crit']}]")
        print(f"             capital {m(x['cap_x'])}   cuota {m(x['cuo_x'])}")
        print(f"             filas {x['filas_xls']} → {', '.join(x['sifcos'])}")
        print(f"             inversionistas: {' | '.join(x['inv_xls']) or '—'}")
        print(f"    DB     : capital {m(x['cap_db'])}   cuota {m(x['cuo_db'])}   "
              f"suma inv {m(x['cap_inv'])}")
        print(f"             última cuota pagada #{x['db_ult_cuota']} (vence {f(x['db_ult_vence'])})"
              + (f", parcial en #{x['db_parcial']} por {float(x['db_parcial_monto']):,.2f}"
                 if x["db_parcial"] else ""))
        print(f"             inversionistas: {x['inv_db']}")
        print(f"    Estado hoy en Excel: hoja {x['hoja_ult']} (cuota #{x['ncuota_ult']}) → {x['pagado_ult']}")

    print("\n" + "=" * 132)
    print(f"RESUMEN  ({len(filas)} créditos de Blokfund, tolerancia Q{TOL:.2f})")
    print("=" * 132)
    print(f"  Cuadran exacto capital y cuota : {len(ok)}")
    print(f"  A observación                  : {len(malos)}")


if __name__ == "__main__":
    main()
