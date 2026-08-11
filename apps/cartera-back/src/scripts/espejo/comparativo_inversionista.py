#!/usr/bin/env python3
"""Comparativo de la cartera de UN inversionista: DB de cartera vs el Excel de cartera.

Es la versión parametrizable del comparativo que se hizo para Autocash. Compara dos
cosas por crédito:

  1. CAPITAL del crédito   — creditos.capital        vs suma de "Capital restante"
  2. LO DEL INVERSIONISTA  — creditos_inversionistas vs la(s) fila(s) del Excel cuyo
                             "Inversionista" es el que se está revisando

Criterio de la foto (qué mes se toma como bueno): la ÚLTIMA hoja mensual donde el
crédito aparece **efectivamente cobrado**, o sea con `Pagado = "Si"` **o** donde el
"Monto boleta" iguala a la "Cuota" (hay meses cobrados que quedaron sin marcar).

NO vale "Atrasado": cuando el cliente cae en mora el Excel congela el capital de casi
todos los inversionistas, pero al que se le sigue sirviendo la cuota le sigue
amortizando. Tomar esas hojas mete diferencias que no son descuadres.

Un crédito son 1..N filas del Excel (SIFCO base + sufijo _2/_3, una por inversionista);
se suman. Y a veces el Excel lo parte en SIFCO correlativos distintos ("companions"),
que se detectan agrupando por (nombre, # de cuota).

Los CSV de la DB viven FUERA del repo (~/Descargas/datos-cuadre-autocash/): traen
nombres de clientes, montos y participaciones. Se regeneran con el COPY del README.

Uso:
  python3 src/scripts/espejo/comparativo_inversionista.py \
      --csv ~/Descargas/datos-cuadre-autocash/db_inv_97.csv \
      --nombre "Blokfund S.A." --excel ~/Descargas/Cartera.xlsx

  # a nivel del inversionista (ignora el capital del crédito), omitiendo los que
  # no están en Cartera.xlsx:
  python3 src/scripts/espejo/comparativo_inversionista.py \
      --csv ~/Descargas/datos-cuadre-autocash/db_prod_inv89.csv \
      --nombre "Autocash S.A." --foco-inv \
      --omitir "Centralcarga,01010407301720,Josecollia,Pistolon"
"""
import argparse
import csv
import glob
import importlib.util
import os
import unicodedata
import warnings
from collections import defaultdict

import openpyxl

warnings.filterwarnings("ignore")

AQUI = os.path.dirname(os.path.abspath(__file__))
TOL = 1.00
PAGADO_SI = {"si"}

_spec = importlib.util.spec_from_file_location(
    "base", os.path.join(AQUI, "..", "autocash", "comparativo_autocash.py")
)
B = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(B)


def resolver_excel(patron):
    """Busca por glob: los nombres con tildes vienen en NFD y no matchean literal."""
    if os.path.isfile(patron):
        return patron
    objetivo = unicodedata.normalize("NFC", os.path.basename(patron)).lower()
    for f in glob.glob(os.path.join(os.path.dirname(os.path.expanduser(patron)) or ".", "*.xlsx")):
        if unicodedata.normalize("NFC", os.path.basename(f)).lower() == objetivo:
            return f
    raise SystemExit(f"No encontré el Excel: {patron}")


def cabecera(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [B.norm_hdr(v) for v in row]
        if "# credito sifco" in vals:
            hdr = {}
            for j, v in enumerate(vals):
                if v and v not in hdr:
                    hdr[v] = j
            return i, hdr
    return None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True, help="CSV de la DB (columna es_objetivo)")
    ap.add_argument("--nombre", required=True, help='Nombre del inversionista, p.ej. "Blokfund S.A."')
    ap.add_argument("--excel", default="~/Descargas/Cartera.xlsx")
    ap.add_argument("--omitir", default="",
                    help="SIFCOs a excluir del comparativo, separados por coma")
    ap.add_argument("--foco-inv", action="store_true",
                    help="Reporte enfocado SOLO en el monto del inversionista (ignora capital del crédito)")
    args = ap.parse_args()

    OMITIR = {s.strip() for s in args.omitir.split(",") if s.strip()}

    INV = B.norm_nombre(args.nombre)
    es_inv = lambda n: INV in B.norm_nombre(n)

    ruta_csv = os.path.expanduser(args.csv)
    if not os.path.isabs(ruta_csv):
        ruta_csv = os.path.join(AQUI, ruta_csv)
    db = list(csv.DictReader(open(ruta_csv)))
    en_db = {r["numero_credito_sifco"] for r in db}
    objetivo = [r for r in db if r["es_objetivo"] == "1"]
    omitidos = [r for r in objetivo if r["numero_credito_sifco"] in OMITIR]
    objetivo = [r for r in objetivo if r["numero_credito_sifco"] not in OMITIR]
    print(f"Inversionista : {args.nombre}")
    print(f"DB local      : {len(db)} créditos, {len(objetivo)} suyos"
          + (f"  (omitidos {len(omitidos)})" if omitidos else ""))
    for r in omitidos:
        print(f"                omitido → {r['numero_credito_sifco']}  {r['cliente'][:40]}"
              f"  (suyo {float(r['inv_padre'] or 0):,.2f})")

    ruta = resolver_excel(os.path.expanduser(args.excel))
    print(f"Excel         : {ruta}")
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = B.hojas_mes(wb)
    print(f"Hojas         : {len(hojas)}  ({hojas[0][1]} → {hojas[-1][1]})\n")

    # ── Pase 1: companions + índice de nombres ────────────────────────────────
    comp, nombres = defaultdict(set), {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_nom, c_num = hdr.get("nombre"), hdr.get("#")
        grupos = defaultdict(set)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if not base:
                continue
            nom = B.norm_nombre(row[c_nom]) if (c_nom is not None and c_nom < len(row)) else ""
            if nom:
                nombres[base] = str(row[c_nom]).strip()
            n = row[c_num] if (c_num is not None and c_num < len(row)) else None
            if nom and n is not None:
                grupos[(nom, str(n))].add(base)
        for bases in grupos.values():
            if len(bases) < 2:
                continue
            dentro = [b for b in bases if b in en_db]
            if len(dentro) == 1:
                comp[dentro[0]] |= (bases - {dentro[0]})

    idx = defaultdict(list)
    for base, nom in nombres.items():
        idx[B.norm_nombre(nom)].append(base)

    # ── Match ─────────────────────────────────────────────────────────────────
    dueno, llave_de = {}, {}
    for r in objetivo:
        s = r["numero_credito_sifco"]
        if s in nombres:
            base, llave = s, "SIFCO"
        else:
            cands = idx.get(B.norm_nombre(r["cliente"]), [])
            base, llave = (cands[0], "Nombre") if len(cands) == 1 else (
                (None, f"Nombre ambiguo ({len(cands)})") if cands else (None, "Sin match"))
        llave_de[s] = (base, llave)
        if base:
            for k in {base} | comp.get(base, set()):
                dueno[k] = s

    # ── Pase 2: foto del último mes cobrado ───────────────────────────────────
    def nueva():
        return {"cap": 0.0, "cuota": 0.0, "boleta": 0.0, "pagado": [], "filas": 0,
                "sifcos": [], "inv": [], "mio_cap": None, "mio_cuota": 0.0,
                "otros": defaultdict(float), "ncuota": None}

    pagada, vista = {}, {}
    for _, hoja, _a in hojas:
        ws = wb[hoja]
        hr, hdr = cabecera(ws)
        if hr is None:
            continue
        c_cap, c_cuo = hdr.get("capital restante"), hdr.get("cuota")
        c_pag, c_inv = hdr.get("pagado"), hdr.get("inversionista")
        c_bol, c_num = hdr.get("monto boleta"), hdr.get("#")
        agg = defaultdict(nueva)
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            base = B.base_sifco(row[hdr["# credito sifco"]])
            if base not in dueno:
                continue
            a = agg[dueno[base]]
            a["filas"] += 1
            a["sifcos"].append(str(row[hdr["# credito sifco"]]).strip())
            g = lambda c: (B.num(row[c]) if (c is not None and c < len(row)) else None)
            cap, cuo, bol = g(c_cap), g(c_cuo), g(c_bol)
            if cap is not None:
                a["cap"] += cap
            if cuo is not None:
                a["cuota"] += cuo
            if bol is not None:
                a["boleta"] += bol
            if c_pag is not None and c_pag < len(row) and row[c_pag] is not None:
                a["pagado"].append(B.norm_hdr(row[c_pag]))
            if c_num is not None and c_num < len(row) and a["ncuota"] is None:
                a["ncuota"] = row[c_num]
            nm = str(row[c_inv]).strip() if (c_inv is not None and c_inv < len(row) and row[c_inv]) else ""
            if nm:
                if nm not in a["inv"]:
                    a["inv"].append(nm)
                if es_inv(nm):
                    a["mio_cap"] = (a["mio_cap"] or 0.0) + (cap or 0.0)
                    a["mio_cuota"] += cuo or 0.0
                else:
                    a["otros"][nm] += cap or 0.0
        for s, a in agg.items():
            foto = dict(a, hoja=hoja, tiene_cuota=c_cuo is not None, otros=dict(a["otros"]))
            vista[s] = foto
            cobrado = any(p in PAGADO_SI for p in a["pagado"]) or (
                a["cuota"] > 0 and abs(a["boleta"] - a["cuota"]) < 0.01)
            if cobrado:
                pagada[s] = foto

    # ── Armado + clasificación en lotes ───────────────────────────────────────
    filas, sin_match = [], []
    for r in sorted(objetivo, key=lambda x: x["cliente"]):
        s = r["numero_credito_sifco"]
        base, llave = llave_de[s]
        foto, crit = pagada.get(s), "mes cobrado"
        if foto is None:
            foto = vista.get(s)
            crit = "última aparición (nunca cobrado)" if foto else None
        if foto is None:
            sin_match.append({"sifco": s, "cliente": r["cliente"], "status": r["status"],
                              "llave": llave, "capital": float(r["capital"] or 0),
                              "mio": float(r["inv_padre"] or 0),
                              "inversionistas": r["inversionistas"]})
            continue

        cap_db, mio_db = float(r["capital"] or 0), float(r["inv_padre"] or 0)
        cap_x, mio_x = foto["cap"], foto["mio_cap"]
        d_cap = cap_db - cap_x
        d_mio = (mio_db - mio_x) if mio_x is not None else None
        liquidado = abs(cap_x) < 0.01

        obs = []
        if liquidado and cap_db > TOL:
            obs.append("Excel lo LIQUIDÓ (capital restante 0) y la DB lo tiene vivo")
        else:
            if abs(d_cap) > TOL:
                obs.append(f"capital del crédito ({d_cap:+,.2f})")
            if mio_x is None:
                obs.append(f"{args.nombre} ya no figura en el Excel")
            elif abs(d_mio) > TOL:
                obs.append(f"monto de {args.nombre} ({d_mio:+,.2f})")

        x = {"sifco": s, "cliente": r["cliente"], "status": r["status"], "llave": llave,
             "comp": sorted(comp.get(base, set())) if base else [],
             "hoja": foto["hoja"], "crit": crit, "ncuota": foto["ncuota"],
             "n_inv": int(r["n_inv"] or 0), "inv_db": r["inversionistas"],
             "inv_x": foto["inv"], "otros_x": foto["otros"],
             "cap_db": cap_db, "cap_x": cap_x, "d_cap": d_cap,
             "mio_db": mio_db, "mio_x": mio_x, "d_mio": d_mio,
             "mio_espejo": float(r["inv_espejo"]) if r["inv_espejo"] else None,
             "cuo_db": float(r["cuota"] or 0),
             "cuo_x": foto["cuota"] if foto["tiene_cuota"] else None,
             "liquidado": liquidado, "obs": obs}
        d_m = d_mio or 0.0
        x["lote"] = (1 if liquidado and cap_db > TOL else
                     2 if cap_db <= TOL and cap_x > TOL else
                     3 if abs(d_cap) > 100 else
                     4 if abs(d_cap) <= TOL and abs(d_m) > TOL else
                     5 if obs else 0)
        filas.append(x)

    ok = [x for x in filas if not x["obs"]]
    mal = [x for x in filas if x["obs"]]
    M = lambda v, w=13: (f"{v:>{w},.2f}" if isinstance(v, float) else f"{'—':>{w}}")
    L = lambda n: sorted([x for x in filas if x["lote"] == n], key=lambda y: -abs(y["d_cap"]))

    print("=" * 140)
    print(f"✅ CUADRAN — capital del crédito y monto de {args.nombre}  ({len(ok)} de {len(filas)})")
    print("=" * 140)
    print(f"{'CLIENTE':<36}{'SIFCO':<17}{'STATUS':<12}{'CAP DB':>13}{'CAP XLS':>13}"
          f"{'SUYO DB':>14}{'SUYO XLS':>13}{'#INV':>5}  MES COBRADO")
    print("-" * 140)
    for x in ok:
        print(f"{x['cliente'][:35]:<36}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
              f"{M(x['cap_db'])}{M(x['cap_x'])}{M(x['mio_db'],14)}{M(x['mio_x'])}"
              f"{x['n_inv']:>5}  {x['hoja']} (#{x['ncuota']})")

    LOTES = {1: "EL EXCEL LOS LIQUIDÓ Y EL SISTEMA LOS TIENE VIVOS",
             2: "EL SISTEMA LOS TIENE EN CERO Y EL EXCEL CON SALDO",
             3: "DIFERENCIAS DE CAPITAL DEL CRÉDITO",
             4: "EL CAPITAL CUADRA, CAMBIA EL REPARTO ENTRE INVERSIONISTAS",
             5: "CENTAVOS"}
    print("\n" + "=" * 140)
    print(f"⚠️  NO CUADRAN  ({len(mal)} de {len(filas)})")
    print("=" * 140)
    for n, titulo in LOTES.items():
        grupo = L(n)
        if not grupo:
            continue
        print(f"\n───── LOTE {n} · {titulo}  ({len(grupo)}) " + "─" * 40)
        for x in grupo:
            print(f"\n▸ {x['cliente']}  [{x['status']}]   {x['sifco']}")
            for o in x["obs"]:
                print(f"    → {o}")
            print(f"    Mes usado : {x['hoja']} (cuota #{x['ncuota']}, {x['crit']})"
                  + (f"   companions: {', '.join(x['comp'])}" if x["comp"] else ""))
            print(f"    CAPITAL   : DB {M(x['cap_db'])}   Excel {M(x['cap_x'])}   dif {M(x['d_cap'])}")
            print(f"    SUYO      : DB {M(x['mio_db'])}   Excel {M(x['mio_x'])}   dif {M(x['d_mio'])}"
                  + (f"   espejo {M(x['mio_espejo'])}" if x["mio_espejo"] is not None else ""))
            if x["n_inv"] > 1:
                print(f"    Otros DB  : {x['inv_db']}")
                print(f"    Otros XLS : " + (" | ".join(f"{k}={v:,.2f}" for k, v in x["otros_x"].items()) or "—"))

    if sin_match:
        print("\n" + "=" * 140)
        print(f"❓ NO APARECEN EN EL EXCEL  ({len(sin_match)})")
        print("=" * 140)
        print(f"{'CLIENTE':<40}{'SIFCO':<17}{'STATUS':<12}{'CAPITAL DB':>14}{'SUYO DB':>14}  MOTIVO")
        print("-" * 140)
        for x in sin_match:
            print(f"{x['cliente'][:39]:<40}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
                  f"{M(x['capital'],14)}{M(x['mio'],14)}  {x['llave']}")

    print("\n" + "=" * 140)
    print(f"RESUMEN  ({len(objetivo)} créditos de {args.nombre}, tolerancia Q{TOL:.2f})")
    print("=" * 140)
    print(f"  Encontrados en el Excel : {len(filas)}")
    print(f"  Cuadran capital Y suyo  : {len(ok)}")
    print(f"  No cuadran              : {len(mal)}")
    for n, titulo in LOTES.items():
        if L(n):
            print(f"      lote {n} · {titulo:<58} {len(L(n))}")
    print(f"  No aparecen en el Excel : {len(sin_match)}")
    tot_db = sum(x["mio_db"] for x in filas) + sum(x["mio"] for x in sin_match)
    tot_x = sum(x["mio_x"] for x in filas if x["mio_x"] is not None)
    print(f"\n  TOTAL suyo en la DB     : {tot_db:>16,.2f}")
    print(f"  TOTAL suyo en el Excel  : {tot_x:>16,.2f}")

    if args.foco_inv:
        foco(args, filas, sin_match)


def foco(args, filas, sin_match):
    """Reporte a nivel del inversionista: solo monto_aportado DB vs Excel.

    El capital del crédito no entra. Lo que se persigue acá es el total del
    inversionista, que es lo que conta compara contra su libro.
    """
    CENT = 0.01
    M = lambda v, w=14: (f"{v:>{w},.2f}" if isinstance(v, float) else f"{'—':>{w}}")

    con_dato = [x for x in filas if x["mio_x"] is not None]
    fuera = [x for x in filas if x["mio_x"] is None]          # está el crédito, no él
    dif = sorted([x for x in con_dato if abs(x["d_mio"]) >= CENT],
                 key=lambda y: -abs(y["d_mio"]))
    igual = [x for x in con_dato if abs(x["d_mio"]) < CENT]

    print("\n\n" + "=" * 132)
    print(f"🎯 FOCO — SOLO EL MONTO DE {args.nombre.upper()} (sin mirar capital del crédito)")
    print("=" * 132)

    print(f"\n▸ CUADRAN AL CENTAVO  ({len(igual)})")

    if dif:
        materiales = [x for x in dif if abs(x["d_mio"]) >= 1.00]
        centavos = [x for x in dif if abs(x["d_mio"]) < 1.00]

        print(f"\n▸ DIFERENCIAS MATERIALES (≥ Q1)  ({len(materiales)})")
        print("-" * 132)
        print(f"{'CLIENTE':<34}{'SIFCO':<17}{'STATUS':<12}{'CAPITAL DB':>14}{'CAPITAL XLS':>14}"
              f"{'SUYO DB':>14}{'SUYO XLS':>13}{'DIF':>12}")
        print("-" * 132)
        for x in materiales:
            print(f"{x['cliente'][:33]:<34}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
                  f"{M(x['cap_db'])}{M(x['cap_x'])}{M(x['mio_db'])}{M(x['mio_x'],13)}{M(x['d_mio'],12)}")
        print("-" * 132)
        print(f"{'SUMA MATERIALES':<77}{'':>13}{M(sum(x['d_mio'] for x in materiales),12)}")
        if centavos:
            print(f"{'+ ' + str(len(centavos)) + ' créditos de centavos':<77}{'':>13}"
                  f"{M(sum(x['d_mio'] for x in centavos),12)}")
            print(f"{'TOTAL':<77}{'':>13}{M(sum(x['d_mio'] for x in dif),12)}")

        print(f"\n▸ QUIÉNES ESTÁN EN CADA UNO DE ESOS {len(materiales)} CRÉDITOS")
        print("-" * 132)
        for x in materiales:
            print(f"\n  {x['cliente']}   {x['sifco']}   [{x['status']}]"
                  f"   capital {x['cap_db']:,.2f}   {x['n_inv']} inversionista(s)"
                  f"   mes {x['hoja']} (cuota #{x['ncuota']})")
            print(f"      DB    : {x['inv_db']}")
            print(f"      Excel : " + (" | ".join(f"{k}={v:,.2f}" for k, v in x["otros_x"].items())
                                       or "(solo él)")
                  + f"  +  {args.nombre}={x['mio_x']:,.2f}")

    if fuera:
        print(f"\n▸ EL CRÉDITO ESTÁ EN EL EXCEL PERO {args.nombre.upper()} NO FIGURA COMO INVERSIONISTA  ({len(fuera)})")
        print("-" * 132)
        for x in fuera:
            print(f"{x['cliente'][:33]:<34}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
                  f"{M(x['mio_db'])}{'':>14}{M(-x['mio_db'])}   quién lo tiene en XLS: "
                  + (" | ".join(f"{k}={v:,.2f}" for k, v in x["otros_x"].items()) or "—"))
        print("-" * 132)
        print(f"{'SUMA':<77}{M(-sum(x['mio_db'] for x in fuera))}")

    if sin_match:
        print(f"\n▸ NO APARECEN EN EL EXCEL  ({len(sin_match)})")
        print("-" * 132)
        for x in sin_match:
            print(f"{x['cliente'][:33]:<34}{x['sifco'][:16]:<17}{x['status'][:11]:<12}"
                  f"{M(x['mio'])}{'':>14}{M(-x['mio'])}  {x['llave']}")
        print("-" * 132)
        print(f"{'SUMA':<77}{M(-sum(x['mio'] for x in sin_match))}")

    d_dif = sum(x["d_mio"] for x in dif)
    d_fuera = -sum(x["mio_db"] for x in fuera)
    d_sin = -sum(x["mio"] for x in sin_match)

    print("\n" + "=" * 132)
    print("CIERRE DEL DESCUADRE  (DB − Excel)")
    print("=" * 132)
    print(f"  a) créditos con diferencia de monto      {M(d_dif)}   ({len(dif)} créditos)")
    print(f"  b) el crédito está, pero él ya no        {M(d_fuera)}   ({len(fuera)} créditos)")
    print(f"  c) el crédito no está en el Excel        {M(d_sin)}   ({len(sin_match)} créditos)")
    print("  " + "-" * 74)
    print(f"  TOTAL a+b+c                              {M(d_dif + d_fuera + d_sin)}")
    print(f"\n  Solo (a) — lo que de verdad descuadra    {M(d_dif)}")


if __name__ == "__main__":
    main()
