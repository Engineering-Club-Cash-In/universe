import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import psycopg2
from decimal import Decimal
from datetime import datetime, timezone, timedelta

EXCEL_PATH = r"C:\Users\Kelvin Palacios\Downloads\flujo).xlsx"
SHEET = sys.argv[1] if len(sys.argv) > 1 else "marzo 2026"
INVERSIONISTA_ID = 84

GT = timezone(timedelta(hours=-6))
stamp = datetime.now(GT).strftime("%Y%m%d_%H%M")
OUT_PATH = rf"C:\Users\Kelvin Palacios\Downloads\comparacion_flujocapital_{stamp}.xlsx"

DSN = os.environ.get("SUPABASE_DB_URL")
if not DSN:
    with open(r"c:\Users\Kelvin Palacios\Documents\universe\apps\cartera-back\.env") as f:
        for line in f:
            if line.startswith("SUPABASE_DB_URL="):
                DSN = line.split("=", 1)[1].strip().strip('"').strip("'")
                break

wb_in = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb_in[SHEET]
excel = {}
total_cap_restante = Decimal("0")
for r in ws.iter_rows(min_row=5, values_only=True):
    sifco, capital_restante = r[0], r[12]
    if sifco is None:
        continue
    s = str(sifco).strip()
    sifco_original = s
    if s.endswith("_2"):
        s = s[:-2]
    cr = Decimal(str(capital_restante)) if capital_restante is not None else Decimal("0")
    excel[s] = (cr, sifco_original)
    total_cap_restante += cr

conn = psycopg2.connect(DSN)
cur = conn.cursor()
cur.execute(
    """
    SELECT c.numero_credito_sifco, cie.monto_aportado, c."statusCredit", u.nombre
    FROM cartera.creditos_inversionistas_espejo cie
    JOIN cartera.creditos c ON c.credito_id = cie.credito_id
    LEFT JOIN cartera.usuarios u ON u.usuario_id = c.usuario_id
    WHERE cie.inversionista_id = %s
    """,
    (INVERSIONISTA_ID,),
)
db = {}
total_aportado = Decimal("0")
for sifco, monto, status, cliente in cur.fetchall():
    s = str(sifco).strip()
    monto = Decimal(str(monto))
    db[s] = (monto, status, cliente)
    total_aportado += monto
cur.close()
conn.close()

all_s = sorted(set(excel) | set(db))
rows_ok, rows_diff, rows_only_excel, rows_only_db = [], [], [], []

for s in all_s:
    e = excel.get(s)
    d = db.get(s)
    if e is None:
        monto, status, cliente = d
        rows_only_db.append([s, cliente, status, None, float(monto), None])
    elif d is None:
        cr, sifco_original = e
        rows_only_excel.append([s, sifco_original, float(cr)])
    else:
        cr, sifco_original = e
        monto, status, cliente = d
        diff = monto - cr
        categoria = "OK" if abs(diff) <= Decimal("0.01") else "DIFERENCIA"
        row = [
            s, sifco_original, cliente, status,
            float(cr), float(monto), float(diff), categoria,
        ]
        if categoria == "OK":
            rows_ok.append(row)
        else:
            rows_diff.append(row)

rows_diff.sort(key=lambda r: abs(r[6]), reverse=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
diff_fill = PatternFill("solid", fgColor="F8CBAD")
ok_fill = PatternFill("solid", fgColor="C6EFCE")

def write_sheet(name, headers, rows, highlight_col=None):
    s = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = s.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = s.cell(r, c, v)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
        if highlight_col is not None and len(row) > highlight_col:
            val = row[highlight_col]
            if isinstance(val, (int, float)) and abs(val) > 0.01:
                for c in range(1, len(row) + 1):
                    s.cell(r, c).fill = diff_fill
    for col_cells in s.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        s.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

resumen = wb.create_sheet("Resumen")
resumen["A1"] = "Comparación Flujocapital — capital_restante (Excel) vs monto_aportado (DB)"
resumen["A1"].font = Font(bold=True, size=14)
resumen["A3"] = "Hoja Excel analizada"
resumen["B3"] = SHEET
resumen["A4"] = "Archivo origen"
resumen["B4"] = EXCEL_PATH
resumen["A5"] = "Inversionista ID"
resumen["B5"] = INVERSIONISTA_ID
resumen["A6"] = "Fecha ejecución"
resumen["B6"] = datetime.now(GT).strftime("%Y-%m-%d %H:%M GT")

rows_sum = [
    ("Créditos en Excel", len(excel), float(total_cap_restante)),
    ("Créditos en DB", len(db), float(total_aportado)),
    ("Coinciden (|diff|<=0.01)", len(rows_ok), None),
    ("Con diferencia", len(rows_diff), sum(r[6] for r in rows_diff)),
    ("Solo Excel", len(rows_only_excel), sum(r[2] for r in rows_only_excel)),
    ("Solo DB", len(rows_only_db), sum(r[4] for r in rows_only_db)),
    ("Diff global (DB - Excel)", "", float(total_aportado - total_cap_restante)),
]
resumen["A8"] = "Métrica"; resumen["B8"] = "Cantidad"; resumen["C8"] = "Monto"
for c in ["A8", "B8", "C8"]:
    resumen[c].font = header_font
    resumen[c].fill = header_fill
for i, (k, qty, amt) in enumerate(rows_sum, 9):
    resumen.cell(i, 1, k)
    resumen.cell(i, 2, qty)
    if amt is not None:
        cell = resumen.cell(i, 3, amt)
        cell.number_format = "#,##0.00"
resumen.column_dimensions["A"].width = 32
resumen.column_dimensions["B"].width = 14
resumen.column_dimensions["C"].width = 18

write_sheet(
    "Diferencias",
    ["SIFCO_base", "SIFCO_excel", "Cliente", "Status", "Excel_capital_restante", "DB_monto_aportado", "Diff (DB-Excel)", "Categoria"],
    rows_diff,
    highlight_col=6,
)
write_sheet(
    "Solo en Excel",
    ["SIFCO_base", "SIFCO_excel", "Excel_capital_restante"],
    rows_only_excel,
)
write_sheet(
    "Solo en DB",
    ["SIFCO", "Cliente", "Status", "Excel_capital_restante", "DB_monto_aportado", "_"],
    [[r[0], r[1], r[2], r[3], r[4], r[5]] for r in rows_only_db],
)
write_sheet(
    "Coinciden",
    ["SIFCO_base", "SIFCO_excel", "Cliente", "Status", "Excel_capital_restante", "DB_monto_aportado", "Diff", "Categoria"],
    rows_ok,
)

wb.save(OUT_PATH)
print(f"OK: {OUT_PATH}")
print(f"Excel total: {total_cap_restante:,.2f} | DB total: {total_aportado:,.2f}")
print(f"OK={len(rows_ok)} DIFF={len(rows_diff)} SoloExcel={len(rows_only_excel)} SoloDB={len(rows_only_db)}")
