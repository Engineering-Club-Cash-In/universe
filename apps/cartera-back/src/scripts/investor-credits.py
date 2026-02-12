import openpyxl
import json
import glob
import os

# Buscar el primer archivo Excel en la carpeta actual del script
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_files = glob.glob(os.path.join(script_dir, "*.xlsx")) + glob.glob(os.path.join(script_dir, "*.xls"))

if not excel_files:
    print("No se encontró ningún archivo Excel en la carpeta de scripts.")
    exit(1)

excel_path = excel_files[0]
print(f"Leyendo archivo: {os.path.basename(excel_path)}")

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Tomar la última hoja
last_sheet_name = wb.sheetnames[-1]
print(f"Hoja: {last_sheet_name}")
ws = wb[last_sheet_name]

# Nombre del inversionista en C2
inversionista = ws["C2"].value

# Leer créditos desde fila 5 hasta que no haya datos en columna B (cliente)
creditos = []
row = 5
while True:
    cliente = ws.cell(row=row, column=2).value  # B - CLIENTE
    if not cliente:
        break

    meses_en_credito = ws.cell(row=row, column=1).value   # A - MESES EN CRÉDITO
    capital = ws.cell(row=row, column=3).value             # C - CAPITAL
    inversor = ws.cell(row=row, column=5).value            # E - % INVERSOR
    interes_inversor = ws.cell(row=row, column=7).value    # G - INTERÉS INVERSOR
    iva = ws.cell(row=row, column=8).value                 # H - IVA
    inversionista_neto = ws.cell(row=row, column=11).value # K - % INVERSIONISTA NETO

    creditos.append({
        "meses_en_credito": int(meses_en_credito) if meses_en_credito is not None else None,
        "cliente": str(cliente).strip(),
        "capital": round(float(capital), 2) if capital is not None else None,
        "inversor": float(inversor) if inversor is not None else None,
        "interes_inversor": round(float(interes_inversor), 2) if interes_inversor is not None else None,
        "iva": round(float(iva), 2) if iva is not None else None,
        "inversionista_neto": round(float(inversionista_neto), 2) if inversionista_neto is not None else None,
    })

    row += 1

resultado = {
    "inversionista": inversionista,
    "creditos": creditos
}

# Guardar JSON en results_json/ con el nombre del inversionista
results_dir = os.path.join(script_dir, "results_json")
os.makedirs(results_dir, exist_ok=True)

safe_name = inversionista.strip().replace("/", "-") if inversionista else "sin_nombre"
output_path = os.path.join(results_dir, f"{safe_name}.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(resultado, f, ensure_ascii=False, indent=2)

print(f"\nResultado guardado en: {output_path}")
print(json.dumps(resultado, ensure_ascii=False, indent=2))
